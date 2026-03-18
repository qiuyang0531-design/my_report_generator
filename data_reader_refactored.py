"""
协议驱动型数据提取器 - 重构版本
=====================

核心设计理念：
1. 配置驱动：所有表格识别规则、字段映射、后处理逻辑都由配置定义
2. 解耦识别与解析：通过指纹识别表格类型，通过协议定义提取逻辑
3. 插件化扩展：新增表格类型只需添加配置，无需修改核心代码

架构层次：
- TABLE_PROTOCOLS: 协议配置层（定义所有表格类型的识别规则和处理逻辑）
- TableFingerprint: 表格指纹识别器（负责识别表格类型）
- ProtocolExtractor: 协议提取器（负责根据协议提取数据）
- ExcelDataReader: 高层接口（提供统一的数据获取接口）
"""

import openpyxl
import re
import sys
import os
from datetime import datetime
from typing import Dict, List, Any, Optional, Callable
from dataclasses import dataclass, field

# 导入 ReportConfig 以支持 quantification_methods
try:
    from report_config import ReportConfig
    HAS_REPORT_CONFIG = True
except ImportError:
    HAS_REPORT_CONFIG = False


# ==================== 工具函数 ====================

def excel_date_to_string(date_value):
    """将Excel日期序列号转换为 'YYYY年MM月DD日' 格式"""
    if date_value is None:
        return None
    if isinstance(date_value, str):
        return date_value
    try:
        # Excel日期基准是1899-12-30
        delta = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_value) - 2)
        return delta.strftime('%Y年%m月%d日')
    except (ValueError, TypeError):
        return str(date_value)


# ==================== 协议配置层 ====================

@dataclass
class FieldMapping:
    """字段映射配置"""
    keyword: str           # Excel列关键词
    default: Any          # 默认值
    dtype: str            # 数据类型: 'str', 'float', 'int'
    alt_keywords: List[str] = field(default_factory=list)  # 备选关键词


@dataclass
class TableProtocol:
    """表格协议配置"""
    name: str                     # 协议名称
    description: str              # 描述
    required_keywords: set        # 必需关键词（必须全部匹配）
    optional_keywords: set        # 可选关键词（部分匹配即可）
    field_mappings: Dict[str, FieldMapping]  # 字段映射
    output_var: str               # 输出变量名
    ffill_fields: List[str] = field(default_factory=list)  # 需要前向填充的字段
    min_match_ratio: float = 0.3  # 最小匹配度（可选关键词）
    post_process: Optional[Callable] = None  # 后处理函数
    sheet_name_patterns: List[str] = field(default_factory=list)  # 工作表名称匹配模式
    header_rows_to_check: int = 2  # 表头行检查数量（默认检查2行）


# ==================== 后处理函数 ====================

def group_by_emission_category(items: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    将排放因子数据按类别分组

    分组规则：
    - 固定燃烧 -> scope1_stationary_combustion_emissions_items
    - 移动燃烧 -> scope1_mobile_combustion_emissions_items
    - 制程排放 -> scope1_process_emissions_items
    - 逸散排放 -> scope1_fugitive_emissions_items
    """
    grouped = {
        'scope1_stationary_combustion_emissions_items': [],
        'scope1_mobile_combustion_emissions_items': [],
        'scope1_fugitive_emissions_items': [],
        'scope1_process_emissions_items': [],
    }

    for item in items:
        category = item.get('category', '')
        category_normalized = category.replace(' ', '')

        if '固定燃烧' in category_normalized:
            grouped['scope1_stationary_combustion_emissions_items'].append(item)
        elif '移动燃烧' in category_normalized or '移动汽油' in category_normalized or '移动柴油' in category_normalized:
            grouped['scope1_mobile_combustion_emissions_items'].append(item)
        elif '制冷产品加工使用等排放' in category_normalized or '制程' in category_normalized:
            grouped['scope1_process_emissions_items'].append(item)
        elif '逸散' in category_normalized:
            grouped['scope1_fugitive_emissions_items'].append(item)

    return grouped


def group_scope1_emissions(items: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    将范围一排放数据按类别分组

    与 group_by_emission_category 相同的分组逻辑
    """
    return group_by_emission_category(items)


# ==================== 核心协议配置 ====================

# ==================== 协议优先级配置 ====================
# 优先级高的协议放在前面，先匹配
_PROTOCOL_ORDER = [
    'EmissionFactorProtocol',      # 排放因子表优先级最高（有独特的"低位发热量"+"氧化率"组合）
    'ActivitySummaryProtocol',      # 活动数据汇总表（基于位置）
    'ActivitySummaryMarketProtocol', # 活动数据汇总表（基于市场）
    'GWPProtocol',                  # GWP值表（有独特的"GWP"关键词）
    'GHGInventoryProtocol',         # 温室气体盘查表
    'Scope1EmissionsProtocol',      # 范围一排放源表（最宽泛，放最后）
]

TABLE_PROTOCOLS: Dict[str, TableProtocol] = {
    'EmissionFactorProtocol': TableProtocol(
        name='排放因子表',
        description='包含低位发热量、氧化率、基于热值排放系数等信息的表格',
        required_keywords={'低位发热量', '氧化率'},
        optional_keywords={'基于热值排放系数', '排放因子', 'GHG排放类别', '计算值', '排放系数'},
        field_mappings={
            'category': FieldMapping('GHG排放类别', '', 'str'),
            'number': FieldMapping('编号', '', 'str'),
            'emission_source': FieldMapping('排放源', '', 'str'),
            'facility': FieldMapping('设施', '', 'str'),
            'ncv': FieldMapping('低位发热量', 0, 'float'),
            'unit': FieldMapping('单位', '', 'str'),
            'ox_rate': FieldMapping('氧化率', 0, 'float'),
            # 基于热值排放系数（CO2/CH4/N2O来自"基于热值排放系数"部分）
            'CO2_emission_cv_factor': FieldMapping('CO2', 0, 'float', ['基于热值']),
            'CH4_emission_cv_factor': FieldMapping('CH4', 0, 'float', ['基于热值']),
            'N2O_emission_cv_factor': FieldMapping('N2O', 0, 'float', ['基于热值']),
            # 排放因子（CO2/CH4/N2O来自"排放因子"部分）
            'CO2_emission_factor': FieldMapping('CO2', 0, 'float', ['排放因子']),
            'CH4_emission_factor': FieldMapping('CH4', 0, 'float', ['排放因子']),
            'N2O_emission_factor': FieldMapping('N2O', 0, 'float', ['排放因子']),
        },
        output_var='pro_ef_items',
        ffill_fields=['category'],
    ),

    'GWPProtocol': TableProtocol(
        name='GWP值表',
        description='全球变暖潜势(GWP)值参考表',
        required_keywords={'GWP'},
        optional_keywords={'GWP(HFCs)', 'GWP(PFCs)', '工业名称', '中文名称', '化学分子式'},
        field_mappings={
            'gas_name': FieldMapping('工业名称', '', 'str'),
            'chinese_name': FieldMapping('中文名称/化学分子式', '', 'str'),
            'formula': FieldMapping('中文名称/化学分子式', '', 'str'),
            'composition_ratio': FieldMapping('组成比例', None, 'float'),
            'gwp_value': FieldMapping('GWP', 0, 'float'),
            'gwp_hfcs': FieldMapping('GWP(HFCs)', None, 'float'),
            'gwp_pfcs': FieldMapping('GWP(PFCs)', None, 'float'),
            'source': FieldMapping('来源', '', 'str'),
            'note': FieldMapping('备注', '', 'str'),
        },
        output_var='gwp_items',
    ),

    'GHGInventoryProtocol': TableProtocol(
        name='温室气体盘查表',
        description='温室气体排放盘查汇总表',
        required_keywords={'GHG排放类别', '排放量'},
        optional_keywords={'排放源', '设施', 'GWP', 'EF', '活动数据'},
        field_mappings={
            'category': FieldMapping('GHG排放类别', '', 'str'),
            'emission_source': FieldMapping('排放源', '', 'str'),
            'facility': FieldMapping('设施', '', 'str'),
            'activity_data': FieldMapping('活动数据', 0, 'float'),
            'activity_data_unit': FieldMapping('单位', '', 'str'),
            'emission_factor': FieldMapping('EF', 0, 'float'),
            'gwp': FieldMapping('GWP', 1, 'float'),
            'emissions': FieldMapping('排放量', 0, 'float'),
            'emissions_unit': FieldMapping('tCO2e', '', 'str'),
        },
        output_var='ghg_inventory_items',
    ),

    'Scope1EmissionsProtocol': TableProtocol(
        name='范围一直接排放源表',
        description='从附表1-温室气体盘查表中提取范围一直接排放源数据',
        required_keywords={'编号', 'GHG排放类别'},
        optional_keywords={'排放源', '设施', 'CO2', 'CH4', 'N2O', 'HFCs', 'PFCs', 'SF6', 'NF3', '总量'},
        field_mappings={
            'number': FieldMapping('编号', '', 'str'),
            'category': FieldMapping('GHG排放类别', '', 'str'),
            'emission_source': FieldMapping('排放源', '', 'str'),
            'facility': FieldMapping('设施', '', 'str'),
            'CO2_emissions': FieldMapping('CO2', 0, 'float', ['CO2排放量']),
            'CH4_emissions': FieldMapping('CH4', 0, 'float', ['CH4排放量']),
            'N2O_emissions': FieldMapping('N2O', 0, 'float', ['N2O排放量']),
            'HFCs_emissions': FieldMapping('HFCs', 0, 'float', ['HFCs排放量']),
            'PFCs_emissions': FieldMapping('PFCs', 0, 'float', ['PFCs排放量']),
            'SF6_emissions': FieldMapping('SF6', 0, 'float', ['SF6排放量']),
            'NF3_emissions': FieldMapping('NF3', 0, 'float', ['NF3排放量']),
            'total_green_house_gas_emissions': FieldMapping('总量', 0, 'float', ['总计']),
        },
        output_var='scope1_emissions_items',
        ffill_fields=['category'],
        min_match_ratio=0.5,  # 提高匹配度要求，避免与排放因子表混淆
    ),

    'ActivitySummaryProtocol': TableProtocol(
        name='活动数据汇总表',
        description='基于位置的活动数据汇总表',
        required_keywords={'编号', '排放源'},  # 修正：实际表头是"编号"不是"序号"
        optional_keywords={'GHG', '基于位置', 'CO2', 'CH4', 'N2O', '报告边界', '活动数据'},
        field_mappings={
            'number': FieldMapping('编号', '', 'str', ['序号']),  # 修正：实际表头是"编号"
            'category': FieldMapping('GHG排放类别', '', 'str'),
            'emission_source': FieldMapping('排放源', '', 'str'),
            'report_boundary': FieldMapping('报告边界', '', 'str'),
            'activity_data': FieldMapping('活动数据', 0, 'float'),  # 主表头中的"活动数据"列
            'unit': FieldMapping('计量单位', '', 'str'),  # 子表头中的"计量单位"列
            'CO2_emissions': FieldMapping('CO2', 0, 'float'),
            'CH4_emissions': FieldMapping('CH4', 0, 'float'),
            'N2O_emissions': FieldMapping('N2O', 0, 'float'),
            'HFCs_emissions': FieldMapping('HFCs', 0, 'float'),
            'PFCs_emissions': FieldMapping('PFCs', 0, 'float'),
            'SF6_emissions': FieldMapping('SF6', 0, 'float'),
            'NF3_emissions': FieldMapping('NF3', 0, 'float'),
            'total_green_house_gas_emissions': FieldMapping('总计', 0, 'float', ['总量']),
        },
        output_var='activity_summary_items',
        ffill_fields=['category'],
        min_match_ratio=0.2,  # 降低匹配度要求，因为活动数据汇总表的差异较大
        sheet_name_patterns=['活动数据汇总表', '位置'],  # 基于位置的活动数据汇总表
        header_rows_to_check=3,  # 活动数据汇总表需要检查3行表头
    ),

    'ActivitySummaryMarketProtocol': TableProtocol(
        name='活动数据汇总表（市场法）',
        description='基于市场的活动数据汇总表',
        required_keywords={'编号', '排放源'},  # 修正：实际表头是"编号"不是"序号"
        optional_keywords={'GHG', '基于市场', 'CO2', 'CH4', 'N2O', '报告边界', '活动数据'},
        field_mappings={
            'number': FieldMapping('编号', '', 'str', ['序号']),  # 修正：实际表头是"编号"
            'category': FieldMapping('GHG排放类别', '', 'str'),
            'emission_source': FieldMapping('排放源', '', 'str'),
            'report_boundary': FieldMapping('报告边界', '', 'str'),
            'activity_data': FieldMapping('活动数据', 0, 'float'),  # 主表头中的"活动数据"列
            'unit': FieldMapping('计量单位', '', 'str'),  # 子表头中的"计量单位"列
            'CO2_emissions': FieldMapping('CO2', 0, 'float'),
            'CH4_emissions': FieldMapping('CH4', 0, 'float'),
            'N2O_emissions': FieldMapping('N2O', 0, 'float'),
            'HFCs_emissions': FieldMapping('HFCs', 0, 'float'),
            'PFCs_emissions': FieldMapping('PFCs', 0, 'float'),
            'SF6_emissions': FieldMapping('SF6', 0, 'float'),
            'NF3_emissions': FieldMapping('NF3', 0, 'float'),
            'total_green_house_gas_emissions': FieldMapping('总计', 0, 'float', ['总量']),
        },
        output_var='activity_summary_market_items',
        ffill_fields=['category'],
        min_match_ratio=0.2,  # 降低匹配度要求，因为活动数据汇总表的差异较大
        sheet_name_patterns=['活动数据汇总表', '市场'],  # 基于市场的活动数据汇总表
        header_rows_to_check=3,  # 活动数据汇总表需要检查3行表头
    ),
}


# ==================== 表格指纹识别器 ====================

class TableFingerprint:
    """表格指纹识别器"""

    def __init__(self, protocols: Dict[str, TableProtocol]):
        self.protocols = protocols

    def identify(self, sheet: openpyxl.worksheet.Worksheet,
                 sheet_name: str = None,
                 check_rows: int = 20) -> Optional[str]:
        """
        识别工作表的表格类型

        按照协议优先级顺序进行匹配，找到第一个匹配的协议就返回

        Args:
            sheet: openpyxl工作表对象
            sheet_name: 工作表名称
            check_rows: 检查前N行

        Returns:
            匹配的协议名称，未匹配返回None
        """
        # 收集表格的唯一字符串值
        unique_strings = self._extract_unique_strings(sheet, check_rows)

        actual_sheet_name = sheet_name or sheet.title

        # 先尝试精确的工作表名称匹配（优先级最高）
        for protocol_name in _PROTOCOL_ORDER:
            if protocol_name not in self.protocols:
                continue

            protocol = self.protocols[protocol_name]

            # 检查工作表名称模式
            if protocol.sheet_name_patterns:
                # 检查所有模式是否都匹配
                all_patterns_match = all(pattern and pattern in actual_sheet_name for pattern in protocol.sheet_name_patterns)
                if all_patterns_match:
                    print(f"[表格识别] '{actual_sheet_name}' 通过名称匹配到 {protocol.name} (patterns: {protocol.sheet_name_patterns})")
                    return protocol_name

        # 如果没有名称匹配，则按关键词匹配
        for protocol_name in _PROTOCOL_ORDER:
            if protocol_name not in self.protocols:
                continue

            protocol = self.protocols[protocol_name]

            # 跳过有名称模式但未匹配的协议（因为它们在前面已经检查过了）
            if protocol.sheet_name_patterns:
                continue

            # 检查必需关键词是否全部匹配
            if not protocol.required_keywords.issubset(unique_strings):
                continue

            # 计算可选关键词匹配度
            optional_matched = len(protocol.optional_keywords & unique_strings)
            optional_total = len(protocol.optional_keywords) if protocol.optional_keywords else 1
            match_score = optional_matched / optional_total if optional_total > 0 else 1.0

            # 检查是否达到最小匹配度
            if match_score >= protocol.min_match_ratio:
                print(f"[表格识别] {sheet_name or sheet.title} 匹配到 {protocol.name} "
                      f"(必需关键词: {len(protocol.required_keywords)}/{len(protocol.required_keywords)}, "
                      f"可选关键词: {optional_matched}/{optional_total})")
                return protocol_name

        if sheet_name:
            print(f"[表格识别] {sheet_name} 未匹配到已知协议类型")
        return None

    def _extract_unique_strings(self, sheet: openpyxl.worksheet.Worksheet,
                               check_rows: int) -> set:
        """提取表格中的唯一字符串"""
        unique_strings = set()
        for row_idx in range(1, min(check_rows + 1, sheet.max_row + 1)):
            for cell in sheet[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    cleaned = str(cell.value).strip()
                    if cleaned:
                        unique_strings.add(cleaned)
        return unique_strings

    def _calculate_match_score(self, unique_strings: set,
                              protocol: TableProtocol) -> float:
        """计算协议匹配度"""
        # 检查必需关键词
        if not protocol.required_keywords.issubset(unique_strings):
            return 0.0

        # 计算可选关键词匹配度
        optional_matched = len(protocol.optional_keywords & unique_strings)
        optional_total = len(protocol.optional_keywords) if protocol.optional_keywords else 1

        match_ratio = optional_matched / optional_total if optional_total > 0 else 1.0
        return match_ratio


# ==================== 协议提取器 ====================

class ProtocolExtractor:
    """协议数据提取器"""

    def __init__(self):
        self.fingerprint = TableFingerprint(TABLE_PROTOCOLS)

    def extract_from_sheet(self, sheet: openpyxl.worksheet.Worksheet,
                          protocol_name: str) -> List[Dict[str, Any]]:
        """
        根据协议从工作表提取数据

        Args:
            sheet: openpyxl工作表对象
            protocol_name: 协议名称

        Returns:
            提取的数据列表
        """
        if protocol_name not in TABLE_PROTOCOLS:
            print(f"[数据提取] 未知协议: {protocol_name}")
            return []

        protocol = TABLE_PROTOCOLS[protocol_name]

        # 查找表头行
        header_row = self._find_header_row(sheet, protocol)
        if not header_row:
            print(f"[数据提取] 未找到表头行")
            return []

        # 获取列映射
        column_map = self._build_column_map(sheet, header_row, protocol)
        if not column_map:
            print(f"[数据提取] 无法创建列映射")
            return []

        # 提取数据
        data_items = self._extract_data_rows(sheet, header_row, column_map, protocol)

        # 应用前向填充
        if protocol.ffill_fields:
            data_items = self._apply_ffill(data_items, protocol.ffill_fields)

        # 应用后处理
        if protocol.post_process:
            processed = protocol.post_process(data_items)
            if processed is not None:
                data_items = processed

        print(f"[数据提取] 提取到 {len(data_items)} 行数据")
        return data_items

    def _find_header_row(self, sheet: openpyxl.worksheet.Worksheet,
                        protocol: TableProtocol, max_row: int = 20) -> Optional[int]:
        """查找表头行"""
        all_keywords = protocol.required_keywords | protocol.optional_keywords

        for row_idx in range(1, min(max_row + 1, sheet.max_row + 1)):
            row_values = set()
            for cell in sheet[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    row_values.add(str(cell.value).strip())

            # 至少匹配2个关键词
            matched = len(all_keywords & row_values)
            if matched >= min(2, len(protocol.required_keywords)):
                return row_idx

        return None

    def _build_column_map(self, sheet: openpyxl.worksheet.Worksheet,
                         header_row: int,
                         protocol: TableProtocol) -> Dict[str, int]:
        """构建列名到列索引的映射"""
        column_map = {}
        header_cells = {}

        # 收集表头单元格（支持多行表头）
        rows_to_check = list(range(header_row, min(header_row + protocol.header_rows_to_check, sheet.max_row + 1)))
        for row_idx in rows_to_check:
            if row_idx > sheet.max_row:
                continue
            for cell in sheet[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    value = str(cell.value).strip()
                    if value and value not in header_cells:
                        header_cells[value] = cell.column - 1

        # 活动数据汇总表的排放量字段特殊处理：使用固定列位置
        is_activity_summary = protocol.name in ['活动数据汇总表', '活动数据汇总表（市场法）']
        # 排放因子表的特殊处理：使用固定列位置区分两组 CO2/CH4/N2O
        is_emission_factor = protocol.name == '排放因子表'

        if is_activity_summary:
            # 活动数据汇总表的排放量列固定位置（基于表格结构）
            # CO2: Col 30, CH4: Col 31, N2O: Col 32, HFCs: Col 33
            # PFCs: Col 34, SF6: Col 35, NF3: Col 36, 总计: Col 37
            emission_column_map = {
                'CO2_emissions': 29,  # Col 30 (0-based)
                'CH4_emissions': 30,  # Col 31
                'N2O_emissions': 31,  # Col 32
                'HFCs_emissions': 32,  # Col 33
                'PFCs_emissions': 33,  # Col 34
                'SF6_emissions': 34,  # Col 35
                'NF3_emissions': 35,  # Col 36
                'total_green_house_gas_emissions': 36,  # Col 37
            }
        elif is_emission_factor:
            # 排放因子表的固定列位置
            # 基于热值排放系数: CO2=Col9, CH4=Col10, N2O=Col11
            # 排放因子: CO2=Col12, CH4=Col13, N2O=Col14
            emission_column_map = {
                'CO2_emission_cv_factor': 8,   # Col 9 (基于热值排放系数 CO2)
                'CH4_emission_cv_factor': 9,   # Col 10 (基于热值排放系数 CH4)
                'N2O_emission_cv_factor': 10,  # Col 11 (基于热值排放系数 N2O)
                'CO2_emission_factor': 11,     # Col 12 (排放因子 CO2)
                'CH4_emission_factor': 12,     # Col 13 (排放因子 CH4)
                'N2O_emission_factor': 13,     # Col 14 (排放因子 N2O)
            }

        # 为每个字段查找列
        for field_name, field_mapping in protocol.field_mappings.items():
            # 活动数据汇总表的排放量字段：使用固定列位置
            if is_activity_summary and field_name in emission_column_map:
                column_map[field_name] = emission_column_map[field_name]
            # 排放因子表的排放系数字段：使用固定列位置
            elif is_emission_factor and field_name in emission_column_map:
                column_map[field_name] = emission_column_map[field_name]
            else:
                # 常规匹配
                if field_mapping.keyword in header_cells:
                    column_map[field_name] = header_cells[field_mapping.keyword]
                else:
                    # 模糊匹配
                    for header_value in header_cells:
                        if field_mapping.keyword in header_value or header_value in field_mapping.keyword:
                            column_map[field_name] = header_cells[header_value]
                            break
                    # 尝试备选关键词
                    if field_name not in column_map:
                        for alt_keyword in field_mapping.alt_keywords:
                            if alt_keyword in header_cells:
                                column_map[field_name] = header_cells[alt_keyword]
                                break

        return column_map

    def _extract_data_rows(self, sheet: openpyxl.worksheet.Worksheet,
                          header_row: int,
                          column_map: Dict[str, int],
                          protocol: TableProtocol) -> List[Dict[str, Any]]:
        """提取数据行"""
        data_items = []

        # 活动数据汇总表的特殊处理：跳过子表头行
        is_activity_summary = protocol.name in ['活动数据汇总表', '活动数据汇总表（市场法）']

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_idx]

            # 检查空行
            if not any(cell.value is not None for cell in row):
                continue

            # 活动数据汇总表：跳过序号列非数字的行（子表头行）
            if is_activity_summary and 'number' in protocol.field_mappings:
                number_col_idx = column_map.get('number', 0)
                if number_col_idx < len(row):
                    number_cell = row[number_col_idx]
                    # 如果序号列不是数字，跳过该行
                    if number_cell.value is None or not isinstance(number_cell.value, (int, float)):
                        try:
                            float(number_cell.value)
                        except (ValueError, TypeError):
                            continue

            # 提取字段值
            item = {}
            for field_name, field_mapping in protocol.field_mappings.items():
                if field_name in column_map:
                    col_idx = column_map[field_name]
                    cell = row[col_idx] if col_idx < len(row) else None

                    value = self._convert_cell_value(cell, field_mapping)
                    item[field_name] = value
                else:
                    item[field_name] = field_mapping.default

            # 检查是否有效数据行
            if self._is_valid_row(item, protocol):
                data_items.append(item)

        return data_items

    def _convert_cell_value(self, cell, field_mapping: FieldMapping) -> Any:
        """转换单元格值"""
        if cell is None or cell.value is None:
            return field_mapping.default

        try:
            if field_mapping.dtype == 'float':
                return float(cell.value)
            elif field_mapping.dtype == 'int':
                return int(cell.value)
            else:
                return str(cell.value).strip()
        except (ValueError, TypeError):
            return field_mapping.default

    def _is_valid_row(self, item: Dict[str, Any], protocol: TableProtocol) -> bool:
        """检查是否为有效数据行"""
        for field_name, field_mapping in protocol.field_mappings.items():
            value = item.get(field_name)
            if value and value != field_mapping.default:
                return True
        return False

    def _apply_ffill(self, data_items: List[Dict[str, Any]],
                    field_names: List[str]) -> List[Dict[str, Any]]:
        """应用前向填充"""
        if not data_items:
            return data_items

        result = []
        last_values = {field: None for field in field_names}

        for item in data_items:
            new_item = item.copy()
            for field in field_names:
                value = item.get(field)
                if value and str(value).strip():
                    last_values[field] = value
                    new_item[field] = value
                elif last_values[field]:
                    new_item[field] = last_values[field]
            result.append(new_item)

        return result


# ==================== 高层数据读取器 ====================

class ExcelDataReaderRefactored:
    """重构后的Excel数据读取器"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.extractor = ProtocolExtractor()
        self.fingerprint = TableFingerprint(TABLE_PROTOCOLS)

    def get_all_context(self) -> Dict[str, Any]:
        """
        获取所有渲染上下文数据

        这是唯一的高层接口，自动识别所有表格并提取数据

        Returns:
            包含所有提取数据的字典
        """
        result = {}

        print(f"[数据读取] 开始处理工作簿，共 {len(self.workbook.sheetnames)} 个工作表")

        # ========== 首先提取基本信息 ==========
        basic_info = self._extract_basic_info()
        result.update(basic_info)

        # ========== 提取范围三类别数据 ==========
        scope3_data = self._extract_scope3_categories()
        result.update(scope3_data)

        # ========== 提取范围二数据 ==========
        scope2_data = self._extract_scope2_data()
        result.update(scope2_data)

        # ========== 添加 quantification_methods ==========
        if HAS_REPORT_CONFIG and result.get('company_name'):
            report_config = ReportConfig(
                company_name=result['company_name'],
                reporting_period=result.get('reporting_period', '2024年')
            )
            result['quantification_methods'] = report_config.get_quantification_methods()
        else:
            # 提供默认的空字典
            result['quantification_methods'] = {}

        # ========== 提取范围三详细数据（从温室气体盘查清册）==========
        scope3_detail_data = self._extract_scope3_detail_data()
        result.update(scope3_detail_data)

        # ========== 提取表1和表2的数据（从表1温室气体盘查表）==========
        table1_table2_data = self._extract_table1_table2_data()
        result.update(table1_table2_data)

        # ========== 提取范围二数据（从温室气体盘查表）==========
        scope2_items_data = self._extract_scope2_items()
        result.update(scope2_items_data)

        # ========== 遍历所有工作表，识别并提取表格数据 ==========
        for sheet in self.workbook.worksheets:
            sheet_name = sheet.title

            # 识别表格类型
            protocol_name = self.fingerprint.identify(sheet, sheet_name)

            if protocol_name:
                protocol = TABLE_PROTOCOLS[protocol_name]
                output_var = protocol.output_var

                # 排放因子表特殊处理：多子表提取
                if protocol_name == 'EmissionFactorProtocol' and '附表2-EF' in sheet_name:
                    data_items = self._extract_emission_factor_subtables(sheet)
                else:
                    # 提取数据
                    data_items = self.extractor.extract_from_sheet(sheet, protocol_name)

                # 存储结果
                result[output_var] = data_items
                print(f"[数据读取] {sheet_name} -> {output_var}: {len(data_items)} 行")

        # 确保所有输出变量都被初始化
        for protocol in TABLE_PROTOCOLS.values():
            if protocol.output_var not in result:
                result[protocol.output_var] = []

        # ========== 后处理：类别分组 ==========
        # 1. 处理排放因子表的类别分组 - 分别创建三个表格的数据
        if 'pro_ef_items' in result and result['pro_ef_items']:
            # 按子表类型分组数据
            combustion_items = []  # 表格22：燃烧排放因子表（固定燃烧、移动燃烧）
            process_items = []       # 表格23：制程排放因子表
            fugitive_items = []      # 表格24：逸散排放表

            for item in result['pro_ef_items']:
                category = item.get('category', '')

                # 根据类别判断子表类型
                if '固定燃烧' in category or '移动燃烧' in category:
                    # 燃烧排放因子表（表格22）
                    mapped_item = item.copy()
                    # 添加模板期望的字段映射
                    mapped_item['emission_source_type_dir'] = item.get('category', '')
                    mapped_item['emission_source_dir'] = item.get('emission_source', '')
                    mapped_item['emission_facilities_dir'] = item.get('facility', '')
                    mapped_item['ncv_dir'] = item.get('ncv', '')
                    mapped_item['emission_unit_dir'] = item.get('unit', '')
                    mapped_item['emission_oa_dir'] = item.get('ox_rate', '')
                    mapped_item['CO2_emission_cv_factor'] = item.get('CO2_emission_cv_factor', '')
                    mapped_item['CH4_emission_cv_factor'] = item.get('CH4_emission_cv_factor', '')
                    mapped_item['N2O_emission_cv_factor'] = item.get('N2O_emission_cv_factor', '')
                    # 排放因子字段已在数据中
                    combustion_items.append(mapped_item)

                elif '制程排放' in category:
                    # 制程排放因子表（表格23）
                    mapped_item = item.copy()
                    # 添加模板期望的字段映射
                    mapped_item['emission_source_type_dir'] = item.get('category', '')
                    mapped_item['emission_source_dir'] = item.get('emission_source', '')
                    mapped_item['emission_facilities_dir'] = item.get('facility', '')
                    # 制程排放只有CO2排放因子
                    mapped_item['CO2_emission_factor'] = item.get('CO2_emission_factor', '')
                    process_items.append(mapped_item)

                elif '逸散排放' in category:
                    # 逸散排放表（表格24）
                    mapped_item = item.copy()
                    # 添加模板期望的字段映射
                    mapped_item['emission_source_type_dir'] = item.get('category', '')
                    mapped_item['emission_source_dir'] = item.get('emission_source', '')
                    mapped_item['emission_facilities_dir'] = item.get('facility', '')
                    # 逸散排放特有字段
                    # 注意：需要从原始数据中获取 HFCs/PFCs, MCF, Bo 等字段
                    # 这些字段在 _extract_fugitive_ef_row 中可能没有正确提取
                    mapped_item['HFCs_PCFs_emission_factor'] = item.get('CO2_emission_factor', '')  # 使用CO2_emission_factor作为默认
                    mapped_item['MCF'] = ''  # 需要补充
                    mapped_item['Bo'] = ''    # 需要补充
                    mapped_item['emission_factor'] = item.get('CO2_emission_factor', '')
                    mapped_item['emission_unit_dir'] = item.get('unit', '')
                    fugitive_items.append(mapped_item)

            # 设置三个表格的数据
            result['emission_factor_combustion_items'] = combustion_items
            result['emission_factor_process_items'] = process_items
            result['emission_factor_fugitive_items'] = fugitive_items

            # 向后兼容：设置 emission_factor_items（包含所有数据）
            result['emission_factor_items'] = combustion_items + process_items + fugitive_items

            print("[后处理] 排放因子表数据分组:")
            print(f"  emission_factor_combustion_items (表格22-燃烧): {len(combustion_items)} 条")
            print(f"  emission_factor_process_items (表格23-制程): {len(process_items)} 条")
            print(f"  emission_factor_fugitive_items (表格24-逸散): {len(fugitive_items)} 条")
            print(f"  emission_factor_items (总计): {len(result['emission_factor_items'])} 条")

            # ========== 处理外购能源间接排放因子（范围二排放因子）==========
            # 从附表2-EF中提取外购能源间接排放因子数据
            scope2_ef_raw_items = [item for item in result.get('pro_ef_items', [])
                                   if '范围二' in item.get('category', '') and '外购能源' in item.get('category', '')]

            # ========== 处理外购能源间接排放因子（范围二排放因子）==========
            # 创建模板期望的indir_ef_items格式
            indir_ef_items = []
            for item in scope2_ef_raw_items:
                mapped_item = {
                    'number': item.get('number', ''),
                    'emission_source_type_indir': item.get('category', ''),
                    'emission_source_indir': item.get('emission_source', ''),
                    'emission_facilities_indir': item.get('facility', ''),
                    'elec_emission_factor': item.get('CO2_emission_factor', ''),
                    'elec_emission_unit': item.get('unit', ''),
                    'elec_emission_source': item.get('emission_source_reference', ''),
                }
                indir_ef_items.append(mapped_item)

            result['indir_ef_items'] = indir_ef_items
            print(f"  indir_ef_items (外购能源间接排放因子): {len(indir_ef_items)} 条")

            # ========== 处理范围三所有类别排放因子（cat1-cat15）==========
            # 支持新格式："范围三 类别N XXX"（N为1-15）
            # 兼容旧格式：直接使用类别名称

            # 首先按类别分组，提取类别编号
            category_groups = {}  # {cat_num: [items]}
            for item in result.get('pro_ef_items', []):
                category = item.get('category', '')
                cat_num = None

                # 尝试从类别名称中提取编号 (新格式)
                # 匹配 "范围三 类别N" 或 "范围3 类别N" (N为1-15)
                # 使用更精确的匹配，避免"类别1"匹配到"类别11"
                for cat_id in range(15, 0, -1):  # 从大到小匹配，避免部分匹配
                    if f'范围三 类别{cat_id}' in category or f'范围三类别{cat_id}' in category or f'范围3 类别{cat_id}' in category:
                        cat_num = cat_id
                        break

                if not cat_num:
                    # 使用旧格式的映射 (向后兼容)
                    legacy_mapping = {
                        '外购商品和服务的上游排放': 1,
                        '资本货物': 2,
                        '范围一、二之外燃料和能源相关的活动产生的排放': 3,
                        '上下游运输配送产生的排放': 4,
                        '运营中产生的废物排放': 5,
                        '商务旅行产生的排放': 6,
                        '员工通勤': 7,
                        '上游租赁资产': 8,
                        '下游运输配送': 9,
                        '外销产品加工': 10,
                        '外销产品使用': 11,
                        '外售产品报废': 12,
                    }
                    cat_num = legacy_mapping.get(category)

                if cat_num and 1 <= cat_num <= 15:
                    if cat_num not in category_groups:
                        category_groups[cat_num] = []
                    category_groups[cat_num].append(item)

            # 为所有类别1-15创建对应的变量（即使为空）
            for cat_num in range(1, 16):
                cat_prefix = f'cat{cat_num}'
                items = category_groups.get(cat_num, [])

                cat_ef_items = []
                for item in items:
                    # 基础字段映射
                    mapped_item = {
                        'number': item.get('number', ''),
                        f'emission_source_type_{cat_prefix}': item.get('category', ''),
                        f'emission_source_{cat_prefix}': item.get('emission_source', ''),
                        f'emission_name_{cat_prefix}': item.get('activity_name', ''),
                        f'emission_geo_{cat_prefix}': item.get('geography', ''),
                        f'{cat_prefix}_emission_factor': item.get('CO2_emission_factor', ''),
                        f'{cat_prefix}_emission_unit': item.get('unit', ''),
                        f'{cat_prefix}_emission_source': item.get('emission_source_reference', ''),
                    }

                    # 检查是否是燃烧表格式（包含ncv字段）
                    # 如果是，添加燃烧表特有字段
                    if item.get('ncv') is not None and item.get('ncv') != 0:
                        mapped_item[f'ncv_{cat_prefix}'] = item.get('ncv', '')
                        mapped_item[f'emission_unit_{cat_prefix}'] = item.get('unit', '')
                        mapped_item[f'emission_oa_{cat_prefix}'] = item.get('ox_rate', '')
                        # 基于热值的排放因子字段
                        mapped_item[f'CO2_emission_cv_factor'] = item.get('CO2_emission_cv_factor', '')
                        mapped_item[f'CH4_emission_cv_factor'] = item.get('CH4_emission_cv_factor', '')
                        mapped_item[f'N2O_emission_cv_factor'] = item.get('N2O_emission_cv_factor', '')
                        # 基于体积的排放因子字段 - 修复：添加CO2排放因子映射
                        mapped_item[f'CO2_emission_factor'] = item.get('CO2_emission_factor', '')
                        mapped_item[f'CH4_emission_factor'] = item.get('CH4_emission_factor', '')
                        mapped_item[f'N2O_emission_factor'] = item.get('N2O_emission_factor', '')

                    cat_ef_items.append(mapped_item)

                result[f'{cat_prefix}_ef_items'] = cat_ef_items
                if items:
                    category_name = items[0].get('category', '')[:40]
                    print(f"  {cat_prefix}_ef_items (类别{cat_num}): {len(cat_ef_items)} 条")
                else:
                    print(f"  {cat_prefix}_ef_items (类别{cat_num}): 0 条 - 无数据")

            # 按类别分组（用于范围一排放数据）
            grouped_data = group_by_emission_category(result['pro_ef_items'])
            result.update(grouped_data)

        # 2. 处理范围一排放数据的类别分组
        if 'scope1_emissions_items' in result and result['scope1_emissions_items']:
            grouped_data = group_scope1_emissions(result['scope1_emissions_items'])
            result.update(grouped_data)
            print("[后处理] 范围一排放按类别分组:")
            for group_name, items in grouped_data.items():
                if items:
                    print(f"  {group_name}: {len(items)} 条")

        # ========== 提取范围一详细数据（从温室气体盘查清册表）==========
        # 放在后处理之后，以覆盖排放因子的分组数据
        scope1_detail_data = self._extract_scope1_detail_from_inventory_sheet()
        result.update(scope1_detail_data)

        # ========== 注意：不再覆盖 scope2_items ==========
        # scope2_items 由 _extract_table1_table2_data() 从表1温室气体盘查表提取
        # 这里保留原有的 scope2_items，不覆盖为汇总行

        # ========== 设置活动数据汇总表（向后兼容模板）==========
        # activity_summary_items 现在是基于位置的数据
        # activity_summary_market_items 是基于市场的数据

        # 处理基于位置的活动数据汇总表
        if 'activity_summary_items' in result:
            # 字段映射：新版字段名 -> 模板期望的字段名
            location_items = []
            for item in result['activity_summary_items']:
                mapped_item = {
                    'number': item.get('number', ''),
                    # GHG排放类别
                    'emission_source_type_loc': item.get('category', ''),
                    'emission_source_type_location_based': item.get('category', ''),
                    # 排放源
                    'emission_source_loc': item.get('emission_source', ''),
                    'emission_source_location_based': item.get('emission_source', ''),
                    # 报告边界
                    'report_boundary_loc': item.get('report_boundary', ''),
                    'report_boundary_location_based': item.get('report_boundary', ''),
                    # 活动数据数值（模板期望的字段名）
                    'act_summary_loc': item.get('activity_data', ''),
                    'activity_data_location_based': item.get('activity_data', ''),
                    # 活动数据单位（模板期望的字段名）
                    'act_summary_loc_unit': item.get('unit', ''),
                    'activity_data_unit_location_based': item.get('unit', ''),
                }
                # 排放量数据
                for field_name in ['CO2_emissions', 'CH4_emissions', 'N2O_emissions',
                                   'HFCs_emissions', 'PFCs_emissions', 'SF6_emissions',
                                   'NF3_emissions', 'total_green_house_gas_emissions']:
                    if field_name in item:
                        mapped_item[field_name] = item[field_name]
                location_items.append(mapped_item)

            # 设置模板期望的变量名
            result['act_summary_loc'] = location_items
            result['activity_summary_location_items'] = location_items
            print(f"[后处理] act_summary_loc (基于位置): {len(location_items)} 行")

        # 处理基于市场的活动数据汇总表
        if 'activity_summary_market_items' in result:
            # 字段映射：新版字段名 -> 模板期望的字段名
            market_items = []
            for item in result['activity_summary_market_items']:
                mapped_item = {
                    'number': item.get('number', ''),
                    # GHG排放类别
                    'emission_source_type_mar': item.get('category', ''),  # 简短版本（模板期望）
                    'emission_source_type_market_based': item.get('category', ''),  # 完整版本（兼容）
                    # 排放源
                    'emission_source_mar': item.get('emission_source', ''),
                    'emission_source_market_based': item.get('emission_source', ''),
                    # 报告边界
                    'report_boundary_mar': item.get('report_boundary', ''),
                    'report_boundary_market_based': item.get('report_boundary', ''),
                    # 活动数据数值（模板期望的字段名）
                    'act_summary_mar': item.get('activity_data', ''),
                    'activity_data_market_based': item.get('activity_data', ''),
                    # 活动数据单位（模板期望的字段名）
                    'act_summary_mar_unit': item.get('unit', ''),
                    'activity_data_unit_market_based': item.get('unit', ''),
                }
                # 排放量数据
                for field_name in ['CO2_emissions', 'CH4_emissions', 'N2O_emissions',
                                   'HFCs_emissions', 'PFCs_emissions', 'SF6_emissions',
                                   'NF3_emissions', 'total_green_house_gas_emissions']:
                    if field_name in item:
                        mapped_item[field_name] = item[field_name]
                market_items.append(mapped_item)

            # 设置模板期望的变量名
            result['act_summary_mar'] = market_items
            print(f"[后处理] act_summary_mar (基于市场): {len(market_items)} 行")

        # ========== 后处理：更新 Flags 标记 ==========
        result = self._update_flags(result)

        return result

    def _extract_scope3_detail_data(self) -> Dict[str, Any]:
        """从温室气体盘查清册中提取范围三详细数据"""
        result = {}
        for i in range(1, 16):
            result[f'scope3_category{i}'] = []

        # 查找表1温室气体盘查表
        table1_sheet = None
        for sheet in self.workbook.worksheets:
            if '表1' in sheet.title and '温室气体盘查表' in sheet.title:
                table1_sheet = sheet
                break

        if not table1_sheet:
            print("[范围三详细] 未找到表1温室气体盘查表")
            return result

        print("[范围三详细] 从表1温室气体盘查表提取范围三详细数据")

        # 范围三类别名称映射
        category_names = {
            '类别1': '外购商品和服务上游排放',
            '类别2': '资本货物上游排放',
            '类别3': '燃料和能源相关活动未包含在范围一和范围二中的上游排放',
            '类别4': '上游运输和配送',
            '类别5': '运营中产生的废弃物',
            '类别6': '员工商务旅行',
            '类别7': '员工通勤',
            '类别8': '上游租赁资产',
            '类别9': '下游运输和配送',
            '类别10': '销售产品的加工',
            '类别11': '销售产品的使用',
            '类别12': '售出产品的加工',
            '类别13': '下游租赁资产',
            '类别14': '特许经营',
            '类别15': '投资'
        }

        # 收集详细排放源行（按类别分组）
        category_detail_rows = {}
        for row_idx, row in enumerate(table1_sheet.iter_rows(min_row=5, values_only=True), start=5):
            row_vals = [str(v) if v is not None else '' for v in row[:10]]

            # 检查是否是数据行（第1列是编号）
            if not row_vals[0] or not row_vals[0].strip():
                continue

            # 检查是否是范围三数据（第5列包含"范围三"）
            if len(row_vals) > 4 and '范围三' in row_vals[4]:
                # 提取类别编号
                category_match = re.search(r'类别\s*(\d+)', row_vals[4])
                if category_match:
                    category_num = int(category_match.group(1))
                    if category_num not in category_detail_rows:
                        category_detail_rows[category_num] = []
                    category_detail_rows[category_num].append({
                        'row_idx': row_idx,
                        'data': row
                    })

        # 为每个类别创建数据项
        for category_num in sorted(category_detail_rows.keys()):
            detail_rows = category_detail_rows[category_num]
            category_key = f'类别{category_num}'
            category_var_name = f'scope3_category{category_num}'

            category_items = []
            sub_num = 0
            for row_info in detail_rows:
                row = row_info['data']
                row_vals = [str(v) if v is not None else '' for v in row[:10]]

                # 提取数据
                number = row_vals[0] if len(row_vals) > 0 else ''
                category = row_vals[1] if len(row_vals) > 1 else ''
                emission_source = row_vals[2] if len(row_vals) > 2 else category
                facility = row_vals[3] if len(row_vals) > 3 else ''
                activity_data = self._safe_float(row[5]) if len(row) > 5 else 0
                emission_factor = self._safe_float(row[7]) if len(row) > 7 else 0
                factor_unit = row_vals[8] if len(row) > 8 else ''

                # 计算排放量
                calculated_emission = activity_data * emission_factor

                if calculated_emission > 0.01:
                    sub_num += 1
                    total_formatted = f"{calculated_emission:,.2f}"

                    # 构造排放源名称（包含设施信息）
                    emission_source_name = emission_source
                    if facility and facility != 'None':
                        emission_source_name = f"{emission_source}（{facility}）"

                    category_items.append({
                        'number': f'3.{category_num}.{sub_num}',
                        'emission_source': emission_source_name,
                        'total_green_house_gas_emissions': total_formatted,
                        'CO2_emissions': total_formatted,
                        'CH4_emissions': '0.00',
                        'N2O_emissions': '0.00',
                        'HFCs_emissions': '0.00',
                        'PFCs_emissions': '0.00',
                        'SFs_emissions': '0.00',
                        'NF3_emissions': '0.00'
                    })

            result[category_var_name] = category_items
            print(f"  {category_var_name}: {len(category_items)} 行")

        return result

    def _extract_table1_table2_data(self) -> Dict[str, Any]:
        """从表1温室气体盘查表中提取表1和表2的数据"""
        result = {
            'scope1_items': [],  # 表1：范围一直接排放源
            'scope2_3_items': [],  # 表2：范围二三间接排放源
        }

        # 查找表1温室气体盘查表
        table1_data = None
        for sheet in self.workbook.worksheets:
            if '表1' in sheet.title and '温室气体盘查表' in sheet.title:
                table1_data = sheet
                break

        if table1_data:
            print(f"[表1表2] 找到表1温室气体盘查表: {table1_data.title}")

            # 维护类别变量用于前向填充（ffill）
            current_category = ""

            # 提取数据（从第5行开始）
            for row in table1_data.iter_rows(min_row=5):
                if len(row) < 7:
                    continue

                # 获取各列数据
                seq = row[0].value  # 序号
                ghg_category = row[1].value  # GHG排放类别（第2列，索引为1）
                emission_source = row[2].value  # 排放源
                facility = row[3].value  # 设施
                boundary = row[4].value  # 组织边界/排放边界

                # 实现前向填充逻辑（ffill）：如果当前行的类别为空，使用上一行的类别
                if ghg_category:
                    current_category = str(ghg_category).strip()
                # 即使 ghg_category 为空，也使用 current_category

                # 跳过空行或标题行
                if not seq and not current_category:
                    continue

                seq_str = str(seq).strip() if seq else ''
                # 使用 current_category（前向填充后的类别）
                ghg_str = current_category if current_category else ''
                source_str = str(emission_source).strip() if emission_source else ''
                facility_str = str(facility).strip() if facility else ''
                boundary_str = str(boundary).strip() if boundary else ''

                # 跳过标题行
                if seq_str == '序号' or ghg_str == 'GHG排放类别':
                    continue

                # 表1：范围一
                if '范围一' in boundary_str:
                    result['scope1_items'].append({
                        'name': ghg_str,  # GHG排放类别（使用前向填充的值）
                        'number': seq_str,  # 序号
                        'emission_source': source_str,  # 排放源
                        'facility': facility_str  # 设施
                    })

                # 表2：范围二三
                elif '范围二' in boundary_str or '范围三' in boundary_str:
                    result['scope2_3_items'].append({
                        'name': ghg_str,  # GHG排放类别（使用前向填充的值）
                        'number': seq_str,  # 序号
                        'emission_source': source_str,  # 排放源
                        'facility': facility_str  # 设施
                    })

            print(f"[表1表2] scope1_items: {len(result['scope1_items'])} 行")
            print(f"[表1表2] scope2_3_items: {len(result['scope2_3_items'])} 行")

        return result

    def _extract_scope2_items(self) -> Dict[str, Any]:
        """从温室气体盘查表中提取范围二输入能源的间接排放清册数据"""
        result = {'scope2_items': []}

        # 查找温室气体盘查表
        pandata_sheet = None
        for sheet in self.workbook.worksheets:
            if '盘查表' in str(sheet.title):
                pandata_sheet = sheet
                break

        if pandata_sheet:
            print(f"[范围二] 找到温室气体盘查表: {pandata_sheet.title}")
            scope2_items = []

            # 查找包含"汇总"和"外购电力"的行
            location_total = None
            market_total = None

            # 尝试匹配包含"汇总"或"Total"的行
            total_keywords = ['汇总', '总计', 'Total', 'TOTAL']

            for row_idx, row in enumerate(pandata_sheet.iter_rows(min_row=1, values_only=True), start=1):
                first_col = str(row[0]).strip() if row[0] else ''
                second_col = str(row[1]).strip() if len(row) > 1 else ''
                fourth_col = str(row[4]).strip() if len(row) > 4 else ''

                # 检查是否是汇总行
                is_total_row = any(keyword in first_col or first_col == keyword for keyword in total_keywords)
                has_electricity = '外购' in second_col or '电力' in second_col

                if is_total_row and has_electricity:
                    # 检查第四列来判断是基于位置还是基于市场
                    is_location = '位置' in fourth_col or 'Location' in fourth_col or 'location' in fourth_col
                    is_market = '市场' in fourth_col or 'Market' in fourth_col or 'market' in fourth_col

                    if is_location:
                        location_total = row
                        print(f"  找到外购电力（基于位置）汇总: Row {row_idx}")
                    elif is_market:
                        market_total = row
                        print(f"  找到外购电力（基于市场）汇总: Row {row_idx}")

            # 构建两行数据
            if location_total:
                total = self._safe_float(location_total[2]) if len(location_total) > 2 else 0
                co2 = self._safe_float(location_total[3]) if len(location_total) > 3 else total

                total_formatted = f"{total:,.2f}" if total > 0 else "0.00"
                co2_formatted = f"{co2:,.2f}" if co2 > 0 else "0.00"

                scope2_items.append({
                    'number': '2.1',
                    'emission_source': '外购电力（基于位置）',
                    'total_green_house_gas_emissions': total_formatted,
                    'CO2_emissions': co2_formatted,
                    'CH4_emissions': '0.00',
                    'N2O_emissions': '0.00',
                    'HFCs_emissions': '0.00',
                    'PFCs_emissions': '0.00',
                    'SFs_emissions': '0.00',
                    'NF3_emissions': '0.00'
                })
                print(f"  提取2.1 外购电力（基于位置）: {total_formatted} tCO2e")

            if market_total:
                total = self._safe_float(market_total[2]) if len(market_total) > 2 else 0
                co2 = self._safe_float(market_total[3]) if len(market_total) > 3 else total

                total_formatted = f"{total:,.2f}" if total > 0 else "0.00"
                co2_formatted = f"{co2:,.2f}" if co2 > 0 else "0.00"

                scope2_items.append({
                    'number': '2.2',
                    'emission_source': '外购电力（基于市场）',
                    'total_green_house_gas_emissions': total_formatted,
                    'CO2_emissions': co2_formatted,
                    'CH4_emissions': '0.00',
                    'N2O_emissions': '0.00',
                    'HFCs_emissions': '0.00',
                    'PFCs_emissions': '0.00',
                    'SFs_emissions': '0.00',
                    'NF3_emissions': '0.00'
                })
                print(f"  提取2.2 外购电力（基于市场）: {total_formatted} tCO2e")

            result['scope2_items'] = scope2_items
            print(f"[范围二] 提取到范围二排放明细: {len(scope2_items)} 行")

        return result

    def _update_flags(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """更新 Flags 标记系统"""
        if 'flags' not in data:
            data['flags'] = {}

        def safe_float(value):
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0.0

        data['flags']['has_scope_1'] = safe_float(data.get('scope_1_emissions', 0)) > 0
        data['flags']['has_scope_2_location'] = safe_float(data.get('scope_2_location_based_emissions', 0)) > 0
        data['flags']['has_scope_2_market'] = safe_float(data.get('scope_2_market_based_emissions', 0)) > 0
        data['flags']['has_scope_3'] = safe_float(data.get('scope_3_emissions', 0)) > 0

        for i in range(1, 16):
            key = f'scope_3_category_{i}_emissions'
            flag_key = f'has_scope_3_category_{i}'
            data['flags'][flag_key] = safe_float(data.get(key, 0)) > 0

        return data

    def _extract_basic_info(self) -> Dict[str, Any]:
        """提取基本信息"""
        result = {
            'company_name': None,
            'company_profile': None,
            'legal_person': None,
            'registered_address': None,
            'date_of_establishment': None,
            'registered_capital': None,
            'Unified_Social_Credit_Identifier': None,
            'deadline': None,
            'evaluation_level': None,
            'evaluation_score': None,
            'scope_of_business': None,
            'rule_file': None,
            'GWP_Value_Reference_Document': None,
            'document_number': None,
            'posted_time': None,
            'production_address': None,
            'reporting_period': None,
            'report_year': '2024',
        }

        # 查找基本信息表
        basic_info_sheet = None
        for sheet in self.workbook.worksheets:
            if '基本信息' in sheet.title:
                basic_info_sheet = sheet
                break

        if basic_info_sheet:
            print(f"[基本信息] 找到基本信息表: {basic_info_sheet.title}")
            # 读取基本信息：第2列是属性代码(key)，第3列是值(value)
            for row in basic_info_sheet.iter_rows(min_row=2, max_row=50, values_only=True):
                if len(row) >= 3 and row[1] and row[2]:
                    key = str(row[1]).strip()  # 第2列：属性代码
                    value = row[2]  # 第3列：值

                    # 映射到标准字段
                    if 'company_name' in key or '组织名称' in key:
                        result['company_name'] = str(value).strip() if value else None
                    elif key == 'company_profile':
                        if isinstance(value, str):
                            value = re.sub(r'[\n\r]+', ' ', value)
                            value = re.sub(r'\s+', ' ', value).strip()
                        result['company_profile'] = value
                    elif key == 'scope_of_business':
                        if isinstance(value, str):
                            value = re.sub(r'[\n\r]+', ' ', value)
                            value = re.sub(r'\s+', ' ', value).strip()
                        result['scope_of_business'] = value
                    elif key == 'legal_person':
                        result['legal_person'] = str(value).strip() if value else None
                    elif key == 'registered_address':
                        result['registered_address'] = str(value).strip() if value else None
                    elif key == 'production_address':
                        result['production_address'] = str(value).strip() if value else None
                    elif key == 'date_of_establishment':
                        # 处理日期字段
                        result['date_of_establishment'] = excel_date_to_string(value)
                    elif key == 'posted_time':
                        # 处理发布日期
                        result['posted_time'] = excel_date_to_string(value)
                    elif key == 'deadline':
                        # 处理截止日期
                        result['deadline'] = excel_date_to_string(value)
                    elif key == 'registered_capital':
                        result['registered_capital'] = str(value).strip() if value else None
                    elif key == 'Unified_Social_Credit_Identifier':
                        result['Unified_Social_Credit_Identifier'] = str(value).strip() if value else None
                    elif key == 'reporting_period':
                        result['reporting_period'] = str(value).strip() if value else None
                        # 从周期中提取年份
                        year_match = re.search(r'(\d{4})', str(value))
                        result['report_year'] = year_match.group(1) if year_match else '2024'
                    elif key == 'document_number':
                        result['document_number'] = str(value).strip() if value else None
                    elif key in result:
                        result[key] = value

            print(f"[基本信息] 公司名称: {result.get('company_name')}")
        else:
            print("[基本信息] 未找到基本信息表，尝试从温室气体盘查清册提取...")
            # 尝试从温室气体盘查清册提取
            for sheet in self.workbook.worksheets:
                if '温室气体盘查清册' in sheet.title or '清册' in sheet.title:
                    for row in sheet.iter_rows(max_row=20, values_only=True):
                        if len(row) >= 3:
                            if row[1] == '组织名称：' and row[2]:
                                result['company_name'] = row[2]
                            elif row[1] == '组织地址：' and row[2]:
                                result['registered_address'] = row[2]
                                result['production_address'] = row[2]
                            elif row[1] == '盘查覆盖周期:' and row[2]:
                                result['reporting_period'] = row[2]
                                year_match = re.search(r'(\d{4})', str(row[2]))
                                result['report_year'] = year_match.group(1) if year_match else '2024'
                    break

        return result

    def _extract_scope3_categories(self) -> Dict[str, Any]:
        """提取范围三类别排放数据"""
        result = {}
        for i in range(1, 16):
            result[f'scope_3_category_{i}_emissions'] = 0.0

        # 添加范围三类别名称映射（模板需要）
        result['scope_3_category_names'] = {
            1: '外购商品和服务上游排放',
            2: '资本货物上游排放',
            3: '燃料和能源相关活动未包含在范围一和范围二中的上游排放',
            4: '上游运输和配送',
            5: '运营中产生的废弃物',
            6: '员工商务旅行',
            7: '员工通勤',
            8: '上游租赁资产',
            9: '下游运输和配送',
            10: '销售产品的加工',
            11: '销售产品的使用',
            12: '售出产品的加工',
            13: '下游租赁资产',
            14: '特许经营',
            15: '投资'
        }

        # 从"表1温室气体盘查表"中提取范围三各类别数据
        table_sheet = None
        for sheet in self.workbook.worksheets:
            if '表1' in sheet.title and '温室气体盘查表' in sheet.title:
                table_sheet = sheet
                break

        if table_sheet:
            scope3_mapping = {
                '类别1': 'scope_3_category_1_emissions',
                '类别2': 'scope_3_category_2_emissions',
                '类别3': 'scope_3_category_3_emissions',
                '类别4': 'scope_3_category_4_emissions',
                '类别5': 'scope_3_category_5_emissions',
                '类别6': 'scope_3_category_6_emissions',
                '类别7': 'scope_3_category_7_emissions',
                '类别8': 'scope_3_category_8_emissions',
                '类别9': 'scope_3_category_9_emissions',
                '类别10': 'scope_3_category_10_emissions',
                '类别11': 'scope_3_category_11_emissions',
                '类别12': 'scope_3_category_12_emissions',
                '类别13': 'scope_3_category_13_emissions',
                '类别14': 'scope_3_category_14_emissions',
                '类别15': 'scope_3_category_15_emissions',
            }

            for category_key, var_name in scope3_mapping.items():
                for row in table_sheet.iter_rows():
                    a_val = row[0].value if len(row) > 0 else None
                    if a_val and isinstance(a_val, str) and '范围三' in a_val and category_key in a_val:
                        current_row_num = row[0].row
                        if current_row_num + 2 <= table_sheet.max_row:
                            emission_row = table_sheet[current_row_num + 2]
                            b_val = emission_row[1].value if len(emission_row) > 1 else None
                            if b_val and isinstance(b_val, (int, float)):
                                result[var_name] = float(b_val)
                        break

        return result

    def _extract_scope2_data(self) -> Dict[str, Any]:
        """提取范围一、二、三排放数据"""
        result = {
            'scope_1_emissions': None,
            'scope_2_location_based_emissions': None,
            'scope_2_market_based_emissions': None,
            'scope_3_emissions': None,
            'scope_2_location': None,
            'scope_2_market': None,
            'total_emission_location': None,
            'total_emission_market': None,
        }

        # 查找表1温室气体盘查表
        table_sheet = None
        for sheet in self.workbook.worksheets:
            if '表1' in sheet.title or '温室气体盘查表' in sheet.title:
                table_sheet = sheet
                break

        if table_sheet:
            # 动态查找总排放量汇总行
            for row in table_sheet.iter_rows(values_only=True):
                a_val = row[0] if len(row) > 0 else None
                b_val = row[1] if len(row) > 1 else None
                c_val = row[2] if len(row) > 2 else None
                d_val = row[3] if len(row) > 3 else None
                e_val = row[4] if len(row) > 4 else None

                if a_val and isinstance(a_val, str) and '排放量' in a_val:
                    if isinstance(b_val, (int, float)) and isinstance(c_val, (int, float)) and isinstance(d_val, (int, float)):
                        result['scope_1_emissions'] = float(b_val)
                        result['scope_2_location_based_emissions'] = float(c_val)
                        result['scope_3_emissions'] = float(d_val)
                        result['scope_2_location'] = float(c_val)
                        if isinstance(e_val, (int, float)):
                            result['total_emission_location'] = float(e_val)
                    break

            # 动态查找范围二基于市场的排放量
            for row in table_sheet.iter_rows():
                e_val = row[4].value if len(row) > 4 else None
                c_val = row[2].value if len(row) > 2 else None
                if e_val and isinstance(e_val, str) and '基于市场' in e_val:
                    if c_val and isinstance(c_val, (int, float)):
                        result['scope_2_market_based_emissions'] = float(c_val)
                        result['scope_2_market'] = float(c_val)
                        print(f"[范围二] 找到基于市场排放量: {float(c_val)}")
                    break

            # 计算总排放量（基于市场）
            result['total_emission_market'] = (
                result.get('scope_1_emissions', 0) +
                result.get('scope_2_market_based_emissions', 0) +
                result.get('scope_3_emissions', 0)
            )

        return result

    def _extract_scope1_emissions_data_from_sheet1(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        从附表1-温室气体盘查表中提取范围一直接排放源数据

        这是一个特殊的数据提取方法，因为附表1的结构是固定的：
        - 前4行是标题
        - 数据从第5行开始
        - 排放量数据在特定的列位置（30-37列）
        """
        result = {
            'scope1_stationary_combustion_emissions_items': [],
            'scope1_mobile_combustion_emissions_items': [],
            'scope1_fugitive_emissions_items': [],
            'scope1_process_emissions_items': [],
        }

        # 查找附表1
        target_sheet = None
        for sheet in self.workbook.worksheets:
            if '附表1' in sheet.title or ('温室' in sheet.title and '盘查' in sheet.title and '1' in sheet.title):
                target_sheet = sheet
                break

        if not target_sheet:
            return result

        print(f"[范围一排放] 找到工作表: {target_sheet.title}")

        # 数据从第5行开始
        data_start_row = 5

        for row_idx in range(data_start_row, target_sheet.max_row + 1):
            try:
                row = target_sheet[row_idx]

                # 读取各列数据
                number = self._safe_str(row[0].value)
                category = self._safe_str(row[1].value)
                emission_source = self._safe_str(row[2].value)
                facility = self._safe_str(row[3].value)

                # 如果编号为空，跳过
                if not number or number.strip() == '':
                    continue

                # 读取排放量数据（Columns 30-37）
                co2_emissions = self._safe_float(row[30].value) if len(row) > 30 else 0
                ch4_emissions = self._safe_float(row[31].value) if len(row) > 31 else 0
                n2o_emissions = self._safe_float(row[32].value) if len(row) > 32 else 0
                hfcs_emissions = self._safe_float(row[33].value) if len(row) > 33 else 0
                pfcs_emissions = self._safe_float(row[34].value) if len(row) > 34 else 0
                sf6_emissions = self._safe_float(row[35].value) if len(row) > 35 else 0
                nf3_emissions = self._safe_float(row[36].value) if len(row) > 36 else 0
                total_emissions = self._safe_float(row[37].value) if len(row) > 37 else 0

                # 创建数据项
                item = {
                    'number': number,
                    'category': category,
                    'emission_source': emission_source,
                    'facility': facility,
                    'CO2_emissions': co2_emissions,
                    'CH4_emissions': ch4_emissions,
                    'N2O_emissions': n2o_emissions,
                    'HFCs_emissions': hfcs_emissions,
                    'PFCs_emissions': pfcs_emissions,
                    'SF6_emissions': sf6_emissions,
                    'NF3_emissions': nf3_emissions,
                    'total_green_house_gas_emissions': total_emissions,
                }

                # 根据类别分组
                if '固定燃烧' in category:
                    result['scope1_stationary_combustion_emissions_items'].append(item)
                elif '移动燃烧' in category or '移动汽油' in category or '移动柴油' in category:
                    result['scope1_mobile_combustion_emissions_items'].append(item)
                elif '逸散排放' in category or '逸散' in category:
                    result['scope1_fugitive_emissions_items'].append(item)
                elif '制程排放' in category or '制程' in category:
                    result['scope1_process_emissions_items'].append(item)

            except Exception as e:
                continue

        print(f"[范围一排放] 提取完成:")
        for group_name, items in result.items():
            print(f"  {group_name}: {len(items)} 条")

        return result

    def _safe_str(self, value) -> str:
        """安全地转换为字符串"""
        if value is None:
            return ''
        return str(value).strip()

    def _safe_float(self, value) -> float:
        """安全地转换为浮点数"""
        try:
            if value is None:
                return 0.0
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _extract_scope1_detail_from_inventory_sheet(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        从温室气体盘查清册表中提取范围一详细表数据

        参考原始代码 data_reader.py 的 extract_data_from_xlsx_dynamic 方法
        从"温室气体盘查清册表"中提取范围一的详细排放源数据
        """
        result = {
            'scope1_stationary_combustion_emissions_items': [],
            'scope1_mobile_combustion_emissions_items': [],
            'scope1_fugitive_emissions_items': [],
            'scope1_process_emissions_items': [],
        }

        # 查找温室气体盘查清册表
        inventory_sheet = None
        for sheet in self.workbook.worksheets:
            if '盘查清册' in sheet.title or '清册' in sheet.title:
                inventory_sheet = sheet
                break

        if not inventory_sheet:
            print("[范围一详细] 未找到温室气体盘查清册表")
            return result

        print(f"[范围一详细] 找到温室气体盘查清册表: {inventory_sheet.title}")
        scope1_detail_items = []

        # 从第14行开始（第12行是标题，第13行是单位）
        # 维护当前类别用于前向填充（ffill）
        current_category = ""
        current_sub_category = ""

        for row in inventory_sheet.iter_rows(min_row=14):
            if len(row) < 13:
                continue

            # Excel结构：A=空, B=编号/类别名, C=排放源, D=排放设施, E=备注, F=总排放量, G=CO2, H=CH4, I=N2O, J=HFCs, K=PFCs, L=SF6, M=NF3
            # 注意：当是类别行时，B列包含类别名（如"范围一 直接排放"），C列为空
            #       当是数据行时，B列包含编号（如1.1），C列包含排放源名称
            col_b = row[1].value        # 编号或类别名
            col_c = row[2].value        # 排放源（仅数据行有值）
            facility = row[3].value      # 排放设施 (列D)
            note = row[4].value          # 备注 (列E)
            total_emission = row[5].value    # 总排放量 (列F)
            co2_emission = row[6].value      # CO2排放量 (列G)
            ch4_emission = row[7].value      # CH4排放量 (列H)
            n2o_emission = row[8].value      # N2O排放量 (列I)
            hfcs_emission = row[9].value     # HFCs排放量 (列J)
            pfcs_emission = row[10].value    # PFCs排放量 (列K)
            sf6_emission = row[11].value     # SF6排放量 (列L)
            nf3_emission = row[12].value     # NF3排放量 (列M)

            # 跳过空行
            if not col_b and not col_c:
                continue

            # 确定编号和排放源
            number_str = ''
            source_str = ''

            if col_b:
                col_b_str = str(col_b).strip()
                # 检查B列是否是编号格式（如"1.1", "1.1.1"）- 以数字开头
                if col_b_str and col_b_str[0].isdigit():
                    number_str = col_b_str
                    source_str = str(col_c).strip() if col_c else ''
                # B列是类别名称（如"范围一 直接排放"）
                else:
                    # 更新当前类别（用于前向填充）
                    current_category = col_b_str
                    # 跳过类别标题行，但继续处理后续行
                    continue

            facility_str = str(facility).strip() if facility else ''
            note_str = str(note).strip() if note else ''

            # 跳过标题行
            if not number_str or number_str == '编号':
                continue

            # 格式化排放量数字（保留两位小数）
            def format_emission(val):
                if val is None:
                    return ''
                if val == 0:
                    return "0.00"
                try:
                    float_value = float(val)
                    if float_value == 0:
                        return "0.00"
                    return f"{float_value:.2f}"
                except (ValueError, TypeError):
                    return '0.00'

            # 范围一：编号以1开头（如1.1, 1.1.1）
            if number_str.startswith('1.'):
                # 确定子类别（根据编号前缀判断）
                if number_str.startswith('1.1.'):
                    current_sub_category = '固定源燃烧'
                elif number_str.startswith('1.2.'):
                    current_sub_category = '移动源燃烧'
                elif number_str.startswith('1.3.'):
                    current_sub_category = '遗散源'
                elif number_str.startswith('1.4.'):
                    current_sub_category = '工艺排放'

                # 使用编号作为主标识，但保留类别信息
                item = {
                    'name': current_sub_category or number_str,  # 使用子类别名称
                    'number': number_str,
                    'category': current_category,  # 添加类别字段
                    'emission_source': source_str,
                    'facility': facility_str,
                    'note': note_str,
                    'total_green_house_gas_emissions': format_emission(total_emission),
                    'CO2_emissions': format_emission(co2_emission),
                    'CH4_emissions': format_emission(ch4_emission),
                    'N2O_emissions': format_emission(n2o_emission),
                    'HFCs_emissions': format_emission(hfcs_emission),
                    'PFCs_emissions': format_emission(pfcs_emission),
                    'SFs_emissions': format_emission(sf6_emission),
                    'NF3_emissions': format_emission(nf3_emission)
                }
                scope1_detail_items.append(item)

        # 分类范围一数据
        for item in scope1_detail_items:
            number = item.get('number', '')
            if number.startswith('1.1.'):
                result['scope1_stationary_combustion_emissions_items'].append(item)
            elif number.startswith('1.2.'):
                result['scope1_mobile_combustion_emissions_items'].append(item)
            elif number.startswith('1.3.'):
                result['scope1_fugitive_emissions_items'].append(item)
            elif number.startswith('1.4.'):
                result['scope1_process_emissions_items'].append(item)

        print(f"[范围一详细] 提取范围一详细表数据: {len(scope1_detail_items)} 行")
        print(f"  固定源燃烧: {len(result['scope1_stationary_combustion_emissions_items'])} 行")
        print(f"  移动源燃烧: {len(result['scope1_mobile_combustion_emissions_items'])} 行")
        print(f"  逸散源: {len(result['scope1_fugitive_emissions_items'])} 行")
        print(f"  工艺排放: {len(result['scope1_process_emissions_items'])} 行")

        # ========== 计算各类别的汇总值 ==========
        emission_columns = [
            'total_green_house_gas_emissions',
            'CO2_emissions',
            'CH4_emissions',
            'N2O_emissions',
            'HFCs_emissions',
            'PFCs_emissions',
            'SFs_emissions',
            'NF3_emissions'
        ]

        # 辅助函数：计算列表中某列的总和
        def calculate_column_sum(items, column_name):
            total = 0.0
            for item in items:
                emission_str = item.get(column_name, '0')
                if emission_str and emission_str.strip():
                    try:
                        emission_str = emission_str.replace(',', '').replace(' ', '')
                        total += float(emission_str)
                    except (ValueError, TypeError):
                        pass
            return total

        # 计算各分类的汇总值
        # 固定源燃烧汇总
        for col in emission_columns:
            total = calculate_column_sum(result['scope1_stationary_combustion_emissions_items'], col)
            result[f'scope1_stationary_combustion_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        # 移动源燃烧汇总
        for col in emission_columns:
            total = calculate_column_sum(result['scope1_mobile_combustion_emissions_items'], col)
            result[f'scope1_mobile_combustion_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        # 逸散源汇总
        for col in emission_columns:
            total = calculate_column_sum(result['scope1_fugitive_emissions_items'], col)
            result[f'scope1_fugitive_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        # 工艺排放汇总
        for col in emission_columns:
            total = calculate_column_sum(result['scope1_process_emissions_items'], col)
            result[f'scope1_process_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        # 计算范围一的总排放量（各分类汇总之和）
        for col in emission_columns:
            stationary_total = float(result[f'scope1_stationary_combustion_emissions_{col}_sum_formatted'])
            mobile_total = float(result[f'scope1_mobile_combustion_emissions_{col}_sum_formatted'])
            fugitive_total = float(result[f'scope1_fugitive_emissions_{col}_sum_formatted'])
            process_total = float(result[f'scope1_process_emissions_{col}_sum_formatted'])
            total = stationary_total + mobile_total + fugitive_total + process_total
            result[f'scope1_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        print(f"[范围一详细] 计算汇总值完成:")
        print(f"  总排放量: {result.get('scope1_emissions_total_green_house_gas_emissions_sum_formatted', 'N/A')} tCO2e")
        print(f"  CO2排放量: {result.get('scope1_emissions_CO2_emissions_sum_formatted', 'N/A')} tCO2e")

        return result

    def get_protocol_data(self, protocol_name: str) -> List[Dict[str, Any]]:
        """
        获取特定协议的数据

        Args:
            protocol_name: 协议名称

        Returns:
            提取的数据列表
        """
        # 查找匹配该协议的工作表
        for sheet in self.workbook.worksheets:
            if self.fingerprint.identify(sheet, sheet.title) == protocol_name:
                return self.extractor.extract_from_sheet(sheet, protocol_name)
        return []

    def _extract_emission_factor_subtables(self, sheet: openpyxl.worksheet.Worksheet) -> List[Dict[str, Any]]:
        """
        从排放因子表中提取所有子表的数据

        附表2-EF 包含多个子表，每个子表以"编号"开始
        每个子表对应不同的排放类别和不同的列结构

        Returns:
            包含所有子表数据的字典列表
        """
        print(f"[排放因子表] 开始识别子表...")

        # 第一步：找到所有"编号"出现的行（子表开始位置）
        subtable_starts = []
        for row_idx in range(1, sheet.max_row + 1):
            for cell in sheet[row_idx]:
                if cell.value and str(cell.value).strip() == '编号':
                    subtable_starts.append(row_idx)
                    break

        print(f"[排放因子表] 找到 {len(subtable_starts)} 个子表")

        all_data = []

        # 第二步：处理每个子表
        for i, start_row in enumerate(subtable_starts):
            end_row = subtable_starts[i + 1] if i + 1 < len(subtable_starts) else sheet.max_row + 1

            # 读取子表的前几行来识别表头结构
            subtable_data = self._extract_single_ef_subtable(sheet, start_row, end_row)

            if subtable_data:
                all_data.extend(subtable_data)
                print(f"[排放因子表] 子表 {i + 1}: 提取到 {len(subtable_data)} 行数据")

        print(f"[排放因子表] 总共提取到 {len(all_data)} 行数据")
        return all_data

    def _extract_single_ef_subtable(self, sheet: openpyxl.worksheet.Worksheet,
                                   start_row: int, end_row: int) -> List[Dict[str, Any]]:
        """
        从单个排放因子子表中提取数据

        Args:
            sheet: 工作表对象
            start_row: 子表开始行
            end_row: 子表结束行

        Returns:
            该子表的数据列表
        """
        # 读取前5行来识别表头结构
        header_rows = []
        for row_idx in range(start_row, min(start_row + 5, end_row)):
            row = sheet[row_idx]
            row_data = []
            for cell in row:
                value = cell.value if cell.value is not None else ''
                row_data.append(str(value).strip())
            header_rows.append(row_data)

        # 查找第一行数据以获取类别信息
        first_data_row = None
        for row_idx in range(start_row + 2, min(start_row + 10, end_row)):
            row = sheet[row_idx]
            # 检查Col2是否有数字编号
            if len(row) > 1:
                col2_value = row[1].value if row[1].value is not None else ''
                try:
                    float(col2_value)  # 如果是数字，说明这是数据行
                    first_data_row = row
                    break
                except (ValueError, TypeError):
                    continue

        # 如果找到数据行，将其信息也添加到header_rows用于识别
        if first_data_row:
            data_row_data = []
            for cell in first_data_row[:8]:
                value = cell.value if cell.value is not None else ''
                data_row_data.append(str(value).strip())
            header_rows.append(data_row_data)

        # 识别子表类型（通过分析表头结构和数据行）
        subtable_type = self._identify_ef_subtable_type(header_rows)

        # 根据子表类型提取数据
        return self._extract_ef_subtable_data_by_type(sheet, start_row, end_row, subtable_type)

    def _identify_ef_subtable_type(self, header_rows: List[List[str]]) -> str:
        """
        根据表头行识别子表类型

        支持显式类别标识：
        - "范围一 XXX" -> scope1_combustion/scope1_process/scope1_fugitive
        - "范围二 XXX" -> scope2
        - "范围三 类别N XXX" -> scope3_catN (N为1-15)

        Returns:
            子表类型
        """
        # 合并所有表头行进行分析
        all_text = ' '.join([' '.join(row) for row in header_rows])

        # ========== 优先级1：识别范围三 "范围三 类别N" (简单字符串匹配) ==========
        # 支持格式：范围三 类别N, 范围三类别N, 范围3 类别N
        # 改为从大到小匹配，避免"类别1"匹配到"类别12"
        for cat_num in range(15, 0, -1):
            if f'范围三 类别{cat_num}' in all_text or f'范围三类别{cat_num}' in all_text or f'范围3 类别{cat_num}' in all_text:
                return f'scope3_cat{cat_num}'

        # ========== 优先级2：识别范围二 "范围二" ==========
        if '范围二' in all_text or '范围2' in all_text:
            return 'scope2'

        # ========== 优先级3：识别范围一排放因子类型 ==========
        if '范围一' in all_text or '范围1' in all_text:
            # 进一步细分范围一的类型
            if '低位发热量' in all_text and '氧化率' in all_text:
                return 'combustion'
            elif '制程排放' in all_text or '工艺排放' in all_text:
                return 'process'
            elif '逸散排放' in all_text or 'HFCs/PFCs' in all_text:
                return 'fugitive'
            return 'scope1_combustion'  # 默认为燃烧

        # ========== 优先级4：旧格式关键词匹配（向后兼容）==========
        if '低位发热量' in all_text and '氧化率' in all_text:
            return 'combustion'
        elif 'CO2排放因子' in all_text and '制程排放' in all_text:
            return 'process'
        elif 'HFCs/PFCs' in all_text or 'MCF' in all_text or 'Bo' in all_text:
            return 'fugitive'
        elif '外购能源间接排放' in all_text:
            return 'scope2'

        # ========== 优先级5：范围三类别关键词匹配（旧格式） ==========
        scope3_keywords = {
            '外购商品和服务': 'scope3_cat1',
            '铁矿石': 'scope3_cat1',
            '资本货物': 'scope3_cat2',
            '燃料和能源相关': 'scope3_cat3',
            '上下游运输配送': 'scope3_cat4',
            '运营中产生的废物': 'scope3_cat5',
            '商务旅行': 'scope3_cat6',
            '员工通勤': 'scope3_cat7',
            '上游租赁资产': 'scope3_cat8',
            '下游运输配送': 'scope3_cat9',
            '外销产品加工': 'scope3_cat10',
            '外销产品使用': 'scope3_cat11',
            '外售产品报废': 'scope3_cat12',
        }

        for keyword, subtable_type in scope3_keywords.items():
            if keyword in all_text:
                return subtable_type

        return 'unknown'

    def _extract_ef_subtable_data_by_type(self, sheet: openpyxl.worksheet.Worksheet,
                                         start_row: int, end_row: int,
                                         subtable_type: str) -> List[Dict[str, Any]]:
        """
        根据子表类型提取数据

        Args:
            sheet: 工作表对象
            start_row: 子表开始行
            end_row: 子表结束行
            subtable_type: 子表类型

        Returns:
            该子表的数据列表
        """
        data_items = []

        # 检查子表是否有燃烧表格式（用于判断某些范围三类别使用燃烧表结构）
        has_combustion_format = False
        for row_idx in range(start_row, min(start_row + 5, end_row)):
            row = sheet[row_idx]
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '低位发热量' in cell.value:
                    has_combustion_format = True
                    break
            if has_combustion_format:
                break

        # 动态确定数据开始行：跳过表头行，找到第一个有编号的数据行
        # 通常表头只有1-2行，从第2行开始检查
        data_start_row = start_row + 1  # 从第2行开始检查

        # 找到第一个包含数字编号的行作为数据开始行
        for row_idx in range(start_row + 1, min(start_row + 6, end_row)):
            row = sheet[row_idx]
            # 检查Col2是否有数字编号
            if len(row) > 1:
                col2_value = row[1].value if row[1].value is not None else ''
                try:
                    # 尝试转换为数字
                    float(col2_value)
                    data_start_row = row_idx
                    break
                except (ValueError, TypeError):
                    continue

        for row_idx in range(data_start_row, end_row):
            row = sheet[row_idx]

            # 检查是否是空行
            if not any(cell.value is not None for cell in row):
                continue

            # 根据子表类型提取数据
            if subtable_type == 'combustion':
                item = self._extract_combustion_ef_row(row)
            elif subtable_type == 'process':
                item = self._extract_process_ef_row(row)
            elif subtable_type == 'fugitive':
                item = self._extract_fugitive_ef_row(row)
            elif subtable_type == 'scope2':
                item = self._extract_scope2_ef_row(row)
            elif subtable_type.startswith('scope3_cat'):
                # 处理"范围三 类别N"格式
                # 特殊处理：某些类别（如类别11）使用燃烧表格式
                if has_combustion_format:
                    # 使用燃烧表提取方式
                    item = self._extract_combustion_ef_row(row)
                else:
                    # 使用标准范围三提取方式
                    item = self._extract_scope3_ef_row(row, subtable_type)
            elif subtable_type in ['scope3_general', 'scope3_capital', 'scope3_fuel',
                                   'scope3_transport', 'scope3_waste', 'scope3_business',
                                   'scope3_commuting', 'scope3_processing', 'scope3_disposal']:
                item = self._extract_scope3_ef_row(row, subtable_type)
            else:
                continue

            if item and item.get('category'):  # 确保有类别信息
                data_items.append(item)

        return data_items

    def _extract_combustion_ef_row(self, row) -> Dict[str, Any]:
        """提取燃烧排放因子行数据"""
        try:
            number = self._safe_get_cell(row, 1)  # Col2 (0-based index 1)
            category = self._safe_get_cell(row, 2)  # Col3
            emission_source = self._safe_get_cell(row, 3)  # Col4
            facility = self._safe_get_cell(row, 4)  # Col5
            ncv = self._safe_float(self._safe_get_cell(row, 5))  # Col6
            unit = self._safe_get_cell(row, 6)  # Col7
            ox_rate = self._safe_float(self._safe_get_cell(row, 7))  # Col8
            co2_cv = self._safe_float(self._safe_get_cell(row, 8))  # Col9: 基于热值 CO2
            ch4_cv = self._safe_float(self._safe_get_cell(row, 9))  # Col10: 基于热值 CH4
            n2o_cv = self._safe_float(self._safe_get_cell(row, 10))  # Col11: 基于热值 N2O
            co2_ef = self._safe_float(self._safe_get_cell(row, 11))  # Col12: 排放因子 CO2
            ch4_ef = self._safe_float(self._safe_get_cell(row, 12))  # Col13: 排放因子 CH4
            n2o_ef = self._safe_float(self._safe_get_cell(row, 13))  # Col14: 排放因子 N2O

            return {
                'number': number,
                'category': category,
                'emission_source': emission_source,
                'facility': facility,
                'ncv': ncv,
                'unit': unit,
                'ox_rate': ox_rate,
                'CO2_emission_cv_factor': co2_cv,
                'CH4_emission_cv_factor': ch4_cv,
                'N2O_emission_cv_factor': n2o_cv,
                'CO2_emission_factor': co2_ef,
                'CH4_emission_factor': ch4_ef,
                'N2O_emission_factor': n2o_ef,
            }
        except Exception as e:
            return {}

    def _extract_process_ef_row(self, row) -> Dict[str, Any]:
        """提取制程排放因子行数据"""
        try:
            number = self._safe_get_cell(row, 1)  # Col2
            category = self._safe_get_cell(row, 2)  # Col3
            emission_source = self._safe_get_cell(row, 3)  # Col4
            facility = self._safe_get_cell(row, 4)  # Col5
            co2_ef = self._safe_float(self._safe_get_cell(row, 5))  # Col6: CO2排放因子
            unit = self._safe_get_cell(row, 6)  # Col7

            return {
                'number': number,
                'category': category,
                'emission_source': emission_source,
                'facility': facility,
                'ncv': 0,
                'unit': unit,
                'ox_rate': 0,
                'CO2_emission_cv_factor': 0,
                'CH4_emission_cv_factor': 0,
                'N2O_emission_cv_factor': 0,
                'CO2_emission_factor': co2_ef,
                'CH4_emission_factor': 0,
                'N2O_emission_factor': 0,
            }
        except Exception as e:
            return {}

    def _extract_fugitive_ef_row(self, row) -> Dict[str, Any]:
        """提取逸散排放因子行数据"""
        try:
            number = self._safe_get_cell(row, 1)  # Col2
            category = self._safe_get_cell(row, 2)  # Col3
            emission_source = self._safe_get_cell(row, 3)  # Col4
            facility = self._safe_get_cell(row, 4)  # Col5

            # 逸散排放表的特殊字段结构
            # Col6: HFCs/PFCs (EF估计值)
            # Col7: 单位
            # Col8: MCF
            # Col9: Bo
            # Col10: 排放因子
            # Col11: 单位
            # Col12: 排放因子引用源

            hfcs_pfcs = self._safe_float(self._safe_get_cell(row, 5))  # Col6
            unit1 = self._safe_get_cell(row, 6)  # Col7
            mcf = self._safe_float(self._safe_get_cell(row, 7))  # Col8
            bo = self._safe_float(self._safe_get_cell(row, 8))  # Col9
            ef_value = self._safe_float(self._safe_get_cell(row, 9))  # Col10
            unit2 = self._safe_get_cell(row, 10)  # Col11
            source = self._safe_get_cell(row, 11)  # Col12

            # 如果CH4逸散，需要使用MCF和Bo计算
            if 'CH4逸散' in category and ef_value == 0:
                if mcf > 0 and bo > 0:
                    ef_value = mcf * bo

            return {
                'number': number,
                'category': category,
                'emission_source': emission_source,
                'facility': facility,
                'ncv': 0,
                'unit': unit2 or unit1,
                'ox_rate': 0,
                'CO2_emission_cv_factor': 0,
                'CH4_emission_cv_factor': 0,
                'N2O_emission_cv_factor': 0,
                'CO2_emission_factor': ef_value,
                'CH4_emission_factor': 0,
                'N2O_emission_factor': 0,
                # 逸散排放特有字段
                'HFCs_PCFs_emission_factor': hfcs_pfcs,
                'MCF': mcf,
                'Bo': bo,
                'emission_factor': ef_value,
                'emission_source_dir': '',
                'emission_unit_dir': unit2 or unit1,
            }
        except Exception as e:
            return {}

    def _extract_scope2_ef_row(self, row) -> Dict[str, Any]:
        """提取外购能源排放因子行数据"""
        try:
            number = self._safe_get_cell(row, 1)  # Col2
            category = self._safe_get_cell(row, 2)  # Col3
            emission_source = self._safe_get_cell(row, 3)  # Col4
            facility = self._safe_get_cell(row, 4)  # Col5
            ef_value = self._safe_float(self._safe_get_cell(row, 5))  # Col6
            unit = self._safe_get_cell(row, 6)  # Col7
            emission_source_reference = self._safe_get_cell(row, 7)  # Col8

            return {
                'number': number,
                'category': category,
                'emission_source': emission_source,
                'facility': facility,
                'ncv': 0,
                'unit': unit,
                'ox_rate': 0,
                'CO2_emission_cv_factor': 0,
                'CH4_emission_cv_factor': 0,
                'N2O_emission_cv_factor': 0,
                'CO2_emission_factor': ef_value,
                'CH4_emission_factor': 0,
                'N2O_emission_factor': 0,
                'emission_source_reference': emission_source_reference,
            }
        except Exception as e:
            return {}

    def _extract_scope3_ef_row(self, row, subtable_type: str) -> Dict[str, Any]:
        """
        提取范围三排放因子行数据

        支持的子表类型：
        - 'scope3_catN': N为类别编号（1-15）
        - 'scope3_general', 'scope3_capital', 等旧格式（向后兼容）
        """
        try:
            import re

            number = self._safe_get_cell(row, 1)  # Col2
            raw_category = self._safe_get_cell(row, 2)  # Col3 (原始类别，可能是排放源名称)
            emission_source = self._safe_get_cell(row, 3)  # Col4
            activity_name = self._safe_get_cell(row, 4)  # Col5: Activity name

            # 根据subtable_type确定正确的类别名称
            # 如果subtable_type是'scope3_catN'格式，提取类别编号并生成标准类别名称
            category = raw_category  # 默认使用原始类别
            if subtable_type.startswith('scope3_cat'):
                # 提取类别编号，例如 'scope3_cat9' -> 9
                match = re.search(r'scope3_cat(\d+)', subtable_type)
                if match:
                    cat_num = match.group(1)
                    category = f'范围三 类别{cat_num}'  # 设置标准类别名称
                    # 如果原始类别不为空且与编号不同，可能包含有用的描述信息
                    # 此时将原始类别作为排放源的补充信息
                    # 注意：raw_category可能是数字（float），需要先转换为字符串
                    raw_category_str = str(raw_category) if raw_category is not None else ''
                    if raw_category_str and raw_category_str.strip() and raw_category_str != category:
                        # 如果原始类别看起来像排放源名称（不是编号格式），则补充到emission_source
                        if not raw_category_str.replace('.', '').isdigit():
                            if not emission_source or not emission_source.strip():
                                emission_source = raw_category_str

            # 确定列结构：检查第5列（Col6）的值类型来判断
            # 如果Col5是Geography（文本且非空、不是数字或公式），使用标准格式
            # 如果Col5是空的或者是数字，使用资本货物格式
            col5_value = self._safe_get_cell(row, 5)

            # 判断是否有Geography列
            # 如果Col5非空且不能转换为数字（且不是公式），认为是Geography列
            # 注意：col5_value可能是数字，需要先转换为字符串再检查strip()
            col5_str = str(col5_value) if col5_value is not None and col5_value != '' else ''
            if col5_str and col5_str.strip():
                # 排除公式（以=开头）
                if col5_str.startswith('='):
                    has_geography = False  # 是公式，没有Geography列
                else:
                    try:
                        float(col5_value)
                        has_geography = False  # 是数字，没有Geography列
                    except (ValueError, TypeError):
                        has_geography = True   # 是文本，有Geography列
            else:
                has_geography = False  # Col5为空，没有Geography列

            if not has_geography:
                # 资本货物列结构: 编号, 类别, 排放源, Activity name, CO2, 单位, 引用源
                ef_value = self._safe_float(self._safe_get_cell(row, 5))  # Col6: CO2
                unit = self._safe_get_cell(row, 6)  # Col7: 单位
                emission_source_reference = self._safe_get_cell(row, 7)  # Col8: 引用源
                geography = ''
            else:
                # 其他类别列结构: 编号, 类别, 排放源, Activity name, Geography, CO2, 单位, 引用源
                geography = self._safe_get_cell(row, 5)  # Col6: Geography
                col7_value = self._safe_get_cell(row, 6)  # Col7: CO2
                try:
                    ef_value = float(col7_value)
                except:
                    ef_value = 0
                unit = self._safe_get_cell(row, 7)  # Col8: 单位
                emission_source_reference = self._safe_get_cell(row, 8)  # Col9: 引用源

            return {
                'number': number,
                'category': category,
                'emission_source': emission_source,
                'facility': activity_name,  # 保持向后兼容
                'activity_name': activity_name,
                'geography': geography,
                'ncv': 0,
                'unit': unit,
                'ox_rate': 0,
                'CO2_emission_cv_factor': 0,
                'CH4_emission_cv_factor': 0,
                'N2O_emission_cv_factor': 0,
                'CO2_emission_factor': ef_value,
                'CH4_emission_factor': 0,
                'N2O_emission_factor': 0,
                'emission_source_reference': emission_source_reference,
            }
        except Exception as e:
            return {}

    def _safe_get_cell(self, row, col_idx):
        """安全获取单元格值"""
        try:
            if col_idx < len(row):
                cell = row[col_idx]
                return cell.value if cell.value is not None else ''
            return ''
        except:
            return ''

    def _safe_float(self, value) -> float:
        """安全转换为浮点数"""
        try:
            if value is None or value == '':
                return 0.0
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()


# ==================== 使用示例 ====================

if __name__ == "__main__":
    # 创建读取器
    reader = ExcelDataReaderRefactored("test_data.xlsx")

    # 获取所有数据（一键获取）
    context = reader.get_all_context()

    # 打印结果
    for var_name, data in context.items():
        if data:
            print(f"{var_name}: {len(data)} 行")

    reader.close()
