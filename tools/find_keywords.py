import openpyxl

# 加载Excel文件
wb = openpyxl.load_workbook('test_data.xlsx', data_only=True)

# 检查的工作表
sheets_to_check = ['温室气体盘查清册', '表1温室气体盘查表']

# 查找可能包含所需信息的关键词
for sheet_name in sheets_to_check:
    print(f"\n在工作表 '{sheet_name}' 中搜索关键信息:")
    sheet = wb[sheet_name]
    
    # 只搜索前20行，通常表头信息在前面
    for row in sheet.iter_rows(max_row=20):
        for cell in row:
            if cell.value:
                # 将单元格值转换为字符串并检查
                cell_text = str(cell.value)
                # 检查是否包含公司、年份、排放等关键词
                if any(keyword in cell_text for keyword in ['公司', '名称', 'year', '年份', '排放', 'scope', '范围']):
                    print(f"找到匹配: '{cell_text}' 在单元格 {cell.coordinate}")
                    # 如果有关键词右侧的值，也打印出来
                    if cell.column < sheet.max_column:
                        next_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value:
                            print(f"  右侧单元格值: '{next_cell.value}'")

# 关闭工作簿
wb.close()