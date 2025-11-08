import openpyxl

# 打开Excel文件
workbook = openpyxl.load_workbook('test_data.xlsx', data_only=True)

# 检查两个可能的工作表
sheet_names = ['表1温室气体盘查表', '温室气体盘查清册']

for sheet_name in sheet_names:
    if sheet_name in workbook.sheetnames:
        print(f"\n检查工作表: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # 搜索包含'总排放量'的单元格
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == '总排放量':
                    print(f"找到'总排放量'在单元格: {cell.coordinate}")
                    # 检查右侧单元格（值应该在右侧）
                    right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    print(f"右侧单元格: {right_cell.coordinate}")
                    print(f"右侧单元格值: {right_cell.value}")
                    
                    # 检查B14单元格（用户提到的范围一直接排放位置）
                    if 'B14' in [c.coordinate for c in sheet[14]]:
                        b14_cell = sheet['B14']
                        f14_cell = sheet['F14']
                        print(f"B14单元格值: {b14_cell.value}")
                        print(f"F14单元格值: {f14_cell.value}")

print("\n调试完成")