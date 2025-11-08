import openpyxl

# 加载Excel文件
workbook = openpyxl.load_workbook('test_data.xlsx', data_only=True)
# 选择要检查的工作表
table_sheet = '表1温室气体盘查表'
sheet = workbook[table_sheet]

print(f"专门查找工作表 '{table_sheet}' 中的数值数据...\n")

# 1. 查找范围二市场法数据
print("1. 查找范围二市场法相关数据：")
print("-" * 60)
found_market_data = False
for row_idx in range(1, 50):  # 检查前50行
    row_values = []
    for col_idx, cell in enumerate(sheet[row_idx], 1):
        if cell.value is not None:
            # 检查是否包含市场法相关关键词
            cell_str = str(cell.value)
            if '市场' in cell_str or 'Market' in cell_str:
                print(f"在第{row_idx}行第{col_idx}列找到: {cell.value}")
                # 打印该行的所有数值
                numeric_values = []
                for c in sheet[row_idx]:
                    if c.value is not None and isinstance(c.value, (int, float)):
                        numeric_values.append(f"({c.column}) {c.value}")
                if numeric_values:
                    print(f"  该行的数值: {', '.join(numeric_values)}")
                found_market_data = True
                print("  --- ")

# 2. 查找范围三总量数据
print("\n2. 查找范围三总量相关数据：")
print("-" * 60)
found_scope3_data = False
for row_idx in range(175, 185):  # 检查第175-185行
    row_values = []
    for col_idx, cell in enumerate(sheet[row_idx], 1):
        if cell.value is not None:
            cell_str = str(cell.value)
            if '范围三' in cell_str or 'Scope 3' in cell_str or '总量' in cell_str:
                print(f"在第{row_idx}行第{col_idx}列找到: {cell.value}")
                # 检查下方单元格是否为数值
                if row_idx + 1 <= sheet.max_row:
                    below_cell = sheet.cell(row=row_idx+1, column=col_idx)
                    if below_cell.value is not None:
                        print(f"  下方单元格: {below_cell.value} (类型: {type(below_cell.value).__name__})")
                found_scope3_data = True
                print("  --- ")

# 3. 查找所有较大的数值（可能是排放量）
print("\n3. 查找较大的数值（可能是排放量）：")
print("-" * 60)
large_numbers = []
for row_idx in range(1, sheet.max_row + 1):
    for col_idx, cell in enumerate(sheet[row_idx], 1):
        if cell.value is not None and isinstance(cell.value, (int, float)):
            if cell.value > 10000:  # 查找大于10000的数值
                large_numbers.append((row_idx, col_idx, cell.value))

# 排序并显示前20个最大的数值
large_numbers.sort(key=lambda x: x[2], reverse=True)
for i, (row, col, value) in enumerate(large_numbers[:20], 1):
    # 显示该数值周围的上下文
    print(f"{i}. 第{row}行第{col}列: {value}")
    # 显示左侧单元格
    if col > 1:
        left_cell = sheet.cell(row=row, column=col-1).value
        print(f"   左侧: {left_cell}")
    # 显示上方单元格
    if row > 1:
        up_cell = sheet.cell(row=row-1, column=col).value
        print(f"   上方: {up_cell}")
    print("   --- ")

workbook.close()