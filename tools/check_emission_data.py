import openpyxl

# 加载Excel文件
workbook = openpyxl.load_workbook('test_data.xlsx', data_only=True)
# 选择要检查的工作表
table_sheet = '表1温室气体盘查表'
sheet = workbook[table_sheet]

print(f"检查工作表 '{table_sheet}' 中的内容...")
print("\n前30行中的所有非空单元格内容：")
print("-" * 80)

# 检查前30行的内容
for row_idx, row in enumerate(sheet.iter_rows(max_row=30), 1):
    row_values = []
    for col_idx, cell in enumerate(row, 1):
        if cell.value is not None:
            # 只显示非空单元格
            row_values.append(f"({col_idx}) {cell.value}")
    if row_values:
        print(f"行 {row_idx}: {', '.join(row_values)}")

print("\n-" * 80)
print("搜索包含'市场'、'范围三'、'总量'、'合计'等关键词的所有单元格：")
print("-" * 80)

# 搜索整个工作表中的关键词
keywords = ['市场', '范围三', '总量', '合计', 'Scope 3', 'Scope 2 Market']
for row_idx, row in enumerate(sheet.iter_rows(), 1):
    for col_idx, cell in enumerate(row, 1):
        if cell.value is not None:
            cell_str = str(cell.value)
            for keyword in keywords:
                if keyword in cell_str:
                    # 打印找到的关键词及其上下文
                    # 尝试获取右侧和下方单元格的值
                    right_cell = sheet.cell(row=row_idx, column=col_idx+1).value
                    down_cell = sheet.cell(row=row_idx+1, column=col_idx).value
                    print(f"在第{row_idx}行第{col_idx}列找到关键词 '{keyword}': {cell.value}")
                    if right_cell is not None:
                        print(f"  右侧单元格: {right_cell}")
                    if down_cell is not None:
                        print(f"  下方单元格: {down_cell}")
                    print("  --- ")

workbook.close()