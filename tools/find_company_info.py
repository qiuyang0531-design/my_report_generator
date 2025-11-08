import openpyxl

# 加载Excel文件
wb = openpyxl.load_workbook('test_data.xlsx', data_only=True)

# 检查的工作表
sheets_to_check = ['温室气体盘查清册']

print("查找公司名称和年份信息...")

for sheet_name in sheets_to_check:
    print(f"\n在工作表 '{sheet_name}' 中搜索:")
    sheet = wb[sheet_name]
    
    # 搜索前30行
    for row in sheet.iter_rows(max_row=30):
        for cell in row:
            if cell.value:
                cell_text = str(cell.value).strip()
                # 检查可能的公司信息
                if any(key in cell_text for key in ['公司', '单位', '组织', '企业', '名称', 'Name']):
                    print(f"找到潜在公司信息: '{cell_text}' 在单元格 {cell.coordinate}")
                    # 检查右侧、下方和斜下方的单元格
                    if cell.column < sheet.max_column:
                        right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                        if right_cell.value:
                            print(f"  右侧单元格: '{right_cell.value}'")
                    if cell.row < sheet.max_row:
                        below_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                        if below_cell.value:
                            print(f"  下方单元格: '{below_cell.value}'")
                        # 检查斜下方
                        if cell.column < sheet.max_column:
                            diag_cell = sheet.cell(row=cell.row + 1, column=cell.column + 1)
                            if diag_cell.value:
                                print(f"  斜下方单元格: '{diag_cell.value}'")
                
                # 检查可能的年份信息
                if any(key in cell_text for key in ['年份', '年度', '年', 'Year', 'year']):
                    print(f"找到潜在年份信息: '{cell_text}' 在单元格 {cell.coordinate}")
                    # 检查右侧单元格
                    if cell.column < sheet.max_column:
                        right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                        if right_cell.value:
                            print(f"  右侧单元格: '{right_cell.value}'")

# 专门检查总排放量
print("\n查找总排放量信息...")
sheet = wb['表1温室气体盘查表']
for row in sheet.iter_rows():
    for cell in row:
        if cell.value:
            cell_text = str(cell.value).strip()
            if any(key in cell_text for key in ['总排放', '合计', '总计', 'Total', 'total', 'sum']):
                print(f"找到潜在总排放量信息: '{cell_text}' 在单元格 {cell.coordinate}")
                if cell.column < sheet.max_column:
                    right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    if right_cell.value:
                        print(f"  右侧单元格: '{right_cell.value}'")

# 关闭工作簿
wb.close()