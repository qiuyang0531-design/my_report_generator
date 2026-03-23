"""
协议提取器模块
=====================

根据协议配置从工作表中提取数据。
"""

import openpyxl
from typing import Dict, List, Any

from .protocols import TABLE_PROTOCOLS
from .config import TableProtocol, FieldMapping
from .utils import safe_float


class ProtocolExtractor:
    """协议数据提取器"""

    def __init__(self):
        pass

    def extract_from_sheet(self, sheet: openpyxl.worksheet.Worksheet,
                          protocol_name: str) -> List[Dict[str, Any]]:
        """
        根据协议从工作表提取数据

        Args:
            sheet: openpyxl工作表对象
            protocol_name: 协议名称

        Returns:
            提取的数据列表
        """
        if protocol_name not in TABLE_PROTOCOLS:
            print(f"[数据提取] 未知协议: {protocol_name}")
            return []

        protocol = TABLE_PROTOCOLS[protocol_name]

        # 查找表头行
        header_row = self._find_header_row(sheet, protocol)
        if not header_row:
            print(f"[数据提取] 未找到表头行")
            return []

        # 获取列映射
        column_map = self._build_column_map(sheet, header_row, protocol)
        if not column_map:
            print(f"[数据提取] 无法创建列映射")
            return []

        # 提取数据
        data_items = self._extract_data_rows(sheet, header_row, column_map, protocol)

        # 应用前向填充
        if protocol.ffill_fields:
            data_items = self._apply_ffill(data_items, protocol.ffill_fields)

        # 应用后处理
        if protocol.post_process:
            processed = protocol.post_process(data_items)
            if processed is not None:
                data_items = processed

        print(f"[数据提取] 提取到 {len(data_items)} 行数据")
        return data_items

    def _find_header_row(self, sheet: openpyxl.worksheet.Worksheet,
                        protocol: TableProtocol, max_row: int = 20) -> int:
        """查找表头行"""
        all_keywords = protocol.required_keywords | protocol.optional_keywords

        for row_idx in range(1, min(max_row + 1, sheet.max_row + 1)):
            row_values = set()
            for cell in sheet[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    row_values.add(str(cell.value).strip())

            # 至少匹配2个关键词
            matched = len(all_keywords & row_values)
            if matched >= min(2, len(protocol.required_keywords)):
                return row_idx

        return None

    def _build_column_map(self, sheet: openpyxl.worksheet.Worksheet,
                         header_row: int,
                         protocol: TableProtocol) -> Dict[str, int]:
        """构建列名到列索引的映射"""
        column_map = {}
        header_cells = {}

        # 收集表头单元格（支持多行表头）
        rows_to_check = list(range(header_row, min(header_row + protocol.header_rows_to_check, sheet.max_row + 1)))
        for row_idx in rows_to_check:
            if row_idx > sheet.max_row:
                continue
            for cell in sheet[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    value = str(cell.value).strip()
                    if value and value not in header_cells:
                        header_cells[value] = cell.column - 1

        # 活动数据汇总表的排放量字段特殊处理：使用固定列位置
        is_activity_summary = protocol.name in ['活动数据汇总表', '活动数据汇总表（市场法）']
        # 排放因子表的特殊处理：使用固定列位置区分两组 CO2/CH4/N2O
        is_emission_factor = protocol.name == '排放因子表'

        if is_activity_summary:
            # 活动数据汇总表的排放量列固定位置（基于表格结构）
            emission_column_map = {
                'CO2_emissions': 29,
                'CH4_emissions': 30,
                'N2O_emissions': 31,
                'HFCs_emissions': 32,
                'PFCs_emissions': 33,
                'SF6_emissions': 34,
                'NF3_emissions': 35,
                'total_green_house_gas_emissions': 36,
            }
        elif is_emission_factor:
            # 排放因子表的固定列位置
            emission_column_map = {
                'CO2_emission_cv_factor': 8,
                'CH4_emission_cv_factor': 9,
                'N2O_emission_cv_factor': 10,
                'CO2_emission_factor': 11,
                'CH4_emission_factor': 12,
                'N2O_emission_factor': 13,
            }
        else:
            emission_column_map = {}

        # 为每个字段查找列
        for field_name, field_mapping in protocol.field_mappings.items():
            # 活动数据汇总表的排放量字段：使用固定列位置
            if is_activity_summary and field_name in emission_column_map:
                column_map[field_name] = emission_column_map[field_name]
            # 排放因子表的排放系数字段：使用固定列位置
            elif is_emission_factor and field_name in emission_column_map:
                column_map[field_name] = emission_column_map[field_name]
            else:
                # 常规匹配
                if field_mapping.keyword in header_cells:
                    column_map[field_name] = header_cells[field_mapping.keyword]
                else:
                    # 模糊匹配
                    for header_value in header_cells:
                        if field_mapping.keyword in header_value or header_value in field_mapping.keyword:
                            column_map[field_name] = header_cells[header_value]
                            break
                    # 尝试备选关键词
                    if field_name not in column_map:
                        for alt_keyword in field_mapping.alt_keywords:
                            if alt_keyword in header_cells:
                                column_map[field_name] = header_cells[alt_keyword]
                                break

        return column_map

    def _extract_data_rows(self, sheet: openpyxl.worksheet.Worksheet,
                          header_row: int,
                          column_map: Dict[str, int],
                          protocol: TableProtocol) -> List[Dict[str, Any]]:
        """提取数据行"""
        data_items = []

        # 活动数据汇总表的特殊处理：跳过子表头行
        is_activity_summary = protocol.name in ['活动数据汇总表', '活动数据汇总表（市场法）']

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row = sheet[row_idx]

            # 检查空行
            if not any(cell.value is not None for cell in row):
                continue

            # 活动数据汇总表：跳过序号列非数字的行（子表头行）
            if is_activity_summary and 'number' in protocol.field_mappings:
                number_col_idx = column_map.get('number', 0)
                if number_col_idx < len(row):
                    number_cell = row[number_col_idx]
                    # 如果序号列不是数字，跳过该行
                    if number_cell.value is None or not isinstance(number_cell.value, (int, float)):
                        try:
                            float(number_cell.value)
                        except (ValueError, TypeError):
                            continue

            # 提取字段值
            item = {}
            for field_name, field_mapping in protocol.field_mappings.items():
                if field_name in column_map:
                    col_idx = column_map[field_name]
                    cell = row[col_idx] if col_idx < len(row) else None

                    value = self._convert_cell_value(cell, field_mapping)
                    item[field_name] = value
                else:
                    item[field_name] = field_mapping.default

            # 检查是否有效数据行
            if self._is_valid_row(item, protocol):
                data_items.append(item)

        return data_items

    def _convert_cell_value(self, cell, field_mapping: FieldMapping) -> Any:
        """转换单元格值"""
        if cell is None or cell.value is None:
            return field_mapping.default

        try:
            if field_mapping.dtype == 'float':
                return float(cell.value)
            elif field_mapping.dtype == 'int':
                return int(cell.value)
            else:
                return str(cell.value).strip()
        except (ValueError, TypeError):
            return field_mapping.default

    def _is_valid_row(self, item: Dict[str, Any], protocol: TableProtocol) -> bool:
        """检查是否为有效数据行"""
        for field_name, field_mapping in protocol.field_mappings.items():
            value = item.get(field_name)
            if value and value != field_mapping.default:
                return True
        return False

    def _apply_ffill(self, data_items: List[Dict[str, Any]],
                    field_names: List[str]) -> List[Dict[str, Any]]:
        """应用前向填充"""
        if not data_items:
            return data_items

        result = []
        last_values = {field: None for field in field_names}

        for item in data_items:
            new_item = item.copy()
            for field in field_names:
                value = item.get(field)
                if value and str(value).strip():
                    last_values[field] = value
                    new_item[field] = value
                elif last_values[field]:
                    new_item[field] = last_values[field]
            result.append(new_item)

        return result


__all__ = ['ProtocolExtractor']
