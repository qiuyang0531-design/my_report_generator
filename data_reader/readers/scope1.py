"""
范围一数据读取模块
=====================

从Excel中提取范围一直接排放源数据。
"""

import openpyxl
from typing import Dict, List, Any

from .base import BaseReader
from ..post_processors import group_by_emission_category


class Scope1Reader(BaseReader):
    """范围一数据读取器"""

    def extract_all(self) -> Dict[str, Any]:
        """
        提取所有范围一相关数据

        Returns:
            包含范围一数据的字典
        """
        result = {}

        # 从附表1-温室气体盘查表提取范围一直接排放源数据
        emissions_data = self._extract_emissions_data_from_sheet1()
        result.update(emissions_data)

        # 从温室气体盘查清册表提取范围一详细数据
        detail_data = self._extract_detail_from_inventory_sheet()
        result.update(detail_data)

        return result

    def _extract_emissions_data_from_sheet1(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        从附表1-温室气体盘查表中提取范围一直接排放源数据

        这是一个特殊的数据提取方法，因为附表1的结构是固定的：
        - 前4行是标题
        - 数据从第5行开始
        - 排放量数据在特定的列位置（30-37列）
        """
        result = {
            'scope1_stationary_combustion_emissions_items': [],
            'scope1_mobile_combustion_emissions_items': [],
            'scope1_fugitive_emissions_items': [],
            'scope1_process_emissions_items': [],
        }

        # 查找附表1
        target_sheet = self.find_sheet_by_name('附表1', '温室', '盘查', '1')

        if not target_sheet:
            return result

        print(f"[范围一排放] 找到工作表: {target_sheet.title}")

        # 数据从第5行开始
        data_start_row = 5

        for row_idx in range(data_start_row, target_sheet.max_row + 1):
            try:
                row = target_sheet[row_idx]

                # 读取各列数据
                number = self.safe_str(row[0].value)
                category = self.safe_str(row[1].value)
                emission_source = self.safe_str(row[2].value)
                facility = self.safe_str(row[3].value)

                # 如果编号为空，跳过
                if not number or number.strip() == '':
                    continue

                # 读取排放量数据（Columns 30-37）
                co2_emissions = self.safe_float(row[30].value) if len(row) > 30 else 0
                ch4_emissions = self.safe_float(row[31].value) if len(row) > 31 else 0
                n2o_emissions = self.safe_float(row[32].value) if len(row) > 32 else 0
                hfcs_emissions = self.safe_float(row[33].value) if len(row) > 33 else 0
                pfcs_emissions = self.safe_float(row[34].value) if len(row) > 34 else 0
                sf6_emissions = self.safe_float(row[35].value) if len(row) > 35 else 0
                nf3_emissions = self.safe_float(row[36].value) if len(row) > 36 else 0
                total_emissions = self.safe_float(row[37].value) if len(row) > 37 else 0

                # 创建数据项
                item = {
                    'number': number,
                    'category': category,
                    'emission_source': emission_source,
                    'facility': facility,
                    'CO2_emissions': co2_emissions,
                    'CH4_emissions': ch4_emissions,
                    'N2O_emissions': n2o_emissions,
                    'HFCs_emissions': hfcs_emissions,
                    'PFCs_emissions': pfcs_emissions,
                    'SF6_emissions': sf6_emissions,
                    'NF3_emissions': nf3_emissions,
                    'total_green_house_gas_emissions': total_emissions,
                }

                # 根据类别分组
                if '固定燃烧' in category:
                    result['scope1_stationary_combustion_emissions_items'].append(item)
                elif '移动燃烧' in category or '移动汽油' in category or '移动柴油' in category:
                    result['scope1_mobile_combustion_emissions_items'].append(item)
                elif '逸散排放' in category or '逸散' in category:
                    result['scope1_fugitive_emissions_items'].append(item)
                elif '制程排放' in category or '制程' in category:
                    result['scope1_process_emissions_items'].append(item)

            except Exception:
                continue

        print(f"[范围一排放] 提取完成:")
        for group_name, items in result.items():
            print(f"  {group_name}: {len(items)} 条")

        return result

    def _extract_detail_from_inventory_sheet(self) -> Dict[str, Any]:
        """
        从温室气体盘查清册表中提取范围一详细表数据，确保排序与工作表一致
        """
        result = {
            'scope1_stationary_combustion_emissions_items': [],
            'scope1_mobile_combustion_emissions_items': [],
            'scope1_fugitive_emissions_items': [],
            'scope1_process_emissions_items': [],
        }

        # 查找所有可能的盘查清册表
        inventory_sheets = []
        for sheet_name in self.workbook.sheetnames:
            if '盘查清册' in sheet_name or '清册' in sheet_name:
                inventory_sheets.append(self.workbook[sheet_name])

        if not inventory_sheets:
            print("[范围一详细] 未找到任何温室气体盘查清册表")
            return result

        sheet_by_title = {s.title: s for s in inventory_sheets}
        preferred_sheet = sheet_by_title.get('温室气体盘查清册') or inventory_sheets[0]
        preferred_title = preferred_sheet.title

        preferred_number_order: List[str] = []
        current_category = ""
        for row in preferred_sheet.iter_rows(min_row=14):
            if len(row) < 13:
                continue

            col_b = row[1].value
            col_c = row[2].value
            if not col_b and not col_c:
                continue

            number_str = ""
            if col_b:
                col_b_str = str(col_b).strip()
                if col_b_str and col_b_str[0].isdigit():
                    number_str = col_b_str
                else:
                    current_category = col_b_str
                    continue

            if not number_str or number_str == '编号':
                continue

            if number_str.startswith('1.') and number_str not in preferred_number_order:
                preferred_number_order.append(number_str)

        def score_item(item: Dict[str, Any], has_error: bool) -> int:
            score = 0
            score += -1000 if has_error else 100
            if item.get('emission_source'):
                score += 20
            if item.get('facility'):
                score += 10
            total_val = self.safe_float(item.get('total_green_house_gas_emissions'))
            if total_val != 0:
                score += 5
            gas_keys = ['CO2_emissions', 'CH4_emissions', 'N2O_emissions', 'HFCs_emissions', 'PFCs_emissions', 'SFs_emissions', 'NF3_emissions']
            if any(self.safe_float(item.get(k)) != 0 for k in gas_keys):
                score += 3
            if item.get('category'):
                score += 1
            return score

        def is_effectively_blank_item(item: Dict[str, Any]) -> bool:
            if item.get('emission_source') or item.get('facility'):
                return False
            keys = ['total_green_house_gas_emissions', 'CO2_emissions', 'CH4_emissions', 'N2O_emissions', 'HFCs_emissions', 'PFCs_emissions', 'SFs_emissions', 'NF3_emissions']
            return all(self.safe_float(item.get(k)) == 0 for k in keys)

        data_pool: Dict[str, Dict[str, Any]] = {}
        data_meta: Dict[str, Dict[str, Any]] = {}

        for inventory_sheet in inventory_sheets:
            print(f"[范围一详细] 正在从 {inventory_sheet.title} 汇总数据...")
            
            # 从第14行开始（第12行是标题，第13行是单位）
            current_category = ""

            for row in inventory_sheet.iter_rows(min_row=14):
                if len(row) < 13:
                    continue

                col_b = row[1].value        # 编号或类别名
                col_c = row[2].value        # 排放源
                
                if not col_b and not col_c:
                    continue

                number_str = ''
                if col_b:
                    col_b_str = str(col_b).strip()
                    if col_b_str and col_b_str[0].isdigit():
                        number_str = col_b_str
                    else:
                        current_category = col_b_str
                        continue

                if not number_str or number_str == '编号':
                    continue

                # 提取数据项
                if number_str.startswith('1.'):
                    # 确定子类别
                    sub_cat = ""
                    if number_str.startswith('1.1.'): sub_cat = '固定源燃烧'
                    elif number_str.startswith('1.2.'): sub_cat = '移动源燃烧'
                    elif number_str.startswith('1.3.'): sub_cat = '遗散源'
                    elif number_str.startswith('1.4.'): sub_cat = '工艺排放'

                    item = {
                        'name': sub_cat or number_str,
                        'number': number_str,
                        'category': current_category,
                        'emission_source': self.safe_str(col_c),
                        'facility': self.safe_str(row[3].value),
                        'note': self.safe_str(row[4].value),
                        'total_green_house_gas_emissions': self.format_emission(row[5].value),
                        'CO2_emissions': self.format_emission(row[6].value),
                        'CH4_emissions': self.format_emission(row[7].value),
                        'N2O_emissions': self.format_emission(row[8].value),
                        'HFCs_emissions': self.format_emission(row[9].value),
                        'PFCs_emissions': self.format_emission(row[10].value),
                        'SFs_emissions': self.format_emission(row[11].value),
                        'NF3_emissions': self.format_emission(row[12].value)
                    }

                    has_error = self.is_error_value(col_c) or self.is_error_value(row[5].value)
                    new_score = score_item(item, has_error)

                    if number_str not in data_pool:
                        data_pool[number_str] = item
                        data_meta[number_str] = {
                            'score': new_score,
                            'has_error': has_error,
                            'sheet_title': inventory_sheet.title,
                        }
                    else:
                        meta = data_meta.get(number_str, {})
                        old_score = int(meta.get('score', -10**9))
                        old_title = str(meta.get('sheet_title', ''))
                        if new_score > old_score or (new_score == old_score and inventory_sheet.title == preferred_title and old_title != preferred_title):
                            data_pool[number_str] = item
                            data_meta[number_str] = {
                                'score': new_score,
                                'has_error': has_error,
                                'sheet_title': inventory_sheet.title,
                            }

        number_order = preferred_number_order if preferred_number_order else sorted(list(data_pool.keys()), key=self.natural_sort_key)

        # 3. 按顺序分配到结果
        for num in number_order:
            if num not in data_pool:
                continue
            item = data_pool[num]
            if is_effectively_blank_item(item):
                continue
            if num.startswith('1.1.'):
                result['scope1_stationary_combustion_emissions_items'].append(item)
            elif num.startswith('1.2.'):
                result['scope1_mobile_combustion_emissions_items'].append(item)
            elif num.startswith('1.3.'):
                result['scope1_fugitive_emissions_items'].append(item)
            elif num.startswith('1.4.'):
                result['scope1_process_emissions_items'].append(item)

        print(f"[范围一详细] 提取完成:")
        print(f"  固定源燃烧: {len(result['scope1_stationary_combustion_emissions_items'])} 行")
        print(f"  移动源燃烧: {len(result['scope1_mobile_combustion_emissions_items'])} 行")
        print(f"  逸散源: {len(result['scope1_fugitive_emissions_items'])} 行")
        print(f"  工艺排放: {len(result['scope1_process_emissions_items'])} 行")

        # 计算各类别的汇总值
        self._calculate_emission_sums(result)

        return result

    def _calculate_emission_sums(self, result: Dict[str, Any]):
        """计算各分类的排放量汇总值"""
        emission_columns = [
            'total_green_house_gas_emissions',
            'CO2_emissions',
            'CH4_emissions',
            'N2O_emissions',
            'HFCs_emissions',
            'PFCs_emissions',
            'SFs_emissions',
            'NF3_emissions'
        ]

        def calculate_column_sum(items, column_name):
            total = 0.0
            for item in items:
                emission_str = item.get(column_name, '0')
                if emission_str and emission_str.strip():
                    try:
                        emission_str = emission_str.replace(',', '').replace(' ', '')
                        total += float(emission_str)
                    except (ValueError, TypeError):
                        pass
            return total

        # 计算各分类的汇总值
        for col in emission_columns:
            total = calculate_column_sum(result['scope1_stationary_combustion_emissions_items'], col)
            result[f'scope1_stationary_combustion_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        for col in emission_columns:
            total = calculate_column_sum(result['scope1_mobile_combustion_emissions_items'], col)
            result[f'scope1_mobile_combustion_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        for col in emission_columns:
            total = calculate_column_sum(result['scope1_fugitive_emissions_items'], col)
            result[f'scope1_fugitive_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        for col in emission_columns:
            total = calculate_column_sum(result['scope1_process_emissions_items'], col)
            result[f'scope1_process_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        # 计算范围一的总排放量（各分类汇总之和）
        for col in emission_columns:
            stationary_total = float(result[f'scope1_stationary_combustion_emissions_{col}_sum_formatted'])
            mobile_total = float(result[f'scope1_mobile_combustion_emissions_{col}_sum_formatted'])
            fugitive_total = float(result[f'scope1_fugitive_emissions_{col}_sum_formatted'])
            process_total = float(result[f'scope1_process_emissions_{col}_sum_formatted'])
            total = stationary_total + mobile_total + fugitive_total + process_total
            result[f'scope1_emissions_{col}_sum_formatted'] = f"{total:.2f}"

        print(f"[范围一详细] 计算汇总值完成:")
        print(f"  总排放量: {result.get('scope1_emissions_total_green_house_gas_emissions_sum_formatted', 'N/A')} tCO2e")
        print(f"  CO2排放量: {result.get('scope1_emissions_CO2_emissions_sum_formatted', 'N/A')} tCO2e")


__all__ = ['Scope1Reader']
