import openpyxl
import csv
import os
import re
from report_config import ReportConfig


# ========== 表格协议配置 (TABLE_PROTOCOLS) ==========
# 基于特征指纹的表格类型定义
# 每个协议定义了：
# - keywords: 用于识别表格的关键词集合
# - required_keywords: 必须包含的关键词
# - field_mapping: 从Excel列到标准字段的映射
TABLE_PROTOCOLS = {
    'EmissionFactorProtocol': {
        'name': '排放因子表',
        'description': '包含低位发热量、氧化率、基于热值排放系数等信息的表格',
        'keywords': {'低位发热量', '氧化率', '基于热值排放系数', '排放因子', 'GHG排放类别', '计算值', '排放系数'},
        'required_keywords': {'低位发热量', '氧化率'},  # 必须包含这2个关键词（与活动数据汇总表区分）
        'field_mapping': {
            # 标准字段名 -> (Excel列关键词, 默认值, 数据类型)
            # 字段名与模板 Jinja2 变量一致: {{item.xxx}}
            'category': ('GHG排放类别', '', 'str'),              # {{item.category}}
            'number': ('编号', '', 'str'),                        # {{item.number}}
            'emission_source': ('排放源', '', 'str'),            # {{item.emission_source}}
            'facility': ('设施', '', 'str'),                      # {{item.facility}}
            'ncv': ('低位发热量', 0, 'float'),                   # {{item.ncv}} - 低位发热量
            'unit': ('单位', '', 'str'),                         # {{item.unit}} - 单位
            'ox_rate': ('氧化率', 0, 'float'),                   # {{item.ox_rate}} - 氧化率
            'ef_val': ('计算值', 0, 'float'),                    # {{item.ef_val}} - 排放系数/计算值
            # 备选匹配词
            'ef_val_alt': ('排放系数', 0, 'float'),              # 备选: 排放系数
            'ef_val_alt2': ('基于热值排放系数', 0, 'float'),     # 备选: 基于热值排放系数
            'CO2_emission_factor': ('CO2', 0, 'float'),          # {{item.CO2_emission_factor}}
            'CH4_emission_factor': ('CH4', 0, 'float'),          # {{item.CH4_emission_factor}}
            'N2O_emission_factor': ('N2O', 0, 'float'),          # {{item.N2O_emission_factor}}
        },
        'output_var': 'pro_ef_items'  # 新的模板变量名
    },

    'GWPProtocol': {
        'name': 'GWP值表',
        'description': '全球变暖潜势(GWP)值参考表',
        'keywords': {'GWP', 'GWP(HFCs)', 'GWP(PFCs)', '工业名称', '中文名称', '化学分子式'},
        'required_keywords': {'GWP'},
        'field_mapping': {
            'gas_name': ('工业名称', '', 'str'),
            'chinese_name': ('中文名称/化学分子式', '', 'str'),
            'formula': ('中文名称/化学分子式', '', 'str'),
            'composition_ratio': ('组成比例', None, 'float'),
            'gwp_value': ('GWP', 0, 'float'),
            'gwp_hfcs': ('GWP(HFCs)', None, 'float'),
            'gwp_pfcs': ('GWP(PFCs)', None, 'float'),
            'source': ('来源', '', 'str'),
            'note': ('备注', '', 'str'),
        },
        'output_var': 'gwp_items'
    },

    'GHGInventoryProtocol': {
        'name': '温室气体盘查表',
        'description': '温室气体排放盘查汇总表',
        'keywords': {'GHG排放类别', '排放源', '设施', 'GWP', 'EF', '活动数据', '排放量'},
        'required_keywords': {'GHG排放类别', '排放量'},
        'field_mapping': {
            'category': ('GHG排放类别', '', 'str'),
            'emission_source': ('排放源', '', 'str'),
            'facility': ('设施', '', 'str'),
            'activity_data': ('活动数据', 0, 'float'),
            'activity_data_unit': ('单位', '', 'str'),
            'emission_factor': ('EF', 0, 'float'),
            'gwp': ('GWP', 1, 'float'),
            'emissions': ('排放量', 0, 'float'),
            'emissions_unit': ('tCO2e', '', 'str'),
        },
        'output_var': 'ghg_inventory_items'
    },

    'ActivitySummaryProtocol': {
        'name': '活动数据汇总表',
        'description': '基于位置或市场的活动数据汇总表',
        'keywords': {'活动数据汇总', 'GHG', '基于位置', '基于市场', 'CO2', 'CH4', 'N2O'},
        'required_keywords': {'活动数据汇总'},
        'field_mapping': {
            'category': ('GHG排放类别', '', 'str'),
            'emission_source': ('排放源', '', 'str'),
            'facility': ('设施', '', 'str'),
            'activity_data': ('活动数据', 0, 'float'),
            'unit': ('单位', '', 'str'),
            'co2_emissions': ('CO2', 0, 'float'),
            'ch4_emissions': ('CH4', 0, 'float'),
            'n2o_emissions': ('N2O', 0, 'float'),
            'total_emissions': ('总计', 0, 'float'),
        },
        'output_var': 'activity_summary_items'
    },

    'UncertaintyProtocol': {
        'name': '不确定性评估表',
        'description': '基于位置或市场的不确定性评估表',
        'keywords': {'不确定性', '评估', '基于位置', '基于市场', '扩展不确定度'},
        'required_keywords': {'不确定性'},
        'field_mapping': {
            'category': ('GHG排放类别', '', 'str'),
            'emission_source': ('排放源', '', 'str'),
            'activity_data_uncertainty': ('活动数据', 0, 'float'),
            'emission_factor_uncertainty': ('排放因子', 0, 'float'),
            'combined_uncertainty': ('合成不确定度', 0, 'float'),
            'extended_uncertainty': ('扩展不确定度', 0, 'float'),
        },
        'output_var': 'uncertainty_items'
    },

    'ReductionActionProtocol': {
        'name': '减排行动统计表',
        'description': '减排行动统计信息表',
        'keywords': {'减排行动', '统计', '项目', '减排量'},
        'required_keywords': {'减排行动'},
        'field_mapping': {
            'project_name': ('项目名称', '', 'str'),
            'implementation_date': ('实施日期', '', 'str'),
            'reduction_type': ('减排类型', '', 'str'),
            'annual_reduction': ('年减排量', 0, 'float'),
            'reduction_unit': ('单位', '', 'str'),
            'description': ('描述', '', 'str'),
        },
        'output_var': 'reduction_action_items'
    },
}


class ExcelDataReader:
    def _safe_float(self, value):
        """安全地将值转换为浮点数"""
        try:
            if value is None:
                return 0.0
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _safe_str(self, value):
        """安全地将值转换为字符串"""
        if value is None:
            return ''
        return str(value).strip()

    def _update_flags(self, data):
        """更新 Flags 标记系统"""
        if 'flags' not in data:
            data['flags'] = {}

        def safe_float(value):
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0.0

        data['flags']['has_scope_1'] = safe_float(data.get('scope_1_emissions', 0)) > 0
        data['flags']['has_scope_2_location'] = safe_float(data.get('scope_2_location_based_emissions', 0)) > 0
        data['flags']['has_scope_2_market'] = safe_float(data.get('scope_2_market_based_emissions', 0)) > 0
        data['flags']['has_scope_3'] = safe_float(data.get('scope_3_emissions', 0)) > 0

        for i in range(1, 16):
            key = f'scope_3_category_{i}_emissions'
            flag_key = f'has_scope_3_category_{i}'
            data['flags'][flag_key] = safe_float(data.get(key, 0)) > 0

        return data

    def _find_activity_summary_sheet(self):
        """
        自动查找包含"活动数据汇总"和"基于位置"关键词的工作表

        Returns:
            找到的工作表名称，如果未找到返回 None
        """
        if not self.workbook:
            return None

        try:
            for sheet_name in self.workbook.sheetnames:
                # 检查sheet名称是否同时包含"活动数据汇总"和"基于位置"
                if '活动数据汇总' in sheet_name and '基于位置' in sheet_name:
                    print(f"找到活动数据汇总表（基于位置）: {sheet_name}")
                    return sheet_name

            # 如果名称中没找到，尝试检查内容
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                has_activity_summary = False
                has_location_based = False

                # 检查前20行内容
                for row in sheet.iter_rows(max_row=20):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell_text = str(cell.value)
                            if '活动数据汇总' in cell_text:
                                has_activity_summary = True
                            if '基于位置' in cell_text or 'Location-based' in cell_text:
                                has_location_based = True

                if has_activity_summary and has_location_based:
                    print(f"通过内容识别活动数据汇总表（基于位置）: {sheet_name}")
                    return sheet_name

            print("警告：未找到活动数据汇总表（基于位置）")
            return None

        except Exception as e:
            print(f"查找活动数据汇总表时出错: {e}")
            return None

    # ========== 基于特征指纹的表格识别和提取 ==========

    def _identify_table_type(self, sheet_name, check_rows=20):
        """
        基于特征指纹识别表格类型

        Args:
            sheet_name: 工作表名称
            check_rows: 检查前N行以识别特征

        Returns:
            匹配的协议类型名称，如 'EmissionFactorProtocol'
            如果未匹配返回 None
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return None

        try:
            sheet = self.workbook[sheet_name]

            # 收集前N行的所有唯一字符串值
            unique_strings = set()
            for row_idx in range(1, min(check_rows + 1, sheet.max_row + 1)):
                for cell in sheet[row_idx]:
                    if cell.value and isinstance(cell.value, str):
                        # 去除空格并添加到集合
                        cleaned_value = str(cell.value).strip()
                        if cleaned_value:
                            unique_strings.add(cleaned_value)

            # 对每个协议类型进行匹配
            for protocol_name, protocol_config in TABLE_PROTOCOLS.items():
                required = protocol_config['required_keywords']

                # 检查是否包含所有必需关键词
                if required.issubset(unique_strings):
                    # 计算匹配度（可选关键词的匹配数）
                    optional_keywords = protocol_config['keywords'] - required
                    matched_optional = len(optional_keywords & unique_strings)
                    total_optional = len(optional_keywords) if optional_keywords else 1

                    match_ratio = matched_optional / total_optional if total_optional > 0 else 1

                    print(f"[表格识别] {sheet_name} 匹配到 {protocol_config['name']} "
                          f"(必需关键词: {len(required)}/{len(required)}, "
                          f"可选关键词: {matched_optional}/{total_optional})")

                    # 如果匹配度超过50%，则认为是该类型
                    if match_ratio >= 0.3:
                        return protocol_name

            print(f"[表格识别] {sheet_name} 未匹配到已知协议类型")
            print(f"  发现的关键词: {sorted(list(unique_strings))[:20]}")
            return None

        except Exception as e:
            print(f"[表格识别] 识别表格类型时出错: {e}")
            return None

    def _find_header_row(self, sheet, keywords, max_row=20):
        """
        查找包含特定关键词的表头行

        Args:
            sheet: openpyxl工作表对象
            keywords: 表头关键词集合
            max_row: 最大搜索行数

        Returns:
            表头行索引（从1开始），未找到返回None
        """
        for row_idx in range(1, min(max_row + 1, sheet.max_row + 1)):
            row_values = set()
            for cell in sheet[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    row_values.add(str(cell.value).strip())

            # 检查是否包含足够的关键词
            matched_count = len(keywords & row_values)
            if matched_count >= min(2, len(keywords)):
                return row_idx

        return None

    def _get_column_mapping(self, sheet, header_row, field_mapping):
        """
        根据表头行创建列索引映射
        支持多行表头（主表头 + 子表头）

        Args:
            sheet: openpyxl工作表对象
            header_row: 表头行索引
            field_mapping: 字段映射配置

        Returns:
            {标准字段名: 列索引} 的字典
        """
        column_map = {}

        # 收集多行表头的所有值（支持主表头和子表头）
        header_cells = {}
        rows_to_check = [header_row, header_row + 1]  # 检查主表头和下一行（子表头）

        for row_idx in rows_to_check:
            if row_idx > sheet.max_row:
                continue
            for cell in sheet[row_idx]:
                if cell.value and isinstance(cell.value, str):
                    header_value = str(cell.value).strip()
                    # 只记录第一次遇到的列（主表头优先）
                    if header_value not in header_cells:
                        header_cells[header_value] = cell.column - 1  # 转换为0-based索引

        # 为每个标准字段查找对应的列
        for standard_field, (keyword, default, dtype) in field_mapping.items():
            # 尝试精确匹配
            if keyword in header_cells:
                column_map[standard_field] = header_cells[keyword]
            else:
                # 尝试模糊匹配
                for header_value in header_cells:
                    if keyword in header_value or header_value in keyword:
                        column_map[standard_field] = header_cells[header_value]
                        break

        return column_map

    def _apply_ffill(self, data_items, field_names):
        """
        对指定字段应用前向填充(ffill)逻辑，处理合并单元格

        Args:
            data_items: 数据项列表
            field_names: 需要应用ffill的字段名列表

        Returns:
            处理后的数据项列表
        """
        if not data_items:
            return data_items

        result = []
        last_values = {field: None for field in field_names}

        for item in data_items:
            new_item = item.copy()

            for field in field_names:
                value = item.get(field)

                if value and str(value).strip():  # 有值
                    last_values[field] = value
                    new_item[field] = value
                elif last_values[field]:  # 空值，使用前一个值
                    new_item[field] = last_values[field]

            result.append(new_item)

        return result

    def _extract_protocol_data(self, sheet_name, protocol_name):
        """
        根据协议类型从工作表中提取标准化数据

        Args:
            sheet_name: 工作表名称
            protocol_name: 协议类型名称

        Returns:
            标准化的数据项列表
        """
        if protocol_name not in TABLE_PROTOCOLS:
            print(f"[数据提取] 未知协议类型: {protocol_name}")
            return []

        protocol_config = TABLE_PROTOCOLS[protocol_name]
        field_mapping = protocol_config['field_mapping']

        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []

        try:
            sheet = self.workbook[sheet_name]

            # 查找表头行
            header_keywords = {v[0] for v in field_mapping.values() if v[0]}
            header_row = self._find_header_row(sheet, header_keywords)

            if not header_row:
                print(f"[数据提取] {sheet_name}: 未找到表头行")
                return []

            print(f"[数据提取] {sheet_name}: 找到表头在第{header_row}行")

            # 获取列映射
            column_map = self._get_column_mapping(sheet, header_row, field_mapping)

            if not column_map:
                print(f"[数据提取] {sheet_name}: 无法创建列映射")
                return []

            print(f"[数据提取] {sheet_name}: 列映射: {column_map}")

            # 提取数据行
            data_items = []
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                row = sheet[row_idx]

                # 检查是否是空行
                row_has_data = False
                for cell in row:
                    if cell.value is not None:
                        row_has_data = True
                        break

                if not row_has_data:
                    continue

                # 提取字段值
                item = {}
                for standard_field, (keyword, default, dtype) in field_mapping.items():
                    if standard_field in column_map:
                        col_idx = column_map[standard_field]
                        cell = row[col_idx] if col_idx < len(row) else None

                        if cell and cell.value is not None:
                            # 根据数据类型转换
                            if dtype == 'float':
                                try:
                                    item[standard_field] = float(cell.value)
                                except (ValueError, TypeError):
                                    item[standard_field] = default
                            elif dtype == 'int':
                                try:
                                    item[standard_field] = int(cell.value)
                                except (ValueError, TypeError):
                                    item[standard_field] = default
                            else:  # str
                                item[standard_field] = str(cell.value).strip()
                        else:
                            item[standard_field] = default
                    else:
                        item[standard_field] = default

                # 检查是否是有效数据行（至少有一个非默认值）
                has_valid_data = False
                for standard_field, (keyword, default, dtype) in field_mapping.items():
                    value = item.get(standard_field)
                    if value and value != default and str(value).strip():
                        has_valid_data = True
                        break

                if has_valid_data:
                    data_items.append(item)

            print(f"[数据提取] {sheet_name}: 提取到 {len(data_items)} 行数据")

            # 对特定字段应用ffill（如类别字段）
            if protocol_name == 'EmissionFactorProtocol':
                # 合并备选排放系数字段到主字段 ef_val
                for item in data_items:
                    if not item.get('ef_val') or item.get('ef_val') == 0:
                        # 尝试使用备选字段
                        for alt_field in ['ef_val_alt', 'ef_val_alt2']:
                            if item.get(alt_field) and item.get(alt_field) != 0:
                                item['ef_val'] = item[alt_field]
                                break

                # 应用 ffill 填充类别字段
                ffill_fields = ['category']
                data_items = self._apply_ffill(data_items, ffill_fields)
                print(f"[数据提取] {sheet_name}: 应用ffill后数据行数: {len(data_items)}")

            return data_items

        except Exception as e:
            print(f"[数据提取] {sheet_name}: 提取数据时出错: {e}")
            import traceback
            traceback.print_exc()
            return []

    def read_protocols(self):
        """
        基于特征指纹读取所有协议表格

        Returns:
            {output_var: data_items} 的字典
        """
        result = {}

        if not self.workbook:
            print("[协议读取] 工作簿未初始化")
            return result

        print(f"[协议读取] 开始识别和提取协议表格，共 {len(self.workbook.sheetnames)} 个工作表")

        for sheet_name in self.workbook.sheetnames:
            # 识别表格类型
            protocol_type = self._identify_table_type(sheet_name)

            if protocol_type:
                protocol_config = TABLE_PROTOCOLS[protocol_type]
                output_var = protocol_config['output_var']

                # 提取数据
                data_items = self._extract_protocol_data(sheet_name, protocol_type)

                # 存储到结果字典
                result[output_var] = data_items

                print(f"[协议读取] {sheet_name} -> {output_var}: {len(data_items)} 行")

        # 确保所有输出变量都被初始化
        for protocol_config in TABLE_PROTOCOLS.values():
            output_var = protocol_config['output_var']
            if output_var not in result:
                result[output_var] = []
                print(f"[协议读取] 初始化空变量: {output_var} = []")

        # ========== 新增：按类别分组 pro_ef_items ==========
        # 将 pro_ef_items 按类别分组，为三个独立的表格提供数据
        if 'pro_ef_items' in result and result['pro_ef_items']:
            # 同时设置 emission_factor_items（用于 Tables 22-24）
            result['emission_factor_items'] = result['pro_ef_items']
            print(f"[协议读取] 设置 emission_factor_items: {len(result['emission_factor_items'])} 条")

            # 按类别分组
            grouped_data = self._group_pro_ef_items_by_category(result['pro_ef_items'])
            result.update(grouped_data)
            print(f"[协议读取] 按类别分组 pro_ef_items:")
            for group_name, items in grouped_data.items():
                print(f"  {group_name}: {len(items)} 条")
        # ========== 类别分组结束 ==========

        # ========== 新增：提取范围一直接排放源数据 ==========
        # 重要：范围一数据包含实际排放量，会覆盖排放因子的分组数据
        scope1_data = self._extract_scope1_emissions_data()
        # 只更新非空的组（保留排放因子数据的备用）
        for group_name, items in scope1_data.items():
            if items:  # 只有当有数据时才覆盖
                result[group_name] = items
                print(f"[协议读取] 覆盖范围一数据 {group_name}: {len(items)} 条")
        # ========== 范围一数据提取结束 ==========

        return result

    def _group_pro_ef_items_by_category(self, pro_ef_items):
        """
        将 pro_ef_items 按类别分组为模板所需的变量

        Args:
            pro_ef_items: 原始排放因子数据列表

        Returns:
            包含分组数据的字典，使用模板期望的变量名
        """
        grouped = {
            'scope1_stationary_combustion_emissions_items': [],  # 固定燃烧
            'scope1_mobile_combustion_emissions_items': [],      # 移动燃烧
            'scope1_fugitive_emissions_items': [],               # 逸散排放
            'scope1_process_emissions_items': [],                # 制程排放
        }

        # 调试：记录每个类别的分配情况
        debug_categories = {}

        for item in pro_ef_items:
            category = item.get('category', '')

            # 移除空格以处理"移 动燃 烧"这种带空格的类别
            category_normalized = category.replace(' ', '')

            # 详细调试：打印第一个"移动燃烧"的匹配过程
            if '移动燃烧' in category:
                print(f"[DEBUG] Processing category: '{category}'")
                print(f"  category_normalized: '{category_normalized}'")
                print(f"  '固定燃烧' in category_normalized: {'固定燃烧' in category_normalized}")
                print(f"  '移动燃烧' in category_normalized: {'移动燃烧' in category_normalized}")
                print(f"  '移动汽油' in category_normalized: {'移动汽油' in category_normalized}")
                print(f"  '移动柴油' in category_normalized: {'移动柴油' in category_normalized}")
                print(f"  '制冷产品加工使用等排放' in category_normalized: {'制冷产品加工使用等排放' in category_normalized}")
                print(f"  '制程' in category_normalized: {'制程' in category_normalized}")
                print(f"  '逸散' in category_normalized: {'逸散' in category_normalized}")

            # 优先匹配更具体的关键词（使用归一化后的类别进行匹配）
            # 固定燃烧
            if '固定燃烧' in category_normalized:
                grouped['scope1_stationary_combustion_emissions_items'].append(item)
                if category not in debug_categories:
                    debug_categories[category] = []
                debug_categories[category].append('stationary')
            # 移动燃烧（包括移动汽油、移动柴油等）
            elif '移动燃烧' in category_normalized or '移动汽油' in category_normalized or '移动柴油' in category_normalized:
                grouped['scope1_mobile_combustion_emissions_items'].append(item)
                if category not in debug_categories:
                    debug_categories[category] = []
                debug_categories[category].append('mobile')
            # 制程排放
            elif '制冷产品加工使用等排放' in category_normalized or '制程' in category_normalized:
                grouped['scope1_process_emissions_items'].append(item)
                if category not in debug_categories:
                    debug_categories[category] = []
                debug_categories[category].append('process')
            # 逸散排放
            elif '逸散' in category_normalized:
                grouped['scope1_fugitive_emissions_items'].append(item)
                if category not in debug_categories:
                    debug_categories[category] = []
                debug_categories[category].append('fugitive')

        # 调试输出
        print(f"[类别分组] 分组统计:")
        for group_name, items in grouped.items():
            print(f"  {group_name}: {len(items)} 条")

        return grouped

    def _extract_scope1_emissions_data(self):
        """
        从附表1-温室气体盘查表中提取范围一直接排放源数据

        Returns:
            包含分组数据的字典:
            - scope1_stationary_combustion_emissions_items (固定燃烧)
            - scope1_mobile_combustion_emissions_items (移动燃烧)
            - scope1_fugitive_emissions_items (逸散排放)
            - scope1_process_emissions_items (制程排放)
        """
        result = {
            'scope1_stationary_combustion_emissions_items': [],
            'scope1_mobile_combustion_emissions_items': [],
            'scope1_fugitive_emissions_items': [],
            'scope1_process_emissions_items': [],
        }

        if not self.workbook:
            print("[范围一排放] 工作簿未初始化")
            return result

        # 找到附表1-温室气体盘查表
        target_sheet = None
        for sheet in self.workbook.worksheets:
            title = sheet.title
            # 使用更宽松的匹配条件
            if '附表1' in title or ('温室' in title and '盘查' in title and '1' in title):
                target_sheet = sheet
                break

        # 如果还没找到，尝试使用索引
        if not target_sheet and len(self.workbook.worksheets) > 4:
            target_sheet = self.workbook.worksheets[4]  # 通常附表1在索引4

        if not target_sheet:
            print("[范围一排放] 未找到附表1-温室气体盘查表")
            return result

        print(f"[范围一排放] 找到工作表: {target_sheet.title}")

        # 数据从第5行开始（前4行是标题）
        data_start_row = 5

        for row_idx in range(data_start_row, target_sheet.max_row + 1):
            try:
                row = target_sheet[row_idx]

                # 读取各列数据
                number = self._safe_str(row[0].value)
                category = self._safe_str(row[1].value)
                emission_source = self._safe_str(row[2].value)
                facility = self._safe_str(row[3].value)

                # 如果编号为空，跳过
                if not number or number.strip() == '':
                    continue

                # 读取排放量数据（Columns 30-37）
                # 注意：这些列包含公式，需要计算值
                co2_emissions = self._safe_float(row[30].value) if len(row) > 30 else 0
                ch4_emissions = self._safe_float(row[31].value) if len(row) > 31 else 0
                n2o_emissions = self._safe_float(row[32].value) if len(row) > 32 else 0
                hfcs_emissions = self._safe_float(row[33].value) if len(row) > 33 else 0
                pfcs_emissions = self._safe_float(row[34].value) if len(row) > 34 else 0
                sf6_emissions = self._safe_float(row[35].value) if len(row) > 35 else 0
                nf3_emissions = self._safe_float(row[36].value) if len(row) > 36 else 0
                total_emissions = self._safe_float(row[37].value) if len(row) > 37 else 0

                # 创建数据项
                item = {
                    'number': number,
                    'category': category,
                    'emission_source': emission_source,
                    'facility': facility,
                    'CO2_emissions': co2_emissions,
                    'CH4_emissions': ch4_emissions,
                    'N2O_emissions': n2o_emissions,
                    'HFCs_emissions': hfcs_emissions,
                    'PFCs_emissions': pfcs_emissions,
                    'SF6_emissions': sf6_emissions,
                    'NF3_emissions': nf3_emissions,
                    'total_green_house_gas_emissions': total_emissions,
                }

                # 根据类别分组
                if '固定燃烧' in category:
                    result['scope1_stationary_combustion_emissions_items'].append(item)
                elif '移动燃烧' in category or '移动汽油' in category or '移动柴油' in category:
                    result['scope1_mobile_combustion_emissions_items'].append(item)
                elif '逸散排放' in category or '逸散' in category:
                    result['scope1_fugitive_emissions_items'].append(item)
                elif '制程排放' in category or '制程' in category:
                    result['scope1_process_emissions_items'].append(item)

            except Exception as e:
                print(f"[范围一排放] 处理Row {row_idx}时出错: {e}")
                continue

        print(f"[范围一排放] 提取完成:")
        for group_name, items in result.items():
            print(f"  {group_name}: {len(items)} 条")

        return result

    def _extract_activity_summary_data(self, sheet_name):
        """
        从活动数据汇总表中提取数据

        Args:
            sheet_name: 工作表名称

        Returns:
            包含活动数据汇总的字典列表
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []

        try:
            sheet = self.workbook[sheet_name]
            result = []

            # 第一步：定位表头行
            # 表头特征：包含"序号"、"排放源"、"报告边界"、"活动数据"等关键词
            # 注意：表头行可能不包含所有关键词，只要包含大部分即可
            header_row_idx = None
            header_keywords = ['序号', '排放源', '报告边界', '活动数据']

            for row_idx in range(1, min(20, sheet.max_row + 1)):
                row = sheet[row_idx]
                row_values = [str(cell.value) if cell.value else '' for cell in row]
                row_text = ''.join(row_values)

                # 检查是否包含表头关键词（至少包含3个）
                match_count = sum(1 for keyword in header_keywords if keyword in row_text)
                if match_count >= 3:
                    header_row_idx = row_idx
                    print(f"找到表头行: 第{row_idx}行 (匹配关键词数: {match_count})")
                    break

            if header_row_idx is None:
                print("警告：未找到活动数据汇总表的表头行")
                return []

            # 第二步：读取表头，建立列名到列索引的映射
            header_row = sheet[header_row_idx]
            column_mapping = {}

            for col_idx, cell in enumerate(header_row, start=1):
                if cell.value:
                    col_name = str(cell.value).strip()
                    column_mapping[col_name] = col_idx

            print(f"表头列映射: {list(column_mapping.keys())}")

            # 第三步：确定各排放量列的位置
            # 根据表格结构：
            # 列30: CO2排放量, 列31: CH4排放量, 列32: N2O排放量
            # 列33: HFCs排放量, 列34: PFCs排放量, 列35: SF6排放量
            # 列36: NF3排放量, 列37: 总量

            # 实际列位置需要通过查找包含"tCO2e"和温室气体名称的表头来确定
            emission_columns = {}
            for row_idx in range(max(1, header_row_idx - 2), header_row_idx + 3):
                row = sheet[row_idx]
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value and isinstance(cell.value, str):
                        val = str(cell.value).strip()
                        # 查找包含温室气体名称且在较后列（通常是排放量列）的单元格
                        # 同时检查是否包含"tCO2e"或位于列20之后
                        if col_idx >= 20:
                            if val == 'CO2' or 'CO2排放' in val:
                                emission_columns['CO2_emissions'] = col_idx
                            elif val == 'CH4' or 'CH4排放' in val:
                                emission_columns['CH4_emissions'] = col_idx
                            elif val == 'N2O' or 'N2O排放' in val:
                                emission_columns['N2O_emissions'] = col_idx
                            elif val == 'HFCs' or 'HFCs排放' in val:
                                emission_columns['HFCs_emissions'] = col_idx
                            elif val == 'PFCs' or 'PFCs排放' in val:
                                emission_columns['PFCs_emissions'] = col_idx
                            elif val == 'SF6' or 'SF6排放' in val:
                                emission_columns['SF6_emissions'] = col_idx
                            elif val == 'NF3' or 'NF3排放' in val:
                                emission_columns['NF3_emissions'] = col_idx
                            elif '总量' in val or '总计' in val:
                                emission_columns['total_green_house_gas_emissions'] = col_idx

            print(f"排放量列映射: {emission_columns}")

            # 第四步：读取数据行
            current_emission_category = ''  # 用于前向填充GHG排放类别

            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                row = sheet[row_idx]

                # 获取序号列的值，判断是否是有效数据行
                number_cell = row[0]  # 序号在第1列
                number_value = number_cell.value

                # 如果序号列为空，跳过
                if number_value is None or str(number_value).strip() == '':
                    continue

                # 尝试将序号转换为数字
                try:
                    number_val = float(str(number_value).strip())
                    if number_val <= 0:  # 跳过非正数序号
                        continue
                except (ValueError, TypeError):
                    continue

                # 获取各列的值
                def get_cell_value(col_idx):
                    if col_idx < 1 or col_idx > len(row):
                        return ''
                    cell = row[col_idx - 1]
                    if cell.value is None:
                        return ''
                    return str(cell.value).strip()

                # GHG排放类别（前向填充）
                emission_category = get_cell_value(2)  # 第2列
                if emission_category and emission_category != '':
                    current_emission_category = emission_category

                # 提取数据 - 使用与模板匹配的字段名
                item = {
                    'number': get_cell_value(1),  # 序号
                    'emission_source_type_loc': current_emission_category,  # GHG排放类别（简短名称，与模板匹配）
                    'emission_source_type_location_based': current_emission_category,  # 完整名称（兼容）
                    'emission_source_loc': get_cell_value(3),  # 排放源（简短名称）
                    'emission_source_location_based': get_cell_value(3),  # 完整名称（兼容）
                    'report_boundary_loc': get_cell_value(4),  # 报告边界（简短名称）
                    'report_boundary_location_based': get_cell_value(4),  # 完整名称（兼容）
                    'act_summary_loc': get_cell_value(5),  # 活动数据数值（模板期望的字段名）
                    'activity_data_location_based': get_cell_value(5),  # 完整名称（兼容）
                    'act_summary_loc_unit': get_cell_value(6),  # 活动数据单位（模板期望的字段名）
                    'activity_data_unit_location_based': get_cell_value(6),  # 完整名称（兼容）
                }

                # 排放量数据
                for field_name, col_idx in emission_columns.items():
                    value = get_cell_value(col_idx)
                    item[field_name] = value
                    # 同时保留带 _formatted 后缀的版本（将在格式化时填充）
                    item[f'{field_name}_formatted'] = value

                result.append(item)

            print(f"从活动数据汇总表提取到 {len(result)} 行数据")
            return result

        except Exception as e:
            print(f"提取活动数据汇总表数据时出错: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _find_activity_summary_sheet_market_based(self):
        """
        自动查找包含"活动数据汇总"和"基于市场"关键词的工作表

        Returns:
            找到的工作表名称，如果未找到返回 None
        """
        if not self.workbook:
            return None

        try:
            for sheet_name in self.workbook.sheetnames:
                # 检查sheet名称是否同时包含"活动数据汇总"和"基于市场"
                if '活动数据汇总' in sheet_name and '基于市场' in sheet_name:
                    print(f"找到活动数据汇总表（基于市场）: {sheet_name}")
                    return sheet_name

            # 如果名称中没找到，尝试检查内容
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                has_activity_summary = False
                has_market_based = False

                # 检查前20行内容
                for row in sheet.iter_rows(max_row=20):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            cell_text = str(cell.value)
                            if '活动数据汇总' in cell_text:
                                has_activity_summary = True
                            if '基于市场' in cell_text or 'Market-based' in cell_text:
                                has_market_based = True

                if has_activity_summary and has_market_based:
                    print(f"通过内容识别活动数据汇总表（基于市场）: {sheet_name}")
                    return sheet_name

            print("警告：未找到活动数据汇总表（基于市场）")
            return None

        except Exception as e:
            print(f"查找活动数据汇总表（基于市场）时出错: {e}")
            return None

    def _extract_activity_summary_data_market_based(self, sheet_name):
        """
        从活动数据汇总表（基于市场）中提取数据

        Args:
            sheet_name: 工作表名称

        Returns:
            包含活动数据汇总的字典列表
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            return []

        try:
            sheet = self.workbook[sheet_name]
            result = []

            # 第一步：定位表头行
            header_row_idx = None
            header_keywords = ['序号', '排放源', '报告边界', '活动数据']

            for row_idx in range(1, min(20, sheet.max_row + 1)):
                row = sheet[row_idx]
                row_values = [str(cell.value) if cell.value else '' for cell in row]
                row_text = ''.join(row_values)

                # 检查是否包含表头关键词（至少包含3个）
                match_count = sum(1 for keyword in header_keywords if keyword in row_text)
                if match_count >= 3:
                    header_row_idx = row_idx
                    print(f"找到表头行: 第{row_idx}行 (匹配关键词数: {match_count})")
                    break

            if header_row_idx is None:
                print("警告：未找到活动数据汇总表（基于市场）的表头行")
                return []

            # 第二步：读取表头，建立列名到列索引的映射
            header_row = sheet[header_row_idx]
            column_mapping = {}

            for col_idx, cell in enumerate(header_row, start=1):
                if cell.value:
                    col_name = str(cell.value).strip()
                    column_mapping[col_name] = col_idx

            print(f"表头列映射: {list(column_mapping.keys())}")

            # 第三步：确定各排放量列的位置
            emission_columns = {}
            for row_idx in range(max(1, header_row_idx - 2), header_row_idx + 3):
                row = sheet[row_idx]
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value and isinstance(cell.value, str):
                        val = str(cell.value).strip()
                        # 查找包含温室气体名称且在较后列（通常是排放量列）的单元格
                        if col_idx >= 20:
                            if val == 'CO2' or 'CO2排放' in val:
                                emission_columns['CO2_emissions'] = col_idx
                            elif val == 'CH4' or 'CH4排放' in val:
                                emission_columns['CH4_emissions'] = col_idx
                            elif val == 'N2O' or 'N2O排放' in val:
                                emission_columns['N2O_emissions'] = col_idx
                            elif val == 'HFCs' or 'HFCs排放' in val:
                                emission_columns['HFCs_emissions'] = col_idx
                            elif val == 'PFCs' or 'PFCs排放' in val:
                                emission_columns['PFCs_emissions'] = col_idx
                            elif val == 'SF6' or 'SF6排放' in val:
                                emission_columns['SF6_emissions'] = col_idx
                            elif val == 'NF3' or 'NF3排放' in val:
                                emission_columns['NF3_emissions'] = col_idx
                            elif '总量' in val or '总计' in val:
                                emission_columns['total_green_house_gas_emissions'] = col_idx

            print(f"排放量列映射: {emission_columns}")

            # 第四步：读取数据行
            current_emission_category = ''  # 用于前向填充GHG排放类别

            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                row = sheet[row_idx]

                # 获取序号列的值，判断是否是有效数据行
                number_cell = row[0]  # 序号在第1列
                number_value = number_cell.value

                # 如果序号列为空，跳过
                if number_value is None or str(number_value).strip() == '':
                    continue

                # 尝试将序号转换为数字
                try:
                    number_val = float(str(number_value).strip())
                    if number_val <= 0:  # 跳过非正数序号
                        continue
                except (ValueError, TypeError):
                    continue

                # 获取各列的值
                def get_cell_value(col_idx):
                    if col_idx < 1 or col_idx > len(row):
                        return ''
                    cell = row[col_idx - 1]
                    if cell.value is None:
                        return ''
                    return str(cell.value).strip()

                # GHG排放类别（前向填充）
                emission_category = get_cell_value(2)  # 第2列
                if emission_category and emission_category != '':
                    current_emission_category = emission_category

                # 提取数据 - 使用与模板匹配的字段名
                item = {
                    'number': get_cell_value(1),  # 序号
                    'emission_source_type_mar': current_emission_category,  # GHG排放类别（简短名称，与模板匹配）
                    'emission_source_type_market_based': current_emission_category,  # 完整名称（兼容）
                    'emission_source_mar': get_cell_value(3),  # 排放源（简短名称，与模板匹配）
                    'emission_source_market_based': get_cell_value(3),  # 完整名称（兼容）
                    'report_boundary_mar': get_cell_value(4),  # 报告边界（简短名称，与模板匹配）
                    'report_boundary_market_based': get_cell_value(4),  # 完整名称（兼容）
                    'act_summary_mar': get_cell_value(5),  # 活动数据数值（模板期望的字段名）
                    'activity_data_market_based': get_cell_value(5),  # 完整名称（兼容）
                    'act_summary_mar_unit': get_cell_value(6),  # 活动数据单位（模板期望的字段名）
                    'activity_data_unit_market_based': get_cell_value(6),  # 完整名称（兼容）
                }

                # 排放量数据
                for field_name, col_idx in emission_columns.items():
                    value = get_cell_value(col_idx)
                    item[field_name] = value
                    # 同时保留带 _formatted 后缀的版本（将在格式化时填充）
                    item[f'{field_name}_formatted'] = value

                result.append(item)

            print(f"从活动数据汇总表（基于市场）提取到 {len(result)} 行数据")
            return result

        except Exception as e:
            print(f"提取活动数据汇总表（基于市场）数据时出错: {e}")
            import traceback
            traceback.print_exc()
            return []

    def get_merged_cell_value(self, sheet, row, col):
        """
        获取单元格的值，处理合并单元格的情况。

        如果指定坐标在合并区域内，返回该合并区域左上角单元格的值。
        否则返回该单元格的直接值。

        Args:
            sheet: openpyxl 工作表对象
            row: 行号（从1开始）
            col: 列号（从1开始）

        Returns:
            单元格的值，如果是合并单元格则返回合并区域左上角的值
        """
        try:
            # 检查是否在合并区域内
            for merged_range in sheet.merged_cells.ranges:
                if merged_range.min_row <= row <= merged_range.max_row and \
                   merged_range.min_col <= col <= merged_range.max_col:
                    # 在合并区域内，返回左上角单元格的值
                    top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    return top_left_cell.value

            # 不在合并区域内，直接返回单元格值
            cell = sheet.cell(row=row, column=col)
            return cell.value
        except Exception as e:
            print(f"获取合并单元格值时出错 (row={row}, col={col}): {e}")
            return None

    def __init__(self, filepath):
        """ 
        初始化时，加载 Excel 工作簿。 
        """ 
        self.workbook = None
        self.filepath = filepath
        self.file_type = None
        self.company_name = None
        self.reporting_period = '2024年'  # 默认报告期
        
        # 检查文件类型
        if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
            self.file_type = 'excel'
            try:
                self.workbook = openpyxl.load_workbook(filepath, data_only=True)
                print(f"成功加载 Excel: {filepath}")
            except FileNotFoundError:
                print(f"错误：找不到文件 {filepath}")
            except Exception as e:
                print(f"加载 Excel 出错: {e}")
        elif filepath.endswith('.csv'):
            self.file_type = 'csv'
            print(f"识别到 CSV 文件: {filepath}")
        else:
            print(f"错误：不支持的文件类型 {filepath}")

    def find_value_by_label(self, sheet_name, label_name, column=None, search_direction='right',
                           exact_match=False, case_sensitive=False, max_rows=None):
        """
        通用函数：遍历 Excel 的某一列或整个工作表，找到包含 label_name 的单元格，然后返回其相邻单元格的值。
        摒弃硬坐标定位，使代码更健壮，能应对 Excel 格式的微调。

        Args:
            sheet_name: 工作表名称
            label_name: 要查找的标签文本
            column: 指定列（如'A', 'B'），如果为None则搜索整个工作表
            search_direction: 搜索方向 - 'right'（右侧）、'left'（左侧）、'below'（下方）、'above'（上方）
            exact_match: 是否要求精确匹配（False表示包含匹配）
            case_sensitive: 是否区分大小写
            max_rows: 最大搜索行数，None表示搜索所有行

        Returns:
            找到的值，如果没找到返回None
        """
        if not self.workbook or self.file_type != 'excel':
            return None

        try:
            sheet = self.workbook[sheet_name]
            if sheet_name not in self.workbook.sheetnames:
                print(f"错误：找不到工作表 {sheet_name}")
                return None

            # 准备标签文本
            search_text = str(label_name) if case_sensitive else str(label_name).lower()

            # 搜索范围
            search_range = []
            if column:
                # 搜索指定列
                column_index = openpyxl.utils.column_index_from_string(column)
                max_search_rows = min(max_rows or sheet.max_row, sheet.max_row)
                for row in range(1, max_search_rows + 1):
                    search_range.append(sheet.cell(row=row, column=column_index))
            else:
                # 搜索整个工作表
                max_search_rows = min(max_rows or sheet.max_row, sheet.max_row)
                for row in range(1, max_search_rows + 1):
                    for col in range(1, sheet.max_column + 1):
                        search_range.append(sheet.cell(row=row, column=col))

            # 搜索匹配的单元格
            matched_cells = []
            for cell in search_range:
                if cell.value is not None:
                    cell_text = str(cell.value)
                    compare_text = cell_text if case_sensitive else cell_text.lower()

                    # 匹配逻辑
                    if exact_match:
                        if search_text == compare_text:
                            matched_cells.append(cell)
                    else:
                        if search_text in compare_text:
                            matched_cells.append(cell)

            if not matched_cells:
                print(f"警告：在 {sheet_name} 中未找到包含 '{label_name}' 的单元格")
                return None

            # 返回第一个匹配单元格相邻的值
            target_cell = matched_cells[0]
            value_cell = None

            if search_direction == 'right':
                value_cell = sheet.cell(row=target_cell.row, column=target_cell.column + 1)
            elif search_direction == 'left':
                if target_cell.column > 1:
                    value_cell = sheet.cell(row=target_cell.row, column=target_cell.column - 1)
            elif search_direction == 'below':
                value_cell = sheet.cell(row=target_cell.row + 1, column=target_cell.column)
            elif search_direction == 'above':
                if target_cell.row > 1:
                    value_cell = sheet.cell(row=target_cell.row - 1, column=target_cell.column)

            return value_cell.value if value_cell and value_cell.value is not None else None

        except Exception as e:
            print(f"查找标签 '{label_name}' 时出错: {e}")
            return None

    def _find_value_next_to(self, sheet_name, keyword): 
        """ 
        私有方法，用于实现向后兼容。
        在指定的工作表中查找一个关键词，并返回其右侧单元格的值。
        """ 
        if not self.workbook: 
            return None 
            
        try: 
            sheet = self.workbook[sheet_name] 
            for row in sheet.iter_rows(): 
                for cell in row: 
                    if cell.value == keyword: 
                        # 找到了关键词！返回它右边一列的值 
                        value_cell = sheet.cell(row=cell.row, column=cell.column + 1) 
                        return value_cell.value 
            print(f"警告：在 {sheet_name} 中未找到关键词 '{keyword}'") 
            return None 
        except KeyError: 
            print(f"错误：找不到工作表 {sheet_name}") 
            return None 
        except Exception as e: 
            print(f"查找关键词 '{keyword}' 时出错: {e}") 
            return None
            
    def _find_value_below(self, sheet_name, keyword):
        """ 
        在指定的工作表中查找关键词，并返回其下方单元格的值。
        """
        if not self.workbook: 
            return None 
            
        try: 
            sheet = self.workbook[sheet_name] 
            for row in sheet.iter_rows(): 
                for cell in row: 
                    if cell.value == keyword: 
                        # 找到了关键词！返回它下方单元格的值 
                        value_cell = sheet.cell(row=cell.row + 1, column=cell.column) 
                        return value_cell.value 
            print(f"警告：在 {sheet_name} 中未找到关键词 '{keyword}'") 
            return None 
        except Exception as e: 
            print(f"查找关键词 '{keyword}' 下方值时出错: {e}") 
            return None
            
    def _find_value_by_content(self, sheet_name, keyword_substring):
        """ 
        在指定的工作表中查找包含关键词子串的单元格，并返回其下方单元格的值。
        用于模糊匹配，如'范围三'可能出现在不同格式的单元格中。
        """
        if not self.workbook: 
            return None 
            
        try: 
            sheet = self.workbook[sheet_name] 
            for row in sheet.iter_rows(): 
                for cell in row: 
                    if cell.value is not None and keyword_substring in str(cell.value): 
                        # 找到了包含关键词的单元格！返回它下方单元格的值 
                        value_cell = sheet.cell(row=cell.row + 1, column=cell.column) 
                        return value_cell.value 
            print(f"警告：在 {sheet_name} 中未找到包含 '{keyword_substring}' 的单元格") 
            return None 
        except Exception as e: 
            print(f"查找包含关键词 '{keyword_substring}' 的单元格时出错: {e}") 
            return None 

    def read_to_list_of_dicts(self, sheet_name=None, header_row=1, start_row=None,
                             end_row=None, skip_empty_rows=True, clean_headers=True):
        """
        将 CSV/Excel 文件中的数据转换为列表字典格式。
        能够处理各种数据格式，支持灵活的表头和数据行配置。

        Args:
            sheet_name: Excel工作表名称（CSV文件不需要）
            header_row: 表头所在行（默认第1行）
            start_row: 数据开始行（默认header_row+1）
            end_row: 数据结束行（默认工作表末尾）
            skip_empty_rows: 是否跳过空行
            clean_headers: 是否清理表头（去空格、标准化）

        Returns:
            列表字典格式: [{"列名1": 值1, "列名2": 值2, ...}, ...]
        """
        result = []

        if self.file_type == 'csv':
            # 处理 CSV 文件
            try:
                # 尝试多种编码
                encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'latin-1']
                csv_data = None

                for encoding in encodings:
                    try:
                        with open(self.filepath, 'r', encoding=encoding, newline='') as csvfile:
                            # 读取所有行
                            lines = csvfile.readlines()

                            # 处理表头行
                            if header_row <= len(lines):
                                header_line = lines[header_row - 1].strip()
                                headers = [h.strip() if clean_headers else h for h in header_line.split(',')]

                                # 确保表头不为空
                                for i, h in enumerate(headers):
                                    if not h or h.strip() == '':
                                        headers[i] = f"column_{i+1}"

                                # 处理数据行
                                data_start = start_row or header_row + 1
                                data_end = min(end_row or len(lines), len(lines))

                                for line_num in range(data_start, data_end + 1):
                                    if line_num <= len(lines):
                                        line = lines[line_num - 1].strip()
                                        if line:  # 非空行
                                            values = [v.strip() for v in line.split(',')]

                                            # 创建行字典
                                            row_dict = {}
                                            for i, header in enumerate(headers):
                                                value = values[i] if i < len(values) else None
                                                row_dict[header] = self._clean_cell_value(value)

                                            # 检查是否跳过空行
                                            if not skip_empty_rows or any(v is not None and v != '' for v in row_dict.values()):
                                                result.append(row_dict)
                                csv_data = True
                                break
                    except UnicodeDecodeError:
                        continue

                if csv_data is None:
                    print(f"无法使用任何编码读取 CSV 文件: {self.filepath}")
                else:
                    print(f"成功从 CSV 文件读取 {len(result)} 行数据")

            except Exception as e:
                print(f"读取 CSV 文件时出错: {e}")

        elif self.file_type == 'excel' and sheet_name:
            # 处理 Excel 文件
            if not self.workbook:
                return result

            try:
                if sheet_name not in self.workbook.sheetnames:
                    print(f"错误：找不到工作表 {sheet_name}")
                    return result

                sheet = self.workbook[sheet_name]

                # 获取表头
                headers = []
                header_row_obj = sheet[header_row]
                for cell in header_row_obj:
                    if cell.value is not None:
                        header_text = str(cell.value).strip()
                        if clean_headers:
                            # 清理表头：去空格、标准化
                            header_text = header_text.replace(' ', '_').replace('\n', '_').strip()
                        headers.append(header_text if header_text else f"column_{cell.column}")
                    else:
                        headers.append(f"column_{cell.column}")

                # 确保表头不为空
                for i, h in enumerate(headers):
                    if not h or h.strip() == '':
                        headers[i] = f"column_{i+1}"

                # 读取数据行
                data_start = start_row or header_row + 1
                data_end = min(end_row or sheet.max_row, sheet.max_row)

                for row in range(data_start, data_end + 1):
                    row_dict = {}
                    has_data = False

                    for col in range(1, len(headers) + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        cleaned_value = self._clean_cell_value(cell_value)
                        row_dict[headers[col-1]] = cleaned_value

                        if cleaned_value is not None and cleaned_value != '':
                            has_data = True

                    # 根据参数决定是否跳过空行
                    if not skip_empty_rows or has_data:
                        result.append(row_dict)

                print(f"成功从 Excel 工作表 {sheet_name} 读取 {len(result)} 行数据")

            except Exception as e:
                print(f"读取 Excel 工作表 {sheet_name} 时出错: {e}")

        else:
            print(f"错误：文件类型 {self.file_type} 或缺少必要参数")

        return result

    def _clean_cell_value(self, value):
        """
        清理和标准化单元格值 - 增强版，彻底去除冗余空格

        Args:
            value: 原始单元格值

        Returns:
            清理后的值
        """
        if value is None:
            return None

        # 处理数字
        if isinstance(value, (int, float)):
            return value

        # 处理字符串
        if isinstance(value, str):
            # 1. 先去除首尾空格和换行符
            cleaned = value.strip()
            
            # 2. 替换换行符为空格，并去除多余空格
            cleaned = cleaned.replace("\n", " ")
            # 3. 将多个连续空格替换为单个空格
            cleaned = re.sub(r'\s+', ' ', cleaned)
            # 4. 再次去除首尾空格（确保不会以空格开头或结尾）
            cleaned = cleaned.strip()

            # 空字符串转为None
            if cleaned == '':
                return None

            # 尝试转换为数字
            try:
                if '.' in cleaned:
                    return float(cleaned)
                else:
                    return int(cleaned)
            except ValueError:
                return cleaned

        return value

    def read_emission_data_csv(self, csv_path='减排行动统计.csv'):
        """
        读取减排行动统计CSV文件，提取所有模板需要的变量

        Args:
            csv_path: CSV文件路径

        Returns:
            包含所有32个模板变量的字典
        """
        import os

        # 如果传入的filepath是CSV文件，使用它
        if self.file_type == 'csv' and self.filepath:
            csv_path = self.filepath

        # 检查文件是否存在
        if not os.path.exists(csv_path):
            print(f"警告：CSV文件不存在: {csv_path}")
            return {}

        data = {}
        encodings = ['gbk', 'gb2312', 'utf-8-sig', 'utf-8']

        for encoding in encodings:
            try:
                with open(csv_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    next(reader)  # 跳过表头
                    for row in reader:
                        if len(row) >= 2 and row[0].strip():
                            key = row[0].strip()
                            value = row[1].strip() if len(row) > 1 else ''
                            data[key] = value

                print(f"成功从CSV读取 {len(data)} 个字段 (编码: {encoding})")

                # 计算scope_3_emissions总和（如果CSV中没有）
                # 保持为 float 类型，不转换为字符串
                if 'scope_3_emissions' not in data:
                    scope3_total = 0.0
                    for i in range(1, 16):
                        key = f'scope_3_category_{i}_emissions'
                        if key in data and data[key]:
                            try:
                                scope3_total += float(data[key])
                            except ValueError:
                                pass
                    data['scope_3_emissions'] = round(scope3_total, 6)
                    print(f"计算得出 scope_3_emissions: {data['scope_3_emissions']}")

                return data

            except (UnicodeDecodeError, Exception) as e:
                continue

        print(f"错误：无法使用任何编码读取CSV文件")
        return {}

    def _parse_csv_sections(self, csv_path='减排行动统计.csv'):
        """
        按区域解析CSV文件，保留行号顺序
        用于提取表格数据（范围一、范围二三的排放源）

        Returns:
            dict: {'scope1_items': [...], 'scope2_3_items': [...]}
        """
        import os
        import re

        # 检查文件是否存在
        if not os.path.exists(csv_path):
            return {'scope1_items': [], 'scope2_3_items': []}

        encodings = ['gbk', 'gb2312', 'utf-8-sig', 'utf-8']

        for encoding in encodings:
            try:
                with open(csv_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    rows = list(reader)

                # 查找区域标记的位置（不写死行号）
                scope1_start = None
                scope2_3_start = None

                for i, row in enumerate(rows):
                    if len(row) >= 1 and row[0]:
                        row_text = str(row[0])
                        # 范围一：匹配"范围一"+"直接"+"排放源"
                        if '范围一' in row_text and '直接' in row_text and '排放源' in row_text:
                            scope1_start = i
                            print(f"找到范围一标记在第 {i+1} 行: {row_text}")
                        # 范围二三：匹配"范围二"或"范围二三"+"排放源"（间接可能有编码问题，用更宽松的匹配）
                        elif ('范围二' in row_text or '范围二三' in row_text) and '排放源' in row_text:
                            # 确保不是范围一
                            if '范围一' not in row_text:
                                scope2_3_start = i
                                print(f"找到范围二三标记在第 {i+1} 行: {row_text}")

                # 解析范围一数据（从"范围一直接排放源"到"范围二三间接排放源"之前）
                scope1_items = []
                if scope1_start is not None:
                    end = scope2_3_start if scope2_3_start else len(rows)

                    # CSV数据结构：第1列=类别，第2列=排放源，第3列=设施
                    # template.docx期望：name=类别, emission=排放源, note=设施
                    for i in range(scope1_start + 2, end):  # +2 跳过标记行和表头行
                        if len(rows[i]) >= 2 and rows[i][0] and rows[i][1]:
                            category = str(rows[i][0]).strip()
                            source = str(rows[i][1]).strip()
                            facility = str(rows[i][2]).strip() if len(rows[i]) >= 3 else ''

                            # 跳过空值和标题行
                            if not source or source == '排放源' or source == category:
                                continue

                            scope1_items.append({
                                'name': category,
                                'emission': source,  # 排放源放在emission字段
                                'note': facility  # 设施放在note字段
                            })

                    print(f"解析范围一数据: {len(scope1_items)} 条记录")

                # 解析范围二三数据（从"范围二三间接排放源"开始）
                scope2_3_items = []
                if scope2_3_start is not None:
                    # CSV数据结构：第1列=类别，第2列=排放源，第3列=设施
                    # template.docx期望：name=类别, emission=排放源, note=设施
                    for i in range(scope2_3_start + 2, len(rows)):  # +2 跳过标记行和表头行
                        if len(rows[i]) >= 2 and rows[i][0] and rows[i][1]:
                            category = str(rows[i][0]).strip()
                            source = str(rows[i][1]).strip()
                            facility = str(rows[i][2]).strip() if len(rows[i]) >= 3 else ''

                            # 跳过空值和标题行
                            if not source or source == '排放源' or source == category:
                                continue

                            # 过滤掉范围一的重复数据（类别名包含"范围一"）
                            if '范围一' in category:
                                continue

                            scope2_3_items.append({
                                'name': category,
                                'emission': source,  # 排放源放在emission字段
                                'note': facility  # 设施放在note字段
                            })

                    print(f"解析范围二三数据: {len(scope2_3_items)} 条记录")

                return {
                    'scope1_items': scope1_items,
                    'scope2_3_items': scope2_3_items
                }

            except (UnicodeDecodeError, Exception) as e:
                print(f"解析CSV区域时出错 (编码 {encoding}): {e}")
                continue

        return {'scope1_items': [], 'scope2_3_items': []}

    def find_multiple_values_by_pattern(self, sheet_name, patterns, search_direction='right',
                                      max_distance=3, require_numeric=False):
        """
        根据模式匹配查找多个值，用于处理复杂的表格结构

        Args:
            sheet_name: 工作表名称
            patterns: 模式列表，支持正则表达式
            search_direction: 搜索方向
            max_distance: 搜索最大距离（单元格数）
            require_numeric: 是否要求数值结果

        Returns:
            找到的值列表
        """
        import re
        results = []

        if not self.workbook or self.file_type != 'excel':
            return results

        try:
            sheet = self.workbook[sheet_name]
            if sheet_name not in self.workbook.sheetnames:
                return results

            for pattern in patterns:
                # 编译正则表达式
                try:
                    regex = re.compile(pattern, re.IGNORECASE)
                except re.error as e:
                    print(f"正则表达式错误 '{pattern}': {e}")
                    continue

                # 搜索匹配的单元格
                matched_cells = []
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and regex.search(str(cell.value)):
                            matched_cells.append(cell)

                # 为每个匹配的单元格查找相邻值
                for cell in matched_cells:
                    for distance in range(1, max_distance + 1):
                        value_cell = None

                        if search_direction == 'right':
                            if cell.column + distance <= sheet.max_column:
                                value_cell = sheet.cell(row=cell.row, column=cell.column + distance)
                        elif search_direction == 'below':
                            if cell.row + distance <= sheet.max_row:
                                value_cell = sheet.cell(row=cell.row + distance, column=cell.column)

                        if value_cell and value_cell.value is not None:
                            # 检查是否需要数值
                            if require_numeric and not isinstance(value_cell.value, (int, float)):
                                # 尝试转换为数字
                                try:
                                    numeric_value = float(str(value_cell.value))
                                    results.append(numeric_value)
                                except (ValueError, TypeError):
                                    continue
                            else:
                                results.append(value_cell.value)
                            break

        except Exception as e:
            print(f"模式匹配查找时出错: {e}")

        return results

    def get_table_data_by_labels(self, sheet_name, row_labels, column_labels,
                                header_row=None, data_start_row=None):
        """
        根据行标签和列标签提取表格数据

        Args:
            sheet_name: 工作表名称
            row_labels: 行标签列表
            column_labels: 列标签列表
            header_row: 表头行位置
            data_start_row: 数据开始行

        Returns:
            字典格式的表格数据
        """
        if not self.workbook or self.file_type != 'excel':
            return {}

        try:
            sheet = self.workbook[sheet_name]
            if sheet_name not in self.workbook.sheetnames:
                return {}

            result = {}

            # 查找行标签位置
            row_positions = {}
            for label in row_labels:
                for row in range(1, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value and label.lower() in str(cell.value).lower():
                            row_positions[label] = col
                            break
                    if label in row_positions:
                        break

            # 查找列标签位置
            if header_row:
                col_positions = {}
                for label in column_labels:
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=header_row, column=col)
                        if cell.value and label.lower() in str(cell.value).lower():
                            col_positions[label] = col
                            break
            else:
                # 如果没有指定表头行，搜索整个工作表
                col_positions = {}
                for label in column_labels:
                    for row in range(1, sheet.max_row + 1):
                        for col in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row, column=col)
                            if cell.value and label.lower() in str(cell.value).lower():
                                col_positions[label] = col
                                break
                        if label in col_positions:
                            break

            # 提取数据
            data_start = data_start_row or 2
            for row_label, row_col in row_positions.items():
                result[row_label] = {}
                for col_label, col_num in col_positions.items():
                    # 从数据开始行向下查找
                    for row in range(data_start, sheet.max_row + 1):
                        if sheet.cell(row=row, column=row_col).value:  # 找到有数据的行
                            cell = sheet.cell(row=row, column=col_num)
                            result[row_label][col_label] = cell.value
                            break
                    else:
                        result[row_label][col_label] = None
            # Initialize Flags
            if 'flags' not in result:
                result['flags'] = {
                    'has_scope_1': False,
                    'has_scope_2_location': False,
                    'has_scope_2_market': False,
                    'has_scope_3': False,
                }

            # Update flags based on data
            result = self._update_flags(result)

            # Add quantification methods data
            # Use ReportConfig for quantification methods
            report_config = ReportConfig(
                company_name or '某公司',
                reporting_period or '2024年'
            )
            result['quantification_methods'] = report_config.get_quantification_methods()

            # Add scope 3 category names mapping
            # Use ReportConfig for scope 3 category names
            report_config_names = ReportConfig()
            result['scope_3_category_names'] = report_config_names.get_all_scope_3_category_names()


            return result

        except Exception as e:
            print(f"提取表格数据时出错: {e}")
            return {}

    def _parse_csv_table_section(self, csv_data, section_label, emission_type='scope1'):
        """
        解析CSV中特定区域的表格数据

        Args:
            csv_data: 原始CSV数据字典 {参数: 值}
            section_label: 区域标签（如"范围一直接排放源"）
            emission_type: 排放类型 ('scope1', 'scope2_3')

        Returns:
            包含name, emission, note的字典列表
        """
        import re
        items = []

        # 查找区域起始位置
        section_start = None
        for key in csv_data.keys():
            if section_label in str(key):
                section_start = key
                break

        if not section_start:
            return items

        # 解析该区域下的数据
        # 根据CSV结构，数据按照 "类别-排放源-具体排放源" 的层级组织
        current_category = None
        emission_sources = []

        # 范围一的类别映射
        scope1_categories = {
            '固定燃烧': '固定燃料燃烧',
            '移动燃烧': '移动燃料燃烧',
            '散逸排放': '散逸排放',
            '过程排放': '工艺过程排放'
        }

        # 范围二、三的类别映射
        scope2_3_categories = {
            '外购热力': '外购热力产生的排放',
            '时间序列': '外购热力产生的排放',
        }

        # 收集所有相关的排放源数据
        for key, value in csv_data.items():
            if not value or value.strip() == '':
                continue

            # 根据排放类型选择相应的类别
            if emission_type == 'scope1':
                for category, display_name in scope1_categories.items():
                    if category in str(key):
                        items.append({
                            'name': display_name,
                            'emission': 0.0,
                            'note': value
                        })
                        break
            else:  # scope2_3
                # 范围二、三的处理
                if '外购电力' in str(key) or '外购热力' in str(key):
                    items.append({
                        'name': value,  # 使用CSV中的值作为排放源名称
                        'emission': 0.0,
                        'note': '外购能源'
                    })
                elif '上游' in str(key) or '产生的排放' in str(key):
                    # 上游排放源
                    items.append({
                        'name': value,
                        'emission': 0.0,
                        'note': '上游排放'
                    })
                elif any(x in str(key) for x in [' purchased_goods', 'category_1', 'category_2']):
                    items.append({
                        'name': value,
                        'emission': 0.0,
                        'note': '商品和服务'
                    })

        return items

    def extract_data(self):
        """
        提取所有模板需要的数据，返回包含32个变量的字典。
        优先从CSV文件读取，如果CSV不存在则从Excel文件提取。
        """
        # 默认值字典 - 用于数据源中不存在的情况
        default_values = {
            'company_profile': '待补充公司简介信息',
            'legal_person': '待补充',
            'registered_address': '待补充注册地址',
            'date_of_establishment': '待补充',
            'registered_capital': '待补充',
            'Unified_Social_Credit_Identifier': '待补充',
            'deadline': '待补充',
            'evaluation_level': '待评估',
            'evaluation_score': '待评估',
            'scope_of_business': '待补充经营范围',
            'source_file': self.filepath if hasattr(self, 'filepath') else '未知',
            'GWP_Value_Reference_Document': '2021年IPCC第六次评估报告AR6',
            'rule_file': '企业温室气体排放核算与报告指南',
        }

        # 初始化数据字典
        data = {}
        data.update(default_values)

        # ========== 从CSV文件读取基本信息 ==========
        import os
        csv_path = '减排行动统计.csv'
        if os.path.exists(csv_path):
            # 只从CSV读取基本信息字段（不读取排放数据）
            try:
                import csv
                encodings = ['utf-8', 'gbk', 'gb2312']
                csv_basic_data = {}
                
                for encoding in encodings:
                    try:
                        with open(csv_path, 'r', encoding=encoding) as f:
                            csv_reader = csv.reader(f)
                            for row in csv_reader:
                                if len(row) >= 2:
                                    key = row[0].strip()
                                    value = row[1].strip()
                                    # 只读取基本信息字段
                                    if key in ['reporting_period', 'document_number', 'posted_time', 
                                              'legal_person', 'registered_capital', 'date_of_establishment',
                                              'registered_address', 'production_address', 'company_profile',
                                              'deadline', 'evaluation_level', 'evaluation_score',
                                              'Unified_Social_Credit_Identifier', 'GWP_Value_Reference_Document',
                                              'rule_file']:
                                        csv_basic_data[key] = value
                        print(f"成功从CSV读取基本信息 (编码: {encoding})")
                        break
                    except:
                        continue
                
                # 更新基本信息到data字典
                data.update(csv_basic_data)
                print(f"从CSV读取了 {len(csv_basic_data)} 个基本信息字段")
            except Exception as e:
                print(f"读取CSV基本信息时出错: {e}")

                # ========== 将排放数据转换为 float 类型（保持数据层纯净）==========
                # 格式化（加逗号、保留小数）是展示层（View Layer/Writer）的事
                def to_float(value, default=0.0):
                    """将值转换为 float 类型，失败则返回默认值"""
                    try:
                        return float(value)
                    except (ValueError, TypeError):
                        return default

                # 确保所有排放数据为 float 类型
                emission_keys = [
                    'scope_1_emissions',
                    'scope_2_location_based_emissions',
                    'scope_2_market_based_emissions',
                    'scope_3_emissions',
                    'scope_3_category_1_emissions',
                    'scope_3_category_2_emissions',
                    'scope_3_category_3_emissions',
                    'scope_3_category_4_emissions',
                    'scope_3_category_5_emissions',
                    'scope_3_category_6_emissions',
                    'scope_3_category_7_emissions',
                    'scope_3_category_9_emissions',
                    'scope_3_category_10_emissions',
                    'scope_3_category_12_emissions',
                ]
                for key in emission_keys:
                    if key in data:
                        data[key] = to_float(data[key])

                # ========== 构建表格数据列表（使用按区域解析的方法）==========
                section_data = self._parse_csv_sections(csv_path)

                # 获取范围一和范围二三的表格数据
                scope1_items = section_data.get('scope1_items', [])
                scope2_3_items_raw = section_data.get('scope2_3_items', [])

                # 构建最终的scope2_3_items列表
                # 1. 首先添加范围二的总量数据（如果有）
                scope2_3_items = []
                scope2_location = data.get('scope_2_location_based_emissions', 0.0)
                scope2_market = data.get('scope_2_market_based_emissions', 0.0)

                if scope2_location > 0:
                    scope2_3_items.append({
                        'name': '范围二：能源间接温室气体排放（基于位置）',
                        'emission': scope2_location,
                        'note': '外购电力和热力'
                    })
                if scope2_market > 0:
                    scope2_3_items.append({
                        'name': '范围二：能源间接温室气体排放（基于市场）',
                        'emission': scope2_market,
                        'note': '外购电力和热力'
                    })

                # 2. 添加范围三分类数据（如果有）
                scope3_total_items = [
                    ('外购商品和服务的上游产生的排放', 'scope_3_category_1_emissions', '原材料采购'),
                    ('资本货物产生的排放', 'scope_3_category_2_emissions', '设备设施建设'),
                    ('燃料和能源相关逸出排放', 'scope_3_category_3_emissions', '外购电力热力上游排放'),
                    ('上下游运输和配送产生的排放', 'scope_3_category_4_emissions', '物流运输'),
                    ('运营中产生的废弃物产生的排放', 'scope_3_category_5_emissions', '废弃物处理'),
                    ('员工商务差旅产生的排放', 'scope_3_category_6_emissions', '商务出行'),
                    ('员工上下班通勤产生的排放', 'scope_3_category_7_emissions', '员工通勤'),
                    ('运营中输入的运输和配送产生的排放', 'scope_3_category_9_emissions', '原材料和产品运输'),
                    ('已售产品的使用过程产生的排放', 'scope_3_category_10_emissions', '产品使用阶段'),
                    ('已售产品的报废处理产生的排放', 'scope_3_category_12_emissions', '产品回收处理'),
                ]

                for name, emission_key, note in scope3_total_items:
                    emission_value = data.get(emission_key, 0.0)
                    if emission_value > 0:
                        scope2_3_items.append({
                            'name': name,
                            'emission': emission_value,
                            'note': note
                        })

                # 3. 最后添加从CSV解析的详细排放源数据
                scope2_3_items.extend(scope2_3_items_raw)

                data['scope1_items'] = scope1_items
                data['scope2_3_items'] = scope2_3_items

                # 为了向后兼容，保留 items 列表（使用范围二三数据）
                data['items'] = scope2_3_items

                # ========== 键名映射：为AIService添加别名（用于AI摘要生成）==========
                # CSV使用的键名 -> AIService期望的键名
                key_mapping = {
                    'scope_1_emissions': 'scope_1',
                    'scope_2_location_based_emissions': 'scope_2_location',
                    'scope_2_market_based_emissions': 'scope_2_market',
                    'scope_3_emissions': 'scope_3',
                }

                for csv_key, ai_key in key_mapping.items():
                    if csv_key in data:
                        data[ai_key] = data[csv_key]

                # ========== 添加布尔值标记系统（Flags）==========
                # 这是为了让后续的程序能看懂"是否需要生成这一节"
                data['flags'] = {
                    'has_scope_1': data.get('scope_1_emissions', 0.0) > 0,
                    'has_scope_2_location': data.get('scope_2_location_based_emissions', 0.0) > 0,
                    'has_scope_2_market': data.get('scope_2_market_based_emissions', 0.0) > 0,
                    'has_scope_3': data.get('scope_3_emissions', 0.0) > 0,
                }

                # 计算总排放量（用于AI摘要）- 保持为 float 类型
                try:
                    s1 = data.get('scope_1_emissions', 0.0)
                    s2_loc = data.get('scope_2_location_based_emissions', 0.0)
                    s3 = data.get('scope_3_emissions', 0.0)
                    total_loc = s1 + s2_loc + s3

                    s2_mkt = data.get('scope_2_market_based_emissions', 0.0)
                    total_mkt = s1 + s2_mkt + s3

                    data['total_emission_location'] = total_loc
                    data['total_emission_market'] = total_mkt

                    # 提取年份
                    period = data.get('reporting_period', '')
                    import re
                    year_match = re.search(r'(\d{4})', str(period))
                    data['report_year'] = year_match.group(1) if year_match else '2024'

                    # 赋值给 self 变量，供 ReportConfig 使用
                    self.company_name = data.get('company_name')
                    self.reporting_period = data.get('reporting_period') or '2024年'
                except (ValueError, TypeError):
                    data['total_emission_location'] = 0.0
                    data['total_emission_market'] = 0.0
                    data['report_year'] = '2024'

                # ========== 从温室气体盘查表提取scope2_items（两行：基于位置和基于市场）==========
                # 和scope3_items（范围三详细排放源）
                if self.workbook and self.file_type == 'excel':
                    import openpyxl
                    import re

                    # 查找温室气体盘查表
                    pandata_sheet = None
                    for sheet in self.workbook.worksheets:
                        if '盘查表' in str(sheet.title):
                            pandata_sheet = sheet
                            break

                    if pandata_sheet:
                        print(f"从温室气体盘查表提取范围二和范围三数据")

                        # ========== 提取scope2_items（两行：基于位置和基于市场）==========
                        scope2_items = []
                        location_total = None
                        market_total = None
                        total_keywords = ['汇总', '总计', 'Total', 'TOTAL']

                        for row_idx, row in enumerate(pandata_sheet.iter_rows(min_row=1, values_only=True), start=1):
                            first_col = str(row[0]).strip() if row[0] else ''
                            second_col = str(row[1]).strip() if len(row) > 1 else ''
                            fourth_col = str(row[4]).strip() if len(row) > 4 else ''

                            is_total_row = any(keyword in first_col or first_col == keyword for keyword in total_keywords)
                            has_electricity = '外购' in second_col or '电力' in second_col

                            if is_total_row and has_electricity:
                                is_location = '位置' in fourth_col or 'Location' in fourth_col
                                is_market = '市场' in fourth_col or 'Market' in fourth_col

                                if is_location:
                                    location_total = row
                                elif is_market:
                                    market_total = row

                        # 构建scope2_items
                        if location_total:
                            total = self._safe_float(location_total[2]) if len(location_total) > 2 else 0
                            co2 = self._safe_float(location_total[3]) if len(location_total) > 3 else total
                            total_formatted = f"{total:,.2f}" if total > 0 else "0.00"
                            co2_formatted = f"{co2:,.2f}" if co2 > 0 else "0.00"
                            scope2_items.append({
                                'number': '2.1',
                                'emission_source': '外购电力（基于位置）',
                                'total_green_house_gas_emissions': total_formatted,
                                'CO2_emissions': co2_formatted,
                                'CH4_emissions': '0.00',
                                'N2O_emissions': '0.00',
                                'HFCs_emissions': '0.00',
                                'PFCs_emissions': '0.00',
                                'SFs_emissions': '0.00',
                                'NF3_emissions': '0.00'
                            })

                        if market_total:
                            total = self._safe_float(market_total[2]) if len(market_total) > 2 else 0
                            co2 = self._safe_float(market_total[3]) if len(market_total) > 3 else total
                            total_formatted = f"{total:,.2f}" if total > 0 else "0.00"
                            co2_formatted = f"{co2:,.2f}" if co2 > 0 else "0.00"
                            scope2_items.append({
                                'number': '2.2',
                                'emission_source': '外购电力（基于市场）',
                                'total_green_house_gas_emissions': total_formatted,
                                'CO2_emissions': co2_formatted,
                                'CH4_emissions': '0.00',
                                'N2O_emissions': '0.00',
                                'HFCs_emissions': '0.00',
                                'PFCs_emissions': '0.00',
                                'SFs_emissions': '0.00',
                                'NF3_emissions': '0.00'
                            })

                        data['scope2_items'] = scope2_items
                        print(f"提取到范围二排放明细: {len(scope2_items)} 行")

                        # ========== 提取scope3_items（范围三详细排放源）==========
                        category_names = {
                            '类别1': '外购商品和服务上游排放',
                            '类别2': '资本货物上游排放',
                            '类别3': '燃料和能源相关活动未包含在范围一和范围二中的上游排放',
                            '类别4': '上游运输和配送',
                            '类别5': '运营中产生的废弃物',
                            '类别6': '员工商务旅行',
                            '类别7': '员工通勤',
                            '类别8': '上游租赁资产',
                            '类别9': '下游运输和配送',
                            '类别10': '销售产品的加工',
                            '类别11': '销售产品的使用',
                            '类别12': '售出产品的加工',
                            '类别13': '下游租赁资产',
                            '类别14': '特许经营',
                            '类别15': '投资'
                        }

                        category_detail_rows = {}
                        for row_idx, row in enumerate(pandata_sheet.iter_rows(min_row=1, values_only=True), start=1):
                            row_vals = [str(v) if v is not None else '' for v in row[:12]]

                            if row_vals[0] and row_vals[0].isdigit():
                                if len(row_vals) > 4 and '范围三' in row_vals[4] and '类别' in row_vals[4]:
                                    category_match = re.search(r'类别\s*(\d+)', row_vals[4])
                                    if category_match:
                                        category_num = int(category_match.group(1))
                                        if category_num not in category_detail_rows:
                                            category_detail_rows[category_num] = []
                                        category_detail_rows[category_num].append({'row_idx': row_idx, 'data': row})

                        # 按类别分组存储范围三详细排放源
                        # 模板期望的数据格式：scope3_category1, scope3_category2, ..., scope3_category15
                        total_scope3_items = 0
                        for category_num in sorted(category_detail_rows.keys()):
                            detail_rows = category_detail_rows[category_num]
                            category_key = f'类别{category_num}'
                            category_var_name = f'scope3_category{category_num}'

                            category_items = []
                            sub_num = 0
                            for row_info in detail_rows:
                                row = row_info['data']
                                row_vals = [str(v) if v is not None else '' for v in row[:12]]

                                emission_source_name = row_vals[2]
                                activity_data = self._safe_float(row[5])
                                emission_factor = self._safe_float(row[7])
                                factor_unit = row_vals[8]

                                factor_in_tons = emission_factor
                                if 'kgCO2' in factor_unit:
                                    factor_in_tons = emission_factor / 1000
                                elif 'kg CO2' in factor_unit:
                                    factor_in_tons = emission_factor / 1000

                                calculated_emission = activity_data * factor_in_tons

                                if calculated_emission > 0.01:
                                    sub_num += 1
                                    total_formatted = f"{calculated_emission:,.2f}"
                                    co2_formatted = f"{calculated_emission:,.2f}"

                                    category_items.append({
                                        'number': f'3.{category_num}.{sub_num}',
                                        'emission_source': emission_source_name,
                                        'total_green_house_gas_emissions': total_formatted,
                                        'CO2_emissions': co2_formatted,
                                        'CH4_emissions': '0.00',
                                        'N2O_emissions': '0.00',
                                        'HFCs_emissions': '0.00',
                                        'PFCs_emissions': '0.00',
                                        'SFs_emissions': '0.00',
                                        'NF3_emissions': '0.00'
                                    })

                            # 将类别数据存储到对应的变量中
                            data[category_var_name] = category_items
                            total_scope3_items += len(category_items)
                            print(f"  提取{category_var_name}: {len(category_items)} 行")

                        # 同时保留scope3_items用于兼容
                        data['scope3_items'] = []
                        for cat_num in range(1, 16):
                            cat_var = f'scope3_category{cat_num}'
                            if cat_var in data:
                                data['scope3_items'].extend(data[cat_var])

                        print(f"提取到范围三详细排放明细总计: {total_scope3_items} 行")

                # ========== 基于特征指纹读取协议表格 ==========
                print("\n[协议读取] 开始基于特征指纹读取协议表格...")
                protocol_data = self.read_protocols()
                data.update(protocol_data)
                print(f"[协议读取] 成功读取 {len(protocol_data)} 个协议变量")
                for var_name, items in protocol_data.items():
                    print(f"  - {var_name}: {len(items)} 行")
                # ========== 协议表格读取结束 ==========

                return data

        # ========== 如果CSV不存在，使用Excel数据（向后兼容） ==========
        # 处理CSV文件（减排行动统计数据）
        if self.file_type == 'csv' and ('减排行动' in str(self.filepath) or 'GHG' in str(self.filepath)):
            emission_reductions = self.read_to_list_of_dicts(skip_empty_rows=True)
            data['emission_reductions'] = emission_reductions
            data['file_type'] = 'csv'
            print(f"从CSV文件提取减排行动数据，共 {len(emission_reductions)} 条记录")

            # CSV文件不支持协议表格，初始化空协议变量
            from data_reader import TABLE_PROTOCOLS
            for protocol_config in TABLE_PROTOCOLS.values():
                output_var = protocol_config['output_var']
                if output_var not in data:
                    data[output_var] = []

            return data

        # 处理Excel文件（温室气体排放数据）
        if not self.workbook or self.file_type != 'excel':
            # 非Excel文件不支持协议表格，初始化空协议变量
            from data_reader import TABLE_PROTOCOLS
            for protocol_config in TABLE_PROTOCOLS.values():
                output_var = protocol_config['output_var']
                if output_var not in data:
                    data[output_var] = []

            return data

        # ========== 从温室气体盘查表提取scope2_items和scope3_items ==========
        import openpyxl
        import re

        # 查找温室气体盘查表
        pandata_sheet = None
        for sheet in self.workbook.worksheets:
            if '盘查表' in str(sheet.title):
                pandata_sheet = sheet
                break

        if pandata_sheet:
            print(f"从温室气体盘查表提取范围二和范围三数据")

            # ========== 提取scope2_items（两行：基于位置和基于市场）==========
            scope2_items = []
            location_total = None
            market_total = None
            total_keywords = ['汇总', '总计', 'Total', 'TOTAL']

            for row_idx, row in enumerate(pandata_sheet.iter_rows(min_row=1, values_only=True), start=1):
                first_col = str(row[0]).strip() if row[0] else ''
                second_col = str(row[1]).strip() if len(row) > 1 else ''
                fourth_col = str(row[4]).strip() if len(row) > 4 else ''

                is_total_row = any(keyword in first_col or first_col == keyword for keyword in total_keywords)
                has_electricity = '外购' in second_col or '电力' in second_col

                if is_total_row and has_electricity:
                    is_location = '位置' in fourth_col or 'Location' in fourth_col
                    is_market = '市场' in fourth_col or 'Market' in fourth_col

                    if is_location:
                        location_total = row
                    elif is_market:
                        market_total = row

            # 构建scope2_items
            if location_total:
                total = self._safe_float(location_total[2]) if len(location_total) > 2 else 0
                co2 = self._safe_float(location_total[3]) if len(location_total) > 3 else total
                total_formatted = f"{total:,.2f}" if total > 0 else "0.00"
                co2_formatted = f"{co2:,.2f}" if co2 > 0 else "0.00"
                scope2_items.append({
                    'number': '2.1',
                    'emission_source': '外购电力（基于位置）',
                    'total_green_house_gas_emissions': total_formatted,
                    'CO2_emissions': co2_formatted,
                    'CH4_emissions': '0.00',
                    'N2O_emissions': '0.00',
                    'HFCs_emissions': '0.00',
                    'PFCs_emissions': '0.00',
                    'SFs_emissions': '0.00',
                    'NF3_emissions': '0.00'
                })

            if market_total:
                total = self._safe_float(market_total[2]) if len(market_total) > 2 else 0
                co2 = self._safe_float(market_total[3]) if len(market_total) > 3 else total
                total_formatted = f"{total:,.2f}" if total > 0 else "0.00"
                co2_formatted = f"{co2:,.2f}" if co2 > 0 else "0.00"
                scope2_items.append({
                    'number': '2.2',
                    'emission_source': '外购电力（基于市场）',
                    'total_green_house_gas_emissions': total_formatted,
                    'CO2_emissions': co2_formatted,
                    'CH4_emissions': '0.00',
                    'N2O_emissions': '0.00',
                    'HFCs_emissions': '0.00',
                    'PFCs_emissions': '0.00',
                    'SFs_emissions': '0.00',
                    'NF3_emissions': '0.00'
                })

            data['scope2_items'] = scope2_items
            print(f"提取到范围二排放明细: {len(scope2_items)} 行")

            # ========== 提取scope3_items（范围三详细排放源，按类别分组）==========
            category_names = {
                '类别1': '外购商品和服务上游排放',
                '类别2': '资本货物上游排放',
                '类别3': '燃料和能源相关活动未包含在范围一和范围二中的上游排放',
                '类别4': '上游运输和配送',
                '类别5': '运营中产生的废弃物',
                '类别6': '员工商务旅行',
                '类别7': '员工通勤',
                '类别8': '上游租赁资产',
                '类别9': '下游运输和配送',
                '类别10': '销售产品的加工',
                '类别11': '销售产品的使用',
                '类别12': '售出产品的加工',
                '类别13': '下游租赁资产',
                '类别14': '特许经营',
                '类别15': '投资'
            }

            category_detail_rows = {}
            for row_idx, row in enumerate(pandata_sheet.iter_rows(min_row=1, values_only=True), start=1):
                row_vals = [str(v) if v is not None else '' for v in row[:12]]

                if row_vals[0] and row_vals[0].isdigit():
                    if len(row_vals) > 4 and '范围三' in row_vals[4] and '类别' in row_vals[4]:
                        category_match = re.search(r'类别\s*(\d+)', row_vals[4])
                        if category_match:
                            category_num = int(category_match.group(1))
                            if category_num not in category_detail_rows:
                                category_detail_rows[category_num] = []
                            category_detail_rows[category_num].append({'row_idx': row_idx, 'data': row})

            # 按类别分组存储
            total_scope3_items = 0
            for category_num in sorted(category_detail_rows.keys()):
                detail_rows = category_detail_rows[category_num]
                category_var_name = f'scope3_category{category_num}'

                category_items = []
                sub_num = 0
                for row_info in detail_rows:
                    row = row_info['data']
                    row_vals = [str(v) if v is not None else '' for v in row[:12]]

                    emission_source_name = row_vals[2]
                    activity_data = self._safe_float(row[5])
                    emission_factor = self._safe_float(row[7])
                    factor_unit = row_vals[8]

                    factor_in_tons = emission_factor
                    if 'kgCO2' in factor_unit:
                        factor_in_tons = emission_factor / 1000
                    elif 'kg CO2' in factor_unit:
                        factor_in_tons = emission_factor / 1000

                    calculated_emission = activity_data * factor_in_tons

                    if calculated_emission > 0.01:
                        sub_num += 1
                        total_formatted = f"{calculated_emission:,.2f}"
                        co2_formatted = f"{calculated_emission:,.2f}"

                        category_items.append({
                            'number': f'3.{category_num}.{sub_num}',
                            'emission_source': emission_source_name,
                            'total_green_house_gas_emissions': total_formatted,
                            'CO2_emissions': co2_formatted,
                            'CH4_emissions': '0.00',
                            'N2O_emissions': '0.00',
                            'HFCs_emissions': '0.00',
                            'PFCs_emissions': '0.00',
                            'SFs_emissions': '0.00',
                            'NF3_emissions': '0.00'
                        })

                # 将类别数据存储到对应的变量中
                data[category_var_name] = category_items
                total_scope3_items += len(category_items)
                if category_items:
                    print(f"  提取{category_var_name}: {len(category_items)} 行")

            # 同时保留scope3_items用于兼容
            data['scope3_items'] = []
            for cat_num in range(1, 16):
                cat_var = f'scope3_category{cat_num}'
                if cat_var in data:
                    data['scope3_items'].extend(data[cat_var])

            print(f"提取到范围三详细排放明细总计: {total_scope3_items} 行")

        # 尝试多个可能的工作表名称
        main_sheet_candidates = ['温室气体盘查清册', '温室气体盘查清册 (2)']
        main_sheet = None
        for candidate in main_sheet_candidates:
            if candidate in self.workbook.sheetnames:
                main_sheet = candidate
                break

        if not main_sheet:
            print("警告：未找到主要工作表")
            return data

        table_sheet = '表1温室气体盘查表'
        
        # 使用新的find_value_by_label方法替代硬坐标定位
        # 从主要工作表中提取元数据
        company_name = self.find_value_by_label(main_sheet, '组织名称：') 
        report_period = self.find_value_by_label(main_sheet, '盘查覆盖周期:') 
        # 从报告周期中提取年份（假设格式为"2024年1月1日至2024年12月31日"）
        report_year = '2024'  # 直接提取年份

        # 赋值给 self 变量，供 ReportConfig 使用
        self.company_name = company_name
        self.reporting_period = report_period or '2024年'  # 默认值

        # 获取范围一排放量
        scope_1 = None
        try:
            sheet = self.workbook[table_sheet]
            # 遍历表格找到'总排放量'所在行
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == '总排放量':
                        # 根据用户反馈，总排放量这一行的数据与正上方单元格一一对应
                        # 因此我们需要获取当前行各列的值，然后将这些值与上一行的标签对应
                        # 这里我们主要关注范围一对应的排放量
                        current_row = cell.row
                        # 假设范围一的标签在B列（根据之前的调试发现）
                        # 检查上一行B列是否包含'范围一'
                        prev_row_cell_b = sheet.cell(row=current_row-1, column=2)
                        if prev_row_cell_b.value and '范围一' in str(prev_row_cell_b.value):
                            # 获取当前行B列的值作为scope_1
                            scope_1 = sheet.cell(row=current_row, column=2).value
                            print(f"从表1温室气体盘查表获取scope_1值(总排放量行上方对应范围一): {scope_1}")
                            break
                if scope_1 is not None:
                    break
            
            # 如果没找到，回退到使用find_value_by_label方法
            if scope_1 is None:
                scope_1 = self.find_value_by_label(table_sheet, '总排放量')
                print(f"回退到查找总排放量右侧值作为scope_1: {scope_1}")
        except Exception as e:
            print(f"获取scope_1值时出错: {e}")
        
        # 提取范围二排放量
        # 先尝试从scope2_items计算（更可靠的方法）
        if 'scope2_items' in data and len(data['scope2_items']) >= 2:
            try:
                # 第1项通常是外购电力（基于位置）
                scope_2_location_str = data['scope2_items'][0].get('total_green_house_gas_emissions', '0')
                scope_2_location = float(str(scope_2_location_str).replace(',', '').replace(' ', ''))
                
                # 第2项通常是外购热力（基于市场）
                scope_2_market_str = data['scope2_items'][1].get('total_green_house_gas_emissions', '0')
                scope_2_market = float(str(scope_2_market_str).replace(',', '').replace(' ', ''))
                
                print(f"从scope2_items计算范围二排放:")
                print(f"  基于位置: {scope_2_location:,.2f}")
                print(f"  基于市场: {scope_2_market:,.2f}")
            except Exception as e:
                print(f"从scope2_items计算范围二失败: {e}")
                # 如果失败，使用原始方法
                scope_2_location = self.find_value_by_label(table_sheet, '基于位置')
                scope_2_market = self.find_value_by_label(table_sheet, '基于市场')
        else:
            # 回退到原始方法
            scope_2_location = self.find_value_by_label(table_sheet, '基于位置')
            scope_2_market = self.find_value_by_label(table_sheet, '基于市场')
        
        # 提取范围三排放量
        scope_3 = None
        try:
            # 优先使用find_value_by_label方法
            scope_3 = self.find_value_by_label(table_sheet, '范围三')
            
            if scope_3 is None:
                # 如果直接查找失败，尝试查找包含"范围三"的单元格
                sheet = self.workbook[table_sheet]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and '范围三' in str(cell.value):
                            # 检查右侧和下方的单元格
                            right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                            below_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                            
                            # 优先尝试右侧单元格（如果是数值或总量）
                            if right_cell.value is not None:
                                if isinstance(right_cell.value, (int, float)):
                                    scope_3 = right_cell.value
                                elif right_cell.value == '总量':
                                    # 如果右侧是总量，获取总量下方的值
                                    total_below = sheet.cell(row=right_cell.row + 1, column=right_cell.column)
                                    if total_below.value is not None:
                                        scope_3 = total_below.value
                            # 如果右侧没有找到，尝试下方单元格
                            elif below_cell.value is not None and isinstance(below_cell.value, (int, float)):
                                scope_3 = below_cell.value
                            break
                    if scope_3 is not None:
                        break
        except Exception as e:
            print(f"查找范围三数据时出错: {e}") 
        
        # 提取总排放量（基于位置）和总排放量（基于市场）
        # 使用find_value_by_label方法替代硬坐标
        total_emission_location = None
        total_emission_market = None
        try:
            # 首先计算预期的总排放量范围，用于验证找到的值是否合理
            expected_total_location = None
            expected_total_market = None
            if scope_1 is not None and scope_2_location is not None and scope_3 is not None:
                expected_total_location = float(scope_1) + float(scope_2_location) + float(scope_3)
                expected_total_market = float(scope_1) + float(scope_2_market) + float(scope_3)
                print(f"预期总排放量范围: 位置={expected_total_location}, 市场={expected_total_market}")
            
            # 使用find_value_by_label方法查找总排放量
            total_emission_location = self.find_value_by_label(table_sheet, '总排放量')
            
            # 尝试查找包含'基于位置'和'总量'的区域
            if total_emission_location is None:
                sheet = self.workbook[table_sheet]
                # 记录所有可能的候选值
                potential_totals = []
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_text = str(cell.value)
                            # 查找包含'总量'的单元格
                            if '总量' in cell_text:
                                # 记录这个位置附近的所有数值
                                for r_offset in range(-4, 5):
                                    for c_offset in range(-4, 5):
                                        check_row = cell.row + r_offset
                                        check_col = cell.column + c_offset
                                        if 1 <= check_row <= sheet.max_row and 1 <= check_col <= sheet.max_column:
                                            check_cell = sheet.cell(row=check_row, column=check_col)
                                            if check_cell.value is not None and isinstance(check_cell.value, (int, float)):
                                                # 检查这个数值是否接近预期的总排放量
                                                is_large_value = check_cell.value > 1000000  # 假设总排放量大于100万
                                                
                                                # 检查是否有'基于位置'或'基于市场'的标注
                                                has_location = False
                                                has_market = False
                                                for context_r in range(max(1, check_row-3), min(sheet.max_row, check_row+4)):
                                                    for context_c in range(max(1, check_col-5), min(sheet.max_column, check_col+6)):
                                                        context_cell = sheet.cell(row=context_r, column=context_c)
                                                        if context_cell.value is not None:
                                                            context_text = str(context_cell.value)
                                                            if '基于位置' in context_text:
                                                                has_location = True
                                                            elif '基于市场' in context_text:
                                                                has_market = True
                                                
                                                # 添加到候选列表
                                                if is_large_value and (has_location or has_market):
                                                    potential_totals.append({
                                                        'value': check_cell.value,
                                                        'row': check_row,
                                                        'col': check_col,
                                                        'is_location': has_location,
                                                        'is_market': has_market
                                                    })
                
                # 从候选值中选择最接近预期值的
                if potential_totals:
                    # 按与预期值的接近程度排序
                    if expected_total_location is not None:
                        potential_totals.sort(key=lambda x: abs(x['value'] - expected_total_location) if x['is_location'] else float('inf'))
                    # 选择第一个合适的基于位置的候选值
                    for candidate in potential_totals:
                        if candidate['is_location']:
                            total_emission_location = candidate['value']
                            print(f"从候选值中选择总排放量（基于位置）在第{candidate['row']}行第{candidate['col']}列: {candidate['value']}")
                            break
                    
                    # 选择第一个合适的基于市场的候选值
                    for candidate in potential_totals:
                        if candidate['is_market']:
                            total_emission_market = candidate['value']
                            print(f"从候选值中选择总排放量（基于市场）在第{candidate['row']}行第{candidate['col']}列: {candidate['value']}")
                            break
            
            # 方法3: 如果仍然找不到，直接使用计算值
            if total_emission_location is None and expected_total_location is not None:
                total_emission_location = expected_total_location
                print(f"使用计算值作为总排放量（基于位置）: {total_emission_location}")
            
            if total_emission_market is None and expected_total_market is not None:
                total_emission_market = expected_total_market
                print(f"使用计算值作为总排放量（基于市场）: {total_emission_market}")
        except Exception as e:
            print(f"获取总排放量时出错: {e}")
        
        # 将所有数据打包，保留之前添加的scope2_items、scope3_category1等变量

        # 计算范围三各类别的排放量（用于模板显示）
        for i in range(1, 16):
            cat_key = f'scope3_category{i}'
            if cat_key in data and data[cat_key]:
                # 计算该类别的总排放量
                total_emission = 0.0
                for item in data[cat_key]:
                    emission_str = item.get('total_green_house_gas_emissions', '0')
                    # 移除逗号和空格
                    emission_str = str(emission_str).replace(',', '').replace(' ', '')
                    try:
                        emission = float(emission_str)
                        total_emission += emission
                    except:
                        pass
                
                # 添加到data字典
                data[f'scope_3_category_{i}_emissions'] = total_emission
                
                if i <= 3:  # 只打印前3个
                    print(f"  category_{i} 总排放量: {total_emission:,.2f}")
        data['company_name'] = company_name
        data['report_year'] = report_year
        data['scope_1'] = scope_1  # 范围一排放量
        data['scope_2_location'] = scope_2_location  # 范围二排放量（基于位置）
        data['scope_2_market'] = scope_2_market      # 范围二排放量（基于市场）
        data['scope_3'] = scope_3                     # 范围三排放量
        data['total_emission_location'] = total_emission_location  # 总排放量（基于位置）
        data['total_emission_market'] = total_emission_market        # 总排放量（基于市场）
        data['file_type'] = 'excel' 
        
        print(f"数据提取完成: {data}") 
        
        # Initialize Flags
        if 'flags' not in data:
            data['flags'] = {
                'has_scope_1': False,
                'has_scope_2_location': False,
                'has_scope_2_market': False,
                'has_scope_3': False,
            }

        # Update flags based on data
        data = self._update_flags(data)

        # Add quantification methods data
        # Use ReportConfig for quantification methods
        report_config = ReportConfig(
            self.company_name or '某公司',
            self.reporting_period or '2024年'
        )
        data['quantification_methods'] = report_config.get_quantification_methods()

        # Add scope 3 category names mapping
        # Use ReportConfig for scope 3 category names
        report_config_names = ReportConfig()
        data['scope_3_category_names'] = report_config_names.get_all_scope_3_category_names()
        # 确保表格数据存在（模板需要这些字段）
        if 'scope1_items' not in data or 'scope2_3_items' not in data:
            print("警告：scope1_items或scope2_3_items缺失，尝试从scope2_items生成...")

            # 如果scope1_items缺失，创建空的列表
            if 'scope1_items' not in data:
                data['scope1_items'] = []

            # 如果scope2_3_items缺失，使用scope2_items创建
            if 'scope2_3_items' not in data and 'scope2_items' in data:
                data['scope2_3_items'] = data['scope2_items']

        # ========== 基于特征指纹读取协议表格 ==========
        print("\n[协议读取] 开始基于特征指纹读取协议表格...")
        protocol_data = self.read_protocols()
        data.update(protocol_data)
        print(f"[协议读取] 成功读取 {len(protocol_data)} 个协议变量")
        for var_name, items in protocol_data.items():
            print(f"  - {var_name}: {len(items)} 行")
        # ========== 协议表格读取结束 ==========

        return data
    def extract_data_from_xlsx_dynamic(self, xlsx_path=None):
        """
        纯 xlsx 数据源动态提取数据(不依赖固定行号，不使用 CSV)
        对于缺失的数据，使用 None

        Args:
            xlsx_path: xlsx 文件路径，如果为 None 则使用初始化时的 filepath

        Returns:
            包含所有模板变量的字典
        """
        import openpyxl
        import re

        filepath = xlsx_path or self.filepath

        # 如果传入的是当前 workbook，直接使用
        if self.workbook and self.file_type == 'excel' and (xlsx_path is None or xlsx_path == self.filepath):
            wb = self.workbook
        else:
            wb = openpyxl.load_workbook(filepath, data_only=True)

        # 默认值（缺失数据为 None）
        data = {
            'company_profile': None,
            'legal_person': None,
            'registered_address': None,
            'date_of_establishment': None,
            'registered_capital': None,
            'Unified_Social_Credit_Identifier': None,
            'deadline': None,
            'evaluation_level': None,
            'evaluation_score': None,
            'scope_of_business': None,
            'rule_file': None,
            'GWP_Value_Reference_Document': None,
            'document_number': None,
            'posted_time': None,
            'production_address': None,
            'source_file': filepath,
            'scope_3_category_1_emissions': 0.0,
            'scope_3_category_2_emissions': 0.0,
            'scope_3_category_3_emissions': 0.0,
            'scope_3_category_4_emissions': 0.0,
            'scope_3_category_5_emissions': 0.0,
            'scope_3_category_6_emissions': 0.0,
            'scope_3_category_7_emissions': 0.0,
            'scope_3_category_8_emissions': 0.0,
            'scope_3_category_9_emissions': 0.0,
            'scope_3_category_10_emissions': 0.0,
            'scope_3_category_11_emissions': 0.0,
            'scope_3_category_12_emissions': 0.0,
            'scope_3_category_13_emissions': 0.0,
            'scope_3_category_14_emissions': 0.0,
            'scope_3_category_15_emissions': 0.0,
        }

        # ========== 基本信息（从"基本信息"工作表读取） ==========
        from datetime import datetime

        # 辅助函数：转换Excel日期序列号为日期字符串
        def excel_date_to_string(date_value):
            """将Excel日期序列号转换为 'YYYY年MM月DD日' 格式"""
            if date_value is None:
                return None
            if isinstance(date_value, str):
                return date_value
            try:
                # Excel日期基准是1899-12-30
                delta = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_value) - 2)
                return delta.strftime('%Y年%m月%d日')
            except (ValueError, TypeError):
                return str(date_value)

        # 检查是否存在"基本信息"工作表
        if '基本信息' in wb.sheetnames:
            basic_info_sheet = wb['基本信息']
            print("找到'基本信息'工作表，开始读取...")

            # 读取基本信息：第2列是属性代码(key)，第3列是值(value)
            for row in basic_info_sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3 and row[1] and row[2]:
                    key = str(row[1]).strip()  # 第2列：属性代码
                    value = row[2]  # 第3列：值

                    # 处理日期字段
                    if key in ['posted_time', 'date_of_establishment']:
                        data[key] = excel_date_to_string(value)
                    # 处理长文本字段（公司简介、经营范围等）：将换行符替换为空格，避免在Word中分段
                    elif key in ['company_profile', 'scope_of_business']:
                        if isinstance(value, str):
                            # 将各种换行符统一替换为空格
                            # 保留段落间的空格，移除多余的换行
                            value = re.sub(r'[\n\r]+', ' ', value)
                            # 移除多余的空格
                            value = re.sub(r'\s+', ' ', value).strip()
                        data[key] = value
                    # 处理其他字段
                    elif key in data:
                        data[key] = value

                    # 特殊字段映射
                    if key == 'company_name':
                        data['company_name'] = value
                    elif key == 'reporting_period':
                        data['reporting_period'] = value
                        # 从周期中提取年份
                        year_match = re.search(r'(\d{4})', str(value))
                        data['report_year'] = year_match.group(1) if year_match else '2024'

            # 赋值给 self 变量，供 ReportConfig 使用
            self.company_name = data.get('company_name')
            self.reporting_period = data.get('reporting_period') or '2024年'

            print(f"从'基本信息'工作表读取完成，公司名称: {data.get('company_name')}")
        else:
            # 如果没有"基本信息"工作表，使用原来的方法
            main_sheet = wb['温室气体盘查清册 (2)']
            for row in main_sheet.iter_rows(max_row=20, values_only=True):
                if len(row) >= 3:
                    if row[1] == '组织名称：' and row[2]:
                        data['company_name'] = row[2]
                    elif row[1] == '组织地址：' and row[2]:
                        data['registered_address'] = row[2]
                        data['production_address'] = row[2]
                    elif row[1] == '盘查覆盖周期:' and row[2]:
                        data['reporting_period'] = row[2]

            # 从周期中提取年份
            if data.get('reporting_period'):
                year_match = re.search(r'(\d{4})', str(data['reporting_period']))
                data['report_year'] = year_match.group(1) if year_match else '2024'
            else:
                data['report_year'] = '2024'

            # 赋值给 self 变量，供 ReportConfig 使用
            self.company_name = data.get('company_name')
            self.reporting_period = data.get('reporting_period') or '2024年'

        # 确保 report_year 存在
        if 'report_year' not in data:
            data['report_year'] = '2024'

        # ========== 排放数据（动态读取） ==========
        table_sheet = wb['表1温室气体盘查表']

        # 动态查找总排放量汇总行
        for row in table_sheet.iter_rows(values_only=True):
            a_val = row[0] if len(row) > 0 else None
            b_val = row[1] if len(row) > 1 else None
            c_val = row[2] if len(row) > 2 else None
            d_val = row[3] if len(row) > 3 else None
            e_val = row[4] if len(row) > 4 else None

            if a_val and isinstance(a_val, str) and '排放量' in a_val:
                if isinstance(b_val, (int, float)) and isinstance(c_val, (int, float)) and isinstance(d_val, (int, float)):
                    data['scope_1_emissions'] = float(b_val)
                    data['scope_2_location_based_emissions'] = float(c_val)
                    data['scope_3_emissions'] = float(d_val)
                    if isinstance(e_val, (int, float)):
                        data['total_emission_location'] = float(e_val)
                break

        # 动态查找范围二基于市场的排放量
        # 查找E列标注"基于市场"的行，取C列值
        for row in table_sheet.iter_rows():
            e_val = row[4].value if len(row) > 4 else None
            c_val = row[2].value if len(row) > 2 else None
            if e_val and isinstance(e_val, str) and '基于市场' in e_val:
                if c_val and isinstance(c_val, (int, float)):
                    data['scope_2_market_based_emissions'] = float(c_val)
                    print(f"找到范围二基于市场排放量: {float(c_val)} (行{row[0].row})")
                break

        # 计算总排放量（基于市场）
        data['total_emission_market'] = (
            data.get('scope_1_emissions', 0) +
            data.get('scope_2_market_based_emissions', 0) +
            data.get('scope_3_emissions', 0)
        )

        # ========== 范围三分类数据（动态查找） ==========
        scope3_mapping = {
            '类别1': 'scope_3_category_1_emissions',
            '类别2': 'scope_3_category_2_emissions',
            '类别3': 'scope_3_category_3_emissions',
            '类别4': 'scope_3_category_4_emissions',
            '类别5': 'scope_3_category_5_emissions',
            '类别6': 'scope_3_category_6_emissions',
            '类别7': 'scope_3_category_7_emissions',
            '类别8': 'scope_3_category_8_emissions',
            '类别9': 'scope_3_category_9_emissions',
            '类别10': 'scope_3_category_10_emissions',
            '类别11': 'scope_3_category_11_emissions',
            '类别12': 'scope_3_category_12_emissions',
            '类别13': 'scope_3_category_13_emissions',
            '类别14': 'scope_3_category_14_emissions',
            '类别15': 'scope_3_category_15_emissions',
        }

        for category_key, var_name in scope3_mapping.items():
            for row in table_sheet.iter_rows():
                a_val = row[0].value if len(row) > 0 else None
                if a_val and isinstance(a_val, str) and '范围三' in a_val and category_key in a_val:
                    current_row_num = row[0].row
                    if current_row_num + 2 <= table_sheet.max_row:
                        emission_row = table_sheet[current_row_num + 2]
                        b_val = emission_row[1].value if len(emission_row) > 1 else None
                        if b_val and isinstance(b_val, (int, float)):
                            data[var_name] = float(b_val)
                    break

        # ========== 添加 flags 标记 ==========
        data['flags'] = {
            'has_scope_1': data.get('scope_1_emissions', 0) > 0,
            'has_scope_2_location': data.get('scope_2_location_based_emissions', 0) > 0,
            'has_scope_2_market': data.get('scope_2_market_based_emissions', 0) > 0,
            'has_scope_3': data.get('scope_3_emissions', 0) > 0,
        }

        # ========== 别名 ==========
        data['scope_1'] = data.get('scope_1_emissions', 0)
        data['scope_2_location'] = data.get('scope_2_location_based_emissions', 0)
        data['scope_2_market'] = data.get('scope_2_market_based_emissions', 0)
        data['scope_3'] = data.get('scope_3_emissions', 0)

        # ========== 从表1温室气体盘查表提取表1和表2的数据 ==========
        try:
            # 查找表1温室气体盘查表
            sheet1_data = None
            for sheet in wb.worksheets:
                if '表1' in sheet.title and '温室气体盘查表' in sheet.title:
                    sheet1_data = sheet
                    break

            if sheet1_data:
                print(f"找到表1温室气体盘查表: {sheet1_data.title}")
                scope1_table_items = []  # 表1：范围一直接排放源
                scope2_3_table_items = []  # 表2：范围二三间接排放源

                # 维护类别变量用于前向填充（ffill）
                current_category = ""

                # 从第5行开始（前4行是标题）
                for row in sheet1_data.iter_rows(min_row=5):
                    if len(row) < 7:
                        continue

                    # 获取各列数据
                    current_row = row[0].row
                    seq = row[0].value  # 序号
                    ghg_category = row[1].value  # GHG排放类别（第2列，索引为1）
                    emission_source = row[2].value  # 排放源
                    facility = row[3].value  # 设施
                    boundary = row[4].value  # 组织边界

                    # 实现前向填充逻辑（ffill）：如果当前行的类别为空，使用上一行的类别
                    if ghg_category:
                        current_category = str(ghg_category).strip()
                    # 即使 ghg_category 为空，也使用 current_category

                    # 跳过空行或标题行
                    if not seq and not current_category:
                        continue

                    seq_str = str(seq).strip() if seq else ''
                    # 使用 current_category（前向填充后的类别）
                    ghg_str = current_category if current_category else ''
                    source_str = str(emission_source).strip() if emission_source else ''
                    facility_str = str(facility).strip() if facility else ''
                    boundary_str = str(boundary).strip() if boundary else ''

                    # 跳过标题行
                    if seq_str == '序号' or ghg_str == 'GHG排放类别':
                        continue

                    # 表1：范围一
                    if '范围一' in boundary_str:
                        scope1_table_items.append({
                            'name': ghg_str,  # GHG排放类别（使用前向填充的值）
                            'number': seq_str,  # 序号
                            'emission_source': source_str,  # 排放源
                            'facility': facility_str  # 设施
                        })

                    # 表2：范围二三
                    elif '范围二' in boundary_str or '范围三' in boundary_str:
                        scope2_3_table_items.append({
                            'name': ghg_str,  # GHG排放类别（使用前向填充的值）
                            'number': seq_str,  # 序号
                            'emission_source': source_str,  # 排放源
                            'facility': facility_str  # 设施
                        })

                data['scope1_items'] = scope1_table_items
                data['scope2_3_items'] = scope2_3_table_items
                print(f"从表1温室气体盘查表提取范围一数据: {len(scope1_table_items)} 行")
                print(f"从表1温室气体盘查表提取范围二三数据: {len(scope2_3_table_items)} 行")

        except Exception as e:
            print(f"从表1温室气体盘查表提取数据时出错: {e}")
            import traceback
            traceback.print_exc()

        # ========== 从温室气体盘查清册表提取范围一详细表数据 ==========
        try:
            # 查找温室气体盘查清册表
            inventory_sheet = None
            for sheet in wb.worksheets:
                if '盘查清册' in sheet.title:
                    inventory_sheet = sheet
                    break

            if inventory_sheet:
                print(f"找到温室气体盘查清册表: {inventory_sheet.title}")
                scope1_detail_items = []  # 范围一详细表数据（从盘查清册提取）

                # 从第14行开始（第12行是标题，第13行是单位）
                # 维护当前类别用于前向填充（ffill）
                current_category = ""
                current_sub_category = ""

                for row in inventory_sheet.iter_rows(min_row=14):
                    if len(row) < 13:
                        continue

                    # Excel结构：A=空, B=编号/类别名, C=排放源, D=排放设施, E=备注, F=总排放量, G=CO2, H=CH4, I=N2O, J=HFCs, K=PFCs, L=SF6, M=NF3
                    # 注意：当是类别行时，B列包含类别名（如"范围一 直接排放"），C列为空
                    #       当是数据行时，B列包含编号（如1.1），C列包含排放源名称
                    col_b = row[1].value        # 编号或类别名
                    col_c = row[2].value        # 排放源（仅数据行有值）
                    facility = row[3].value      # 排放设施 (列D)
                    note = row[4].value          # 备注 (列E)
                    total_emission = row[5].value    # 总排放量 (列F)
                    co2_emission = row[6].value      # CO2排放量 (列G)
                    ch4_emission = row[7].value      # CH4排放量 (列H)
                    n2o_emission = row[8].value      # N2O排放量 (列I)
                    hfcs_emission = row[9].value     # HFCs排放量 (列J)
                    pfcs_emission = row[10].value    # PFCs排放量 (列K)
                    sf6_emission = row[11].value     # SF6排放量 (列L)
                    nf3_emission = row[12].value     # NF3排放量 (列M)

                    # 跳过空行
                    if not col_b and not col_c:
                        continue

                    # 确定编号和排放源
                    number_str = ''
                    source_str = ''

                    if col_b:
                        col_b_str = str(col_b).strip()
                        # 检查B列是否是编号格式（如"1.1", "1.1.1"）- 以数字开头
                        if col_b_str and col_b_str[0].isdigit():
                            number_str = col_b_str
                            source_str = str(col_c).strip() if col_c else ''
                        # B列是类别名称（如"范围一 直接排放"）
                        else:
                            # 更新当前类别（用于前向填充）
                            current_category = col_b_str
                            # 跳过类别标题行，但继续处理后续行
                            continue

                    facility_str = str(facility).strip() if facility else ''
                    note_str = str(note).strip() if note else ''

                    # 跳过标题行
                    if not number_str or number_str == '编号':
                        continue

                    # 格式化排放量数字（保留两位小数）
                    def format_emission(val):
                        if val is None:
                            return ''
                        if val == 0:
                            return "0.00"
                        try:
                            float_value = float(val)
                            if float_value == 0:
                                return "0.00"
                            return f"{float_value:.2f}"
                        except (ValueError, TypeError):
                            return '0.00'

                    # 范围一：编号以1开头（如1.1, 1.1.1）
                    if number_str.startswith('1.'):
                        # 确定子类别（根据编号前缀判断）
                        if number_str.startswith('1.1.'):
                            current_sub_category = '固定源燃烧'
                        elif number_str.startswith('1.2.'):
                            current_sub_category = '移动源燃烧'
                        elif number_str.startswith('1.3.'):
                            current_sub_category = '遗散源'
                        elif number_str.startswith('1.4.'):
                            current_sub_category = '工艺排放'

                        # 使用编号作为主标识，但保留类别信息
                        scope1_detail_items.append({
                            'name': current_sub_category or number_str,  # 使用子类别名称
                            'number': number_str,
                            'category': current_category,  # 添加类别字段
                            'emission_source': source_str,
                            'facility': facility_str,
                            'note': note_str,
                            'total_green_house_gas_emissions': format_emission(total_emission),
                            'CO2_emissions': format_emission(co2_emission),
                            'CH4_emissions': format_emission(ch4_emission),
                            'N2O_emissions': format_emission(n2o_emission),
                            'HFCs_emissions': format_emission(hfcs_emission),
                            'PFCs_emissions': format_emission(pfcs_emission),
                            'SFs_emissions': format_emission(sf6_emission),
                            'NF3_emissions': format_emission(nf3_emission)
                        })

                # 分类范围一数据
                scope1_stationary_combustion = []  # 固定源燃烧（1.1.x）
                scope1_mobile_combustion = []      # 移动源燃烧（1.2.x）
                scope1_fugitive = []                # 遗散源（1.3.x）
                scope1_process = []                 # 工艺排放（如果有的话）

                for item in scope1_detail_items:
                    number = item.get('number', '')
                    if number.startswith('1.1.'):
                        scope1_stationary_combustion.append(item)
                    elif number.startswith('1.2.'):
                        scope1_mobile_combustion.append(item)
                    elif number.startswith('1.3.'):
                        scope1_fugitive.append(item)
                    elif number.startswith('1.4.'):
                        scope1_process.append(item)

                # 注意：scope1_items和scope2_3_items已经从"表1温室气体盘查表"提取，不再覆盖

                # 添加分类列表（注意：模板中使用的是 _items 后缀）
                data['scope1_stationary_combustion_emissions_items'] = scope1_stationary_combustion
                data['scope1_mobile_combustion_emissions_items'] = scope1_mobile_combustion
                data['scope1_fugitive_emissions_items'] = scope1_fugitive
                data['scope1_process_emissions_items'] = scope1_process

                # 计算各类别的汇总值
                emission_columns = [
                    'total_green_house_gas_emissions',
                    'CO2_emissions',
                    'CH4_emissions',
                    'N2O_emissions',
                    'HFCs_emissions',
                    'PFCs_emissions',
                    'SFs_emissions',
                    'NF3_emissions'
                ]

                # 先计算各分类的汇总值
                # 固定源燃烧汇总
                for col in emission_columns:
                    total = 0.0
                    for item in scope1_stationary_combustion:
                        emission_str = item.get(col, '0')
                        if emission_str and emission_str.strip():
                            try:
                                emission_str = emission_str.replace(',', '').replace(' ', '')
                                total += float(emission_str)
                            except (ValueError, TypeError):
                                pass
                    data[f'scope1_stationary_combustion_emissions_{col}_sum_formatted'] = f"{total:.2f}"

                # 移动源燃烧汇总
                for col in emission_columns:
                    total = 0.0
                    for item in scope1_mobile_combustion:
                        emission_str = item.get(col, '0')
                        if emission_str and emission_str.strip():
                            try:
                                emission_str = emission_str.replace(',', '').replace(' ', '')
                                total += float(emission_str)
                            except (ValueError, TypeError):
                                pass
                    data[f'scope1_mobile_combustion_emissions_{col}_sum_formatted'] = f"{total:.2f}"

                # 遗散源汇总
                for col in emission_columns:
                    total = 0.0
                    for item in scope1_fugitive:
                        emission_str = item.get(col, '0')
                        if emission_str and emission_str.strip():
                            try:
                                emission_str = emission_str.replace(',', '').replace(' ', '')
                                total += float(emission_str)
                            except (ValueError, TypeError):
                                pass
                    data[f'scope1_fugitive_emissions_{col}_sum_formatted'] = f"{total:.2f}"

                # 工艺排放汇总
                for col in emission_columns:
                    total = 0.0
                    for item in scope1_process:
                        emission_str = item.get(col, '0')
                        if emission_str and emission_str.strip():
                            try:
                                emission_str = emission_str.replace(',', '').replace(' ', '')
                                total += float(emission_str)
                            except (ValueError, TypeError):
                                pass
                    data[f'scope1_process_emissions_{col}_sum_formatted'] = f"{total:.2f}"

                # 范围一总计 = 各分类汇总的和（避免重复计算类别行的排放量）
                for col in emission_columns:
                    stationary_total = float(data[f'scope1_stationary_combustion_emissions_{col}_sum_formatted'])
                    mobile_total = float(data[f'scope1_mobile_combustion_emissions_{col}_sum_formatted'])
                    fugitive_total = float(data[f'scope1_fugitive_emissions_{col}_sum_formatted'])
                    process_total = float(data[f'scope1_process_emissions_{col}_sum_formatted'])
                    total = stationary_total + mobile_total + fugitive_total + process_total
                    data[f'scope1_emissions_{col}_sum_formatted'] = f"{total:.2f}"

                print(f"从温室气体盘查清册提取范围一详细表数据: {len(scope1_detail_items)} 行")
                print(f"  固定源燃烧: {len(scope1_stationary_combustion)} 行")
                print(f"  移动源燃烧: {len(scope1_mobile_combustion)} 行")
                print(f"  遗散源: {len(scope1_fugitive)} 行")
            else:
                print("警告：未找到温室气体盘查清册表")

        except Exception as e:
            print(f"从温室气体盘查清册提取数据时出错: {e}")
            import traceback
            traceback.print_exc()

        # ========== 从温室气体盘查表提取scope2_items（两行：基于位置和基于市场）==========
        try:
            # 查找温室气体盘查表
            pandata_sheet = None
            for sheet in wb.worksheets:
                if '盘查表' in str(sheet.title):
                    pandata_sheet = sheet
                    break

            if pandata_sheet:
                print(f"找到温室气体盘查表: {pandata_sheet.title}")
                scope2_items = []

                # 查找包含"汇总"和"外购电力"的行
                location_total = None
                market_total = None

                # 尝试匹配包含"汇总"或"Total"的行
                total_keywords = ['汇总', '总计', 'Total', 'TOTAL', '\u6c47\u603b', '\u603b\u8ba1']

                for row_idx, row in enumerate(pandata_sheet.iter_rows(min_row=1, values_only=True), start=1):
                    first_col = str(row[0]).strip() if row[0] else ''
                    second_col = str(row[1]).strip() if len(row) > 1 else ''
                    fourth_col = str(row[4]).strip() if len(row) > 4 else ''

                    # 检查是否是汇总行（使用多种方式匹配）
                    is_total_row = any(keyword in first_col or first_col == keyword for keyword in total_keywords)
                    has_electricity = '外购' in second_col or '电力' in second_col or 'Purchased' in second_col or 'Electricity' in second_col

                    if is_total_row and has_electricity:
                        # 检查第四列来判断是基于位置还是基于市场
                        is_location = '位置' in fourth_col or 'Location' in fourth_col or 'location' in fourth_col
                        is_market = '市场' in fourth_col or 'Market' in fourth_col or 'market' in fourth_col

                        if is_location:
                            location_total = row
                            print(f"  找到外购电力（基于位置）汇总: Row {row_idx}, fourth_col='{fourth_col}'")
                        elif is_market:
                            market_total = row
                            print(f"  找到外购电力（基于市场）汇总: Row {row_idx}, fourth_col='{fourth_col}'")

                # 构建两行数据
                if location_total:
                    total = self._safe_float(location_total[2]) if len(location_total) > 2 else 0
                    co2 = self._safe_float(location_total[3]) if len(location_total) > 3 else total

                    total_formatted = f"{total:,.2f}" if total > 0 else "0.00"
                    co2_formatted = f"{co2:,.2f}" if co2 > 0 else "0.00"

                    scope2_items.append({
                        'number': '2.1',
                        'emission_source': '外购电力（基于位置）',
                        'total_green_house_gas_emissions': total_formatted,
                        'CO2_emissions': co2_formatted,
                        'CH4_emissions': '0.00',
                        'N2O_emissions': '0.00',
                        'HFCs_emissions': '0.00',
                        'PFCs_emissions': '0.00',
                        'SFs_emissions': '0.00',
                        'NF3_emissions': '0.00'
                    })
                    print(f"  提取2.1 外购电力（基于位置）: {total_formatted} tCO2e")

                if market_total:
                    total = self._safe_float(market_total[2]) if len(market_total) > 2 else 0
                    co2 = self._safe_float(market_total[3]) if len(market_total) > 3 else total

                    total_formatted = f"{total:,.2f}" if total > 0 else "0.00"
                    co2_formatted = f"{co2:,.2f}" if co2 > 0 else "0.00"

                    scope2_items.append({
                        'number': '2.2',
                        'emission_source': '外购电力（基于市场）',
                        'total_green_house_gas_emissions': total_formatted,
                        'CO2_emissions': co2_formatted,
                        'CH4_emissions': '0.00',
                        'N2O_emissions': '0.00',
                        'HFCs_emissions': '0.00',
                        'PFCs_emissions': '0.00',
                        'SFs_emissions': '0.00',
                        'NF3_emissions': '0.00'
                    })
                    print(f"  提取2.2 外购电力（基于市场）: {total_formatted} tCO2e")

                data['scope2_items'] = scope2_items
                print(f"提取到范围二输入能源排放明细: {len(scope2_items)} 行")
            else:
                print("警告：未找到温室气体盘查表")
                data['scope2_items'] = []

        except Exception as e:
            print(f"提取scope2_items时出错: {e}")
            import traceback
            traceback.print_exc()
            data['scope2_items'] = []

        # ========== 从温室气体盘查表提取scope3_items（范围三各类别排放明细）==========
        # 提取详细排放源数据，而不是仅提取类别汇总
        try:
            # 查找温室气体盘查表（确保可以找到）
            pandata_sheet = None
            for sheet in wb.worksheets:
                if '盘查表' in str(sheet.title):
                    pandata_sheet = sheet
                    break

            if pandata_sheet:
                print(f"从温室气体盘查表提取范围三详细数据")
                scope3_items = []

                # 范围三类别名称映射
                category_names = {
                    '类别1': '外购商品和服务上游排放',
                    '类别2': '资本货物上游排放',
                    '类别3': '燃料和能源相关活动未包含在范围一和范围二中的上游排放',
                    '类别4': '上游运输和配送',
                    '类别5': '运营中产生的废弃物',
                    '类别6': '员工商务旅行',
                    '类别7': '员工通勤',
                    '类别8': '上游租赁资产',
                    '类别9': '下游运输和配送',
                    '类别10': '销售产品的加工',
                    '类别11': '销售产品的使用',
                    '类别12': '售出产品的加工',
                    '类别13': '下游租赁资产',
                    '类别14': '特许经营',
                    '类别15': '投资'
                }

                # 第一步：收集所有详细排放源行（按类别分组）
                category_detail_rows = {}  # {category_num: [row_data, ...]}

                for row_idx, row in enumerate(pandata_sheet.iter_rows(min_row=1, values_only=True), start=1):
                    row_vals = [str(v) if v is not None else '' for v in row[:12]]

                    # 检查是否是详细排放源行
                    # 特征：第1列是数字（行ID），第4列包含"范围三 类别X"
                    if row_vals[0] and row_vals[0].isdigit():
                        if len(row_vals) > 4 and '范围三' in row_vals[4] and '类别' in row_vals[4]:
                            # 提取类别编号
                            category_match = re.search(r'类别\s*(\d+)', row_vals[4])
                            if category_match:
                                category_num = int(category_match.group(1))
                                if category_num not in category_detail_rows:
                                    category_detail_rows[category_num] = []
                                category_detail_rows[category_num].append({
                                    'row_idx': row_idx,
                                    'data': row
                                })

                # 第二步：为每个类别的每个详细排放源创建数据项，按类别分组存储
                # 模板期望的数据格式：scope3_category1, scope3_category2, ..., scope3_category15
                total_scope3_items = 0
                for category_num in sorted(category_detail_rows.keys()):
                    detail_rows = category_detail_rows[category_num]
                    category_key = f'类别{category_num}'
                    category_var_name = f'scope3_category{category_num}'

                    category_items = []
                    sub_num = 0
                    for row_info in detail_rows:
                        row = row_info['data']
                        row_vals = [str(v) if v is not None else '' for v in row[:12]]

                        # 提取数据
                        # 列0: 行ID, 列1: 类别/组名, 列2: 排放源名称
                        emission_source_name = row_vals[2] if len(row_vals) > 2 and row_vals[2] else category_key

                        # 列5: 活动数据数量, 列6: 活动数据单位, 列7: 排放因子, 列8: 排放因子单位
                        activity_data = self._safe_float(row[5]) if len(row) > 5 and row[5] else 0
                        activity_unit = row_vals[6] if len(row_vals) > 6 else ''
                        emission_factor = self._safe_float(row[7]) if len(row) > 7 and row[7] else 0
                        factor_unit = row_vals[8] if len(row_vals) > 8 else ''

                        # 计算排放量（需要根据单位进行转换）
                        calculated_emission = 0
                        if activity_data > 0 and emission_factor > 0:
                            # 检查是否需要单位转换
                            # 规则：
                            # - kgCO2/kg, kgCO2/t, tCO2/t 等比值单位：不需要转换（1 kg/kg = 1 t/t）
                            # - kgCO2/unit, kg CO2/unit 等绝对值单位：需要除以1000转换为 tCO2/unit
                            factor_in_tons = emission_factor

                            # 只有当单位是 "每单位质量的kgCO2" 且不是比值时才转换
                            # kgCO2/kg, kgCO2/t 是比值，等于 tCO2/t，不需要转换
                            # kgCO2/(其他单位) 需要转换为 tCO2/(其他单位)
                            if factor_unit:
                                # 检查是否是比值单位（分母是质量单位）
                                is_ratio_unit = any(u in factor_unit for u in ['kgCO2/kg', 'kgCO2/t', 'tCO2/kg', 'tCO2/t'])
                                # 检查是否需要转换（kgCO2或kg CO2但不是比值）
                                needs_conversion = ('kgCO2' in factor_unit or 'kg CO2' in factor_unit) and not is_ratio_unit

                                if needs_conversion:
                                    factor_in_tons = emission_factor / 1000

                            calculated_emission = activity_data * factor_in_tons

                        # 只添加有排放量的项目（排放量 > 0.01）
                        if calculated_emission > 0.01:
                            sub_num += 1
                            total_formatted = f"{calculated_emission:,.2f}"
                            co2_formatted = f"{calculated_emission:,.2f}"

                            category_items.append({
                                'number': f'3.{category_num}.{sub_num}',
                                'emission_source': emission_source_name,
                                'total_green_house_gas_emissions': total_formatted,
                                'CO2_emissions': co2_formatted,
                                'CH4_emissions': '0.00',
                                'N2O_emissions': '0.00',
                                'HFCs_emissions': '0.00',
                                'PFCs_emissions': '0.00',
                                'SFs_emissions': '0.00',
                                'NF3_emissions': '0.00'
                            })

                    # 将类别数据存储到对应的变量中
                    data[category_var_name] = category_items
                    total_scope3_items += len(category_items)
                    if category_items:
                        print(f"  提取{category_var_name}: {len(category_items)} 行")

                # 计算每个类别的总排放量（从详细数据汇总）
                # 如果从"表格"sheet没有提取到汇总数据，则从详细数据计算
                print("  验证类别总排放量...")
                calculated_from_detail = []
                for cat_num in range(1, 16):
                    cat_var = f'scope3_category{cat_num}'
                    emission_key = f'scope_3_category_{cat_num}_emissions'

                    # 如果该类别有详细数据
                    if cat_var in data and data[cat_var]:
                        # 汇总该类别所有详细项的排放量
                        category_total = 0
                        for item in data[cat_var]:
                            # 从total_green_house_gas_emissions字段提取数值
                            emission_str = item.get('total_green_house_gas_emissions', '0')
                            # 去除逗号和空格
                            emission_str = emission_str.replace(',', '').replace(' ', '')
                            try:
                                emission_value = float(emission_str)
                                category_total += emission_value
                            except (ValueError, TypeError):
                                pass

                        # 如果"表格"sheet没有提取到数据，使用计算值
                        if data.get(emission_key, 0) == 0 and category_total > 0:
                            data[emission_key] = category_total
                            calculated_from_detail.append(cat_num)
                            print(f"    [+] 类别{cat_num}: {category_total:,.2f} tCO2e (从详细数据计算)")
                        elif category_total > 0:
                            # 验证：表格值 vs 详细数据计算值
                            table_value = data.get(emission_key, 0)
                            diff = abs(table_value - category_total)
                            if diff > 0.01:  # 差异超过0.01时警告
                                print(f"    [!] 类别{cat_num}: 表格值={table_value:,.2f}, 计算值={category_total:,.2f}, 差异={diff:,.2f}")
                            else:
                                print(f"    [OK] 类别{cat_num}: {category_total:,.2f} tCO2e")

                if calculated_from_detail:
                    print(f"  注意: 类别{calculated_from_detail}从详细数据计算得到")

                # 同时保留scope3_items用于兼容
                data['scope3_items'] = []
                for cat_num in range(1, 16):
                    cat_var = f'scope3_category{cat_num}'
                    if cat_var in data:
                        data['scope3_items'].extend(data[cat_var])

                print(f"提取到范围三详细排放明细总计: {total_scope3_items} 行")
            else:
                print("警告：未找到温室气体盘查表，无法提取范围三数据")
                data['scope3_items'] = []

        except Exception as e:
            print(f"提取scope3_items时出错: {e}")
            import traceback
            traceback.print_exc()
            data['scope3_items'] = []

        print(f"纯 xlsx 动态提取完成，范围三类别数据: {sum(1 for i in range(1, 16) if data.get(f'scope_3_category_{i}_emissions', 0) > 0)} 个类别有数据")

        # Add quantification methods data
        # Use ReportConfig for quantification methods
        report_config = ReportConfig(
            self.company_name or '某公司',
            self.reporting_period or '2024年'
        )
        data['quantification_methods'] = report_config.get_quantification_methods()

        # Use ReportConfig for scope 3 category names
        report_config_names = ReportConfig()
        data['scope_3_category_names'] = report_config_names.get_all_scope_3_category_names()

        # ========== 提取活动数据汇总表（基于位置） ==========
        print("[活动数据汇总表] 开始提取活动数据汇总表（基于位置）...")
        activity_summary_sheet = self._find_activity_summary_sheet()
        if activity_summary_sheet:
            data['act_summary_loc'] = self._extract_activity_summary_data(activity_summary_sheet)
            print(f"[活动数据汇总表] 成功提取 {len(data['act_summary_loc'])} 行数据")
        else:
            data['act_summary_loc'] = []
            print("[活动数据汇总表] 未找到活动数据汇总表，使用空列表")

        # ========== 提取活动数据汇总表（基于市场） ==========
        print("[活动数据汇总表] 开始提取活动数据汇总表（基于市场）...")
        activity_summary_sheet_mkt = self._find_activity_summary_sheet_market_based()
        if activity_summary_sheet_mkt:
            data['act_summary_mar'] = self._extract_activity_summary_data_market_based(activity_summary_sheet_mkt)
            print(f"[活动数据汇总表] 成功提取 {len(data['act_summary_mar'])} 行数据")
        else:
            data['act_summary_mar'] = []
            print("[活动数据汇总表] 未找到活动数据汇总表（基于市场），使用空列表")

        # Update flags based on data
        data = self._update_flags(data)

        # ========== 新增：基于特征指纹读取协议表格 ==========
        print("\n[协议读取] 开始基于特征指纹读取协议表格...")
        protocol_data = self.read_protocols()
        data.update(protocol_data)
        print(f"[协议读取] 成功读取 {len(protocol_data)} 个协议变量")
        for var_name, items in protocol_data.items():
            print(f"  - {var_name}: {len(items)} 行")
        # ========== 协议表格读取结束 ==========

        return data


if __name__ == "__main__": 
    # 测试1: Excel数据读取
    reader = ExcelDataReader('test_data.xlsx') 
    data = reader.extract_data() 
    print("--- 测试 data_reader.py ---\n", data)
    
    # 测试2: find_value_by_label方法
    if reader.workbook:
        test_value = reader.find_value_by_label('温室气体盘查清册', '组织名称：')
        print(f"\n--- 测试 find_value_by_label 方法 ---\n组织名称: {test_value}")
    
    # 测试3: 读取CSV文件（如果存在）
    csv_file = '减排行动统计.csv'
    if os.path.exists(csv_file):
        csv_reader = ExcelDataReader(csv_file)
        csv_data = csv_reader.read_to_list_of_dicts()
        print(f"\n--- 测试 CSV 读取功能 ---\n读取到 {len(csv_data)} 行数据")
        if csv_data:
            print("数据示例:", csv_data[:2])  # 打印前两行数据
    
    # 测试4: 读取Excel工作表为列表字典
    if reader.workbook:
        excel_list_data = reader.read_to_list_of_dicts('表1温室气体盘查表')
        print(f"\n--- 测试 Excel 列表字典读取功能 ---\n读取到 {len(excel_list_data)} 行数据")
        if excel_list_data:
            print("数据示例:", excel_list_data[:2])  # 打印前两行数据